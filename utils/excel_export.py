"""
Excel Export Module
===================

Funkcje do generowania plików Excel z danymi zamówień.
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ============================================
# Reusable Styles
# ============================================

_PURPLE = "7B2CBF"
_PURPLE_LIGHT = "F3E8FF"
_GREEN = "D4EDDA"
_GREEN_DARK = "28A745"
_RED = "F8D7DA"
_RED_DARK = "DC3545"
_ORANGE = "FFE5CC"
_VIOLET = "E8DAEF"
_BONUS_GREEN = "D1FAE5"
_GRAY_BG = "F8F9FA"
_GRAY_BORDER = "D0D0D0"
_DARK_TEXT = "212121"


def _styles():
    """Return a dict of reusable openpyxl style objects."""
    thin = Side(style='thin', color=_GRAY_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        'border': border,
        # Title row
        'title_font': Font(bold=True, size=16, color=_PURPLE),
        'subtitle_font': Font(size=11, color="666666"),
        # Section header (e.g. "STATYSTYKI", "SETY")
        'section_font': Font(bold=True, size=12, color=_PURPLE),
        'section_fill': PatternFill(start_color=_PURPLE_LIGHT, end_color=_PURPLE_LIGHT, fill_type="solid"),
        # Table header
        'header_font': Font(bold=True, color="FFFFFF", size=10),
        'header_fill': PatternFill(start_color=_PURPLE, end_color=_PURPLE, fill_type="solid"),
        'header_align': Alignment(horizontal="center", vertical="center", wrap_text=True),
        # Data cells
        'left': Alignment(horizontal="left", vertical="center"),
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'wrap': Alignment(horizontal="left", vertical="center", wrap_text=True),
        # Status fills
        'fulfilled_fill': PatternFill(start_color=_GREEN, end_color=_GREEN, fill_type="solid"),
        'unfulfilled_fill': PatternFill(start_color=_RED, end_color=_RED, fill_type="solid"),
        'fullset_fill': PatternFill(start_color=_VIOLET, end_color=_VIOLET, fill_type="solid"),
        'custom_fill': PatternFill(start_color=_ORANGE, end_color=_ORANGE, fill_type="solid"),
        'bonus_fill': PatternFill(start_color=_BONUS_GREEN, end_color=_BONUS_GREEN, fill_type="solid"),
        'alt_fill': PatternFill(start_color=_GRAY_BG, end_color=_GRAY_BG, fill_type="solid"),
        # Fonts for stats
        'stat_value_font': Font(bold=True, size=14, color=_DARK_TEXT),
        'stat_label_font': Font(size=10, color="666666"),
        'bold': Font(bold=True),
        'bold_green': Font(bold=True, color=_GREEN_DARK),
        'bold_red': Font(bold=True, color=_RED_DARK),
    }


def _write_header_row(ws, row, headers, s):
    """Write a styled table header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = s['header_font']
        cell.fill = s['header_fill']
        cell.alignment = s['header_align']
        cell.border = s['border']
    ws.row_dimensions[row].height = 32


def _write_cell(ws, row, col, value, s, align='left', bold=False, fill=None, fmt=None):
    """Write a single styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = s.get(align, s['left'])
    cell.border = s['border']
    if bold:
        cell.font = s['bold']
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter_or_index: width}."""
    for key, width in widths.items():
        if isinstance(key, int):
            key = get_column_letter(key)
        ws.column_dimensions[key].width = width


# ============================================
# Main: Exclusive Closure Excel (multi-sheet)
# ============================================

def generate_offer_closure_excel(page, summary):
    """
    Generates a multi-sheet Excel file for a closed Offer page.

    Sheets:
        1. Przegląd   - Key metrics, top products, sets breakdown
        2. Produkty    - Aggregated product table
        3. Zamówienia  - Order details (one row per order item)

    Args:
        page: OfferPage model object
        summary: Dict from get_page_summary(include_financials=True)

    Returns:
        BytesIO: Buffer with .xlsx file
    """
    wb = Workbook()
    s = _styles()

    _build_overview_sheet(wb, page, summary, s)
    _build_products_sheet(wb, summary, s)
    _build_orders_sheet(wb, summary, s)

    if page.page_type == 'preorder':
        _build_preorder_summary_sheet(wb, page, s)
        _build_preorder_quantities_sheet(wb, page, s)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ------------------------------------------
# Sheet 1: Przegląd (Overview)
# ------------------------------------------

def _build_overview_sheet(wb, page, summary, s):
    ws = wb.active
    ws.title = "Przegląd"
    ws.sheet_properties.tabColor = _PURPLE

    # --- Page title ---
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = summary.get('page_name', page.name)
    cell.font = s['title_font']
    cell.alignment = s['left']

    # --- Meta info ---
    row = 3
    meta_items = [
        ("Zamknięto:", summary.get('closed_at').strftime('%d.%m.%Y %H:%M') if summary.get('closed_at') else '-'),
        ("Przez:", summary.get('closed_by', '-')),
        ("Okres:", _format_period(summary.get('starts_at'), summary.get('ends_at'))),
    ]
    for label, value in meta_items:
        ws.cell(row=row, column=1, value=label).font = s['bold']
        ws.cell(row=row, column=1).alignment = s['right']
        ws.cell(row=row, column=2, value=value).alignment = s['left']
        row += 1

    # --- Statistics section ---
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="STATYSTYKI")
    cell.font = s['section_font']
    cell.fill = s['section_fill']
    cell.alignment = s['left']
    row += 1

    stats = [
        ("Zamówienia", summary.get('total_orders', 0)),
        ("Klienci", summary.get('unique_customers', 0)),
        ("Produkty (szt.)", summary.get('total_items', 0)),
    ]

    total_revenue = summary.get('total_revenue')
    if total_revenue is not None:
        stats.append(("Przychód (PLN)", f"{total_revenue:,.2f}"))
        avg = summary.get('avg_order_value')
        if avg is not None:
            stats.append(("Śr. wartość zamówienia", f"{avg:,.2f} PLN"))

    fulfillment = summary.get('fulfillment_pct')
    if fulfillment is not None:
        stats.append(("Realizacja setów", f"{fulfillment:.1f}%"))

    # Write stats in 2 columns (label + value pairs, 3 per row)
    col = 1
    for label, value in stats:
        ws.cell(row=row, column=col, value=label).font = s['stat_label_font']
        ws.cell(row=row, column=col, value=label).alignment = s['right']
        val_cell = ws.cell(row=row, column=col + 1, value=value)
        val_cell.font = s['stat_value_font']
        val_cell.alignment = s['left']
        col += 3
        if col > 6:
            col = 1
            row += 1
    row += 1

    # --- Top 5 Products ---
    top_products = summary.get('top_products', [])
    if top_products:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="TOP 5 PRODUKTÓW")
        cell.font = s['section_font']
        cell.fill = s['section_fill']
        cell.alignment = s['left']
        row += 1

        top_headers = ["#", "Produkt", "Ilość", "Zamówień", "Realizacja"]
        if total_revenue is not None:
            top_headers.append("Przychód (PLN)")
        _write_header_row(ws, row, top_headers, s)
        row += 1

        for idx, prod in enumerate(top_products, 1):
            name = prod['product_name']
            if prod.get('is_full_set'):
                name = f"SET: {name}"
            elif prod.get('is_custom'):
                name = f"RĘCZNY: {name}"
            if prod.get('is_bonus'):
                name = f"GRATIS: {name}"

            _write_cell(ws, row, 1, idx, s, align='center')
            _write_cell(ws, row, 2, name, s, align='left')
            _write_cell(ws, row, 3, prod['total_quantity'], s, align='center', bold=True)
            _write_cell(ws, row, 4, prod.get('order_count', 0), s, align='center')

            fp = prod.get('fulfillment_pct')
            if fp is not None:
                _write_cell(ws, row, 5, f"{fp:.0f}%", s, align='center')
            else:
                _write_cell(ws, row, 5, "-", s, align='center')

            if total_revenue is not None:
                rev = prod.get('revenue', 0) or 0
                _write_cell(ws, row, 6, rev, s, align='right', fmt='#,##0.00')

            row += 1

    # --- Sets breakdown ---
    sets_info = summary.get('sets', [])
    if sets_info:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="SETY")
        cell.font = s['section_font']
        cell.fill = s['section_fill']
        cell.alignment = s['left']
        row += 1

        for set_data in sets_info:
            # Set name + fulfillment
            set_name = set_data.get('set_name', 'Set')
            fp = set_data.get('fulfillment_pct')
            fp_str = f" ({fp:.0f}% zrealizowano)" if fp is not None else ""
            total_ord = set_data.get('total_ordered', 0)
            total_ful = set_data.get('total_fulfilled', 0)

            ws.merge_cells(f'A{row}:F{row}')
            cell = ws.cell(row=row, column=1, value=f"{set_name}{fp_str}  |  Zamówiono: {total_ord}  |  Zrealizowano: {total_ful}")
            cell.font = Font(bold=True, size=11)
            row += 1

            # Products in set
            set_products = set_data.get('products', [])
            if set_products:
                set_headers = ["Produkt", "Na set", "Zamówiono", "Zrealizowano", "Brakuje"]
                _write_header_row(ws, row, set_headers, s)
                row += 1

                for sp in set_products:
                    _write_cell(ws, row, 1, sp.get('product_name', ''), s, align='left')
                    _write_cell(ws, row, 2, sp.get('quantity_per_set', 0), s, align='center')
                    _write_cell(ws, row, 3, sp.get('total_ordered', 0), s, align='center')
                    _write_cell(ws, row, 4, sp.get('fulfilled', 0), s, align='center')

                    unfulfilled = sp.get('unfulfilled', 0)
                    c = _write_cell(ws, row, 5, unfulfilled, s, align='center')
                    if unfulfilled > 0:
                        c.font = s['bold_red']
                    else:
                        c.font = s['bold_green']

                    row += 1

            row += 1

    # Column widths
    _set_col_widths(ws, {'A': 22, 'B': 30, 'C': 16, 'D': 16, 'E': 16, 'F': 18})
    ws.freeze_panes = 'A3'


# ------------------------------------------
# Sheet 2: Produkty (Products aggregated)
# ------------------------------------------

def _build_products_sheet(wb, summary, s):
    ws = wb.create_sheet("Produkty")
    ws.sheet_properties.tabColor = "28A745"

    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "Podsumowanie produktów"
    cell.font = s['title_font']
    cell.alignment = s['left']

    # Headers
    has_financials = summary.get('total_revenue') is not None
    headers = ["#", "Produkt", "Typ", "Zamówień", "Ilość łączna", "Zrealizowano", "Niezrealizowano", "Realizacja %"]
    if has_financials:
        headers.append("Przychód (PLN)")

    header_row = 3
    _write_header_row(ws, header_row, headers, s)

    # Data
    products = summary.get('products_aggregated', [])
    for idx, prod in enumerate(products, 1):
        row = header_row + idx

        # Alternating background
        fill = s['alt_fill'] if idx % 2 == 0 else None

        # Type label
        if prod.get('is_bonus'):
            type_label = "Gratis"
        elif prod.get('is_full_set'):
            type_label = "Set"
        elif prod.get('is_custom'):
            type_label = "Ręczny"
        else:
            type_label = "Zwykły"

        _write_cell(ws, row, 1, idx, s, align='center', fill=fill)
        _write_cell(ws, row, 2, prod['product_name'], s, align='left', fill=fill)

        type_cell = _write_cell(ws, row, 3, type_label, s, align='center', fill=fill)
        if type_label == "Set":
            type_cell.fill = s['fullset_fill']
        elif type_label == "Ręczny":
            type_cell.fill = s['custom_fill']
        elif type_label == "Gratis":
            type_cell.fill = s['bonus_fill']

        _write_cell(ws, row, 4, prod.get('order_count', 0), s, align='center', fill=fill)
        _write_cell(ws, row, 5, prod['total_quantity'], s, align='center', bold=True, fill=fill)

        ful = prod.get('fulfilled_quantity', 0)
        ful_cell = _write_cell(ws, row, 6, ful, s, align='center', fill=fill)
        if ful > 0:
            ful_cell.font = s['bold_green']

        unful = prod.get('unfulfilled_quantity', 0)
        unful_cell = _write_cell(ws, row, 7, unful, s, align='center', fill=fill)
        if unful > 0:
            unful_cell.font = s['bold_red']

        fp = prod.get('fulfillment_pct')
        if fp is not None:
            fp_cell = _write_cell(ws, row, 8, fp / 100, s, align='center', fill=fill, fmt='0%')
            if fp >= 100:
                fp_cell.fill = s['fulfilled_fill']
            elif fp < 50:
                fp_cell.fill = s['unfulfilled_fill']
        else:
            _write_cell(ws, row, 8, "-", s, align='center', fill=fill)

        if has_financials:
            rev = prod.get('revenue', 0) or 0
            _write_cell(ws, row, 9, rev, s, align='right', fill=fill, fmt='#,##0.00')

    # Summary row
    if products:
        sum_row = header_row + len(products) + 2
        ws.cell(row=sum_row, column=1, value="SUMA").font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=5, value=sum(p['total_quantity'] for p in products)).font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=5).alignment = s['center']
        ws.cell(row=sum_row, column=6, value=sum(p.get('fulfilled_quantity', 0) for p in products)).font = Font(bold=True, size=11, color=_GREEN_DARK)
        ws.cell(row=sum_row, column=6).alignment = s['center']
        ws.cell(row=sum_row, column=7, value=sum(p.get('unfulfilled_quantity', 0) for p in products)).font = Font(bold=True, size=11, color=_RED_DARK)
        ws.cell(row=sum_row, column=7).alignment = s['center']

        if has_financials:
            total_rev = sum(p.get('revenue', 0) or 0 for p in products)
            c = ws.cell(row=sum_row, column=9, value=total_rev)
            c.font = Font(bold=True, size=11)
            c.number_format = '#,##0.00'
            c.alignment = s['right']

        # Border on summary row
        for col in range(1, len(headers) + 1):
            ws.cell(row=sum_row, column=col).border = Border(
                top=Side(style='medium', color=_PURPLE),
                bottom=Side(style='medium', color=_PURPLE),
            )

    # Column widths
    _set_col_widths(ws, {
        'A': 6, 'B': 35, 'C': 12, 'D': 12, 'E': 14,
        'F': 16, 'G': 18, 'H': 14, 'I': 18,
    })
    ws.freeze_panes = 'A4'


# ------------------------------------------
# Sheet 3: Zamówienia (Order details)
# ------------------------------------------

def _build_orders_sheet(wb, summary, s):
    ws = wb.create_sheet("Zamówienia")
    ws.sheet_properties.tabColor = "4A90D9"

    # Title
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Szczegóły zamówień"
    cell.font = s['title_font']
    cell.alignment = s['left']

    has_financials = summary.get('total_revenue') is not None

    # Headers
    headers = [
        "Nr zamówienia", "Klient", "Email", "Telefon", "Data",
        "Produkt", "Typ", "Ilość", "Status realizacji",
    ]
    if has_financials:
        headers.extend(["Cena jedn. (PLN)", "Wartość (PLN)"])

    header_row = 3
    _write_header_row(ws, header_row, headers, s)

    # Data - one row per order item
    orders = summary.get('orders', [])
    row = header_row + 1
    prev_order_id = None

    for order in orders:
        order_id = order.get('order_id')
        items = order.get('order_items', [])
        if not items:
            items = [None]  # Still show order row with no items

        is_new_order = order_id != prev_order_id
        prev_order_id = order_id

        for item_idx, item in enumerate(items):
            # Only write order info on first item of each order
            if item_idx == 0:
                # Separator line above each new order (except first)
                if is_new_order and row > header_row + 1:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row - 1, column=col).border = Border(
                            bottom=Side(style='thin', color=_PURPLE),
                            left=ws.cell(row=row - 1, column=col).border.left,
                            right=ws.cell(row=row - 1, column=col).border.right,
                            top=ws.cell(row=row - 1, column=col).border.top,
                        )

                _write_cell(ws, row, 1, order.get('order_number', ''), s, align='left', bold=True)

                customer_name = order.get('customer_name') or 'Gość'
                _write_cell(ws, row, 2, customer_name, s, align='left')
                _write_cell(ws, row, 3, order.get('customer_email') or '-', s, align='left')
                _write_cell(ws, row, 4, order.get('customer_phone') or '-', s, align='left')

                created_at = order.get('created_at')
                if created_at and hasattr(created_at, 'strftime'):
                    date_str = created_at.strftime('%d.%m.%Y %H:%M')
                else:
                    date_str = str(created_at) if created_at else '-'
                _write_cell(ws, row, 5, date_str, s, align='center')
            else:
                # Empty cells for repeated order columns
                for col in range(1, 6):
                    _write_cell(ws, row, col, '', s, align='left')

            # Item columns
            if item:
                product_name = item.get('product_name', '-')
                _write_cell(ws, row, 6, product_name, s, align='left')

                # Type
                if item.get('is_bonus'):
                    type_label = "Gratis"
                    type_fill = s['bonus_fill']
                elif item.get('is_full_set'):
                    type_label = "Set"
                    type_fill = s['fullset_fill']
                elif item.get('is_custom'):
                    type_label = "Ręczny"
                    type_fill = s['custom_fill']
                else:
                    type_label = "Zwykły"
                    type_fill = None
                type_cell = _write_cell(ws, row, 7, type_label, s, align='center')
                if type_fill:
                    type_cell.fill = type_fill

                _write_cell(ws, row, 8, item.get('quantity', 0), s, align='center', bold=True)

                # Fulfillment status
                fulfilled = item.get('is_set_fulfilled')
                if fulfilled is None:
                    status_text = "Tak"
                    status_fill = s['fulfilled_fill']
                elif fulfilled:
                    status_text = "Tak"
                    status_fill = s['fulfilled_fill']
                else:
                    status_text = "Nie"
                    status_fill = s['unfulfilled_fill']

                if item.get('is_full_set'):
                    status_text = "Set"
                    status_fill = s['fullset_fill']
                elif item.get('is_custom'):
                    status_text = "Ręczny"
                    status_fill = s['custom_fill']

                status_cell = _write_cell(ws, row, 9, status_text, s, align='center')
                status_cell.fill = status_fill

                if has_financials:
                    price = item.get('price', 0) or 0
                    total = item.get('total', 0) or 0
                    _write_cell(ws, row, 10, price, s, align='right', fmt='#,##0.00')
                    _write_cell(ws, row, 11, total, s, align='right', fmt='#,##0.00')
            else:
                # No items - empty product columns
                for col in range(6, len(headers) + 1):
                    _write_cell(ws, row, col, '-', s, align='center')

            row += 1

    # Order totals summary
    if orders:
        row += 1
        ws.cell(row=row, column=1, value="PODSUMOWANIE").font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Zamówień:").font = s['bold']
        ws.cell(row=row, column=1).alignment = s['right']
        ws.cell(row=row, column=2, value=summary.get('total_orders', len(orders))).font = s['stat_value_font']

        if has_financials:
            total_rev = summary.get('total_revenue', 0)
            ws.cell(row=row, column=4, value="Przychód:").font = s['bold']
            ws.cell(row=row, column=4).alignment = s['right']
            c = ws.cell(row=row, column=5, value=total_rev)
            c.font = s['stat_value_font']
            c.number_format = '#,##0.00'

    # Column widths
    widths = {
        'A': 18, 'B': 25, 'C': 28, 'D': 16, 'E': 18,
        'F': 30, 'G': 12, 'H': 10, 'I': 18,
    }
    if has_financials:
        widths['J'] = 16
        widths['K'] = 16
    _set_col_widths(ws, widths)
    ws.freeze_panes = 'A4'


# ============================================
# Live Dashboard Excel (multi-sheet)
# ============================================

def generate_offer_live_excel(page, summary):
    """
    Generates a multi-sheet Excel file for LIVE Offer dashboard.

    Sheets:
        1. Przegląd   - Key metrics, top products, sets matrix with customer names
        2. Produkty    - Aggregated product table
        3. Zamówienia  - Order details (one row per order item)

    Args:
        page: OfferPage model object
        summary: Dict from get_live_summary(include_financials=True)

    Returns:
        BytesIO: Buffer with .xlsx file
    """
    wb = Workbook()
    s = _styles()

    _build_live_overview_sheet(wb, page, summary, s)
    _build_live_products_sheet(wb, summary, s)
    _build_live_orders_sheet(wb, summary, s)

    if page.page_type == 'preorder':
        _build_preorder_summary_sheet(wb, page, s)
        _build_preorder_quantities_sheet(wb, page, s)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ------------------------------------------
# Live Sheet 1: Przegląd (Overview + Sets Matrix)
# ------------------------------------------

_MATRIX_FILLED = "D4EDDA"       # green bg for filled slot
_MATRIX_EMPTY = "F8F9FA"        # light gray for empty slot
_MATRIX_HEADER = "E8DAEF"       # light violet for set column headers


def _build_live_overview_sheet(wb, page, summary, s):
    ws = wb.active
    ws.title = "Przegląd"
    ws.sheet_properties.tabColor = _PURPLE

    filled_fill = PatternFill(start_color=_MATRIX_FILLED, end_color=_MATRIX_FILLED, fill_type="solid")
    empty_fill = PatternFill(start_color=_MATRIX_EMPTY, end_color=_MATRIX_EMPTY, fill_type="solid")
    matrix_header_fill = PatternFill(start_color=_MATRIX_HEADER, end_color=_MATRIX_HEADER, fill_type="solid")

    # --- Page title ---
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"{summary.get('page_name', page.name)} — LIVE"
    cell.font = s['title_font']
    cell.alignment = s['left']

    # --- Meta info ---
    row = 3
    meta_items = [
        ("Status:", summary.get('status', '-')),
        ("Okres:", _format_period(summary.get('starts_at'), summary.get('ends_at'))),
        ("Eksport:", datetime.now().strftime('%d.%m.%Y %H:%M')),
    ]
    for label, value in meta_items:
        ws.cell(row=row, column=1, value=label).font = s['bold']
        ws.cell(row=row, column=1).alignment = s['right']
        ws.cell(row=row, column=2, value=value).alignment = s['left']
        row += 1

    # --- Statistics section ---
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="STATYSTYKI")
    cell.font = s['section_font']
    cell.fill = s['section_fill']
    cell.alignment = s['left']
    row += 1

    stats = [
        ("Zamówienia", summary.get('total_orders', 0)),
        ("Klienci", summary.get('unique_customers', 0)),
        ("Produkty (szt.)", summary.get('total_items', 0)),
    ]

    total_revenue = summary.get('total_revenue')
    if total_revenue is not None:
        stats.append(("Przychód (PLN)", f"{total_revenue:,.2f}"))
        avg = summary.get('avg_order_value')
        if avg is not None:
            stats.append(("Śr. wartość zamówienia", f"{avg:,.2f} PLN"))

    bonus = summary.get('total_bonus_items', 0)
    if bonus > 0:
        stats.append(("Produkty gratis (szt.)", bonus))

    reservations = summary.get('active_reservations', 0)
    if reservations > 0:
        stats.append(("Aktywne rezerwacje", reservations))

    col = 1
    for label, value in stats:
        ws.cell(row=row, column=col, value=label).font = s['stat_label_font']
        ws.cell(row=row, column=col).alignment = s['right']
        val_cell = ws.cell(row=row, column=col + 1, value=value)
        val_cell.font = s['stat_value_font']
        val_cell.alignment = s['left']
        col += 3
        if col > 6:
            col = 1
            row += 1
    row += 1

    # --- Top 5 Products ---
    top_products = summary.get('top_products', [])
    if top_products:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="TOP 5 PRODUKTÓW")
        cell.font = s['section_font']
        cell.fill = s['section_fill']
        cell.alignment = s['left']
        row += 1

        top_headers = ["#", "Produkt", "Ilość", "Zamówień"]
        if total_revenue is not None:
            top_headers.append("Przychód (PLN)")
        _write_header_row(ws, row, top_headers, s)
        row += 1

        for idx, prod in enumerate(top_products, 1):
            name = prod['product_name']
            if prod.get('is_full_set'):
                name = f"SET: {name}"
            elif prod.get('is_custom'):
                name = f"RĘCZNY: {name}"
            if prod.get('is_bonus'):
                name = f"GRATIS: {name}"

            _write_cell(ws, row, 1, idx, s, align='center')
            _write_cell(ws, row, 2, name, s, align='left')
            _write_cell(ws, row, 3, prod['total_quantity'], s, align='center', bold=True)
            _write_cell(ws, row, 4, prod.get('order_count', 0), s, align='center')

            if total_revenue is not None:
                rev = prod.get('revenue', 0) or 0
                _write_cell(ws, row, 5, rev, s, align='right', fmt='#,##0.00')

            row += 1

    # --- Sets matrix with customer names ---
    sets_info = summary.get('sets', [])
    if sets_info:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="MACIERZ SETÓW")
        cell.font = s['section_font']
        cell.fill = s['section_fill']
        cell.alignment = s['left']
        row += 1

        for set_data in sets_info:
            set_name = set_data.get('set_name', 'Set')
            max_sets = set_data.get('set_max_sets', 0)
            has_limit = set_data.get('has_limit', True)
            ordered_sets = set_data.get('ordered_sets', 0)
            total_sets_sold = set_data.get('total_sets_sold', 0)
            full_set_sold = set_data.get('full_set_sold', 0)

            # Set header
            if has_limit:
                set_header = f"{set_name}  —  {ordered_sets} / {max_sets} kompletnych setów"
            else:
                set_header = f"{set_name}  —  {ordered_sets} kompletnych setów (bez limitu)"

            total_cols = 1 + max_sets  # product name + set columns
            if has_limit:
                total_cols += 1  # locked column

            end_col_letter = get_column_letter(max(total_cols, 2))
            ws.merge_cells(f'A{row}:{end_col_letter}{row}')
            cell = ws.cell(row=row, column=1, value=set_header)
            cell.font = Font(bold=True, size=11)
            row += 1

            products = set_data.get('products', [])
            non_full_set = [p for p in products if not p.get('is_full_set')]
            full_set_products = [p for p in products if p.get('is_full_set')]

            if non_full_set and max_sets > 0:
                # Matrix header row: Produkt | Set 1 | Set 2 | ... | Set N
                cell = ws.cell(row=row, column=1, value="Produkt")
                cell.font = s['header_font']
                cell.fill = s['header_fill']
                cell.alignment = s['header_align']
                cell.border = s['border']

                for col_idx in range(max_sets):
                    cell = ws.cell(row=row, column=col_idx + 2, value=f"Set {col_idx + 1}")
                    cell.font = Font(bold=True, size=10)
                    cell.fill = matrix_header_fill
                    cell.alignment = s['center']
                    cell.border = s['border']

                ws.row_dimensions[row].height = 28
                row += 1

                # Product rows with customer names in slots
                for prod in non_full_set:
                    cell = ws.cell(row=row, column=1, value=prod.get('product_name', ''))
                    cell.font = s['bold']
                    cell.alignment = s['left']
                    cell.border = s['border']

                    slots = prod.get('slots', [])
                    for sl_idx in range(max_sets):
                        slot = slots[sl_idx] if sl_idx < len(slots) else None
                        is_filled = slot and (slot.get('filled') if isinstance(slot, dict) else slot)
                        customer = slot.get('customer', '') if isinstance(slot, dict) else ''

                        cell = ws.cell(row=row, column=sl_idx + 2)
                        cell.border = s['border']
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        if is_filled:
                            cell.value = customer or "✓"
                            cell.fill = filled_fill
                            if customer:
                                cell.font = Font(size=9)
                            else:
                                cell.font = Font(size=12, color=_GREEN_DARK)
                        else:
                            cell.value = ""
                            cell.fill = empty_fill

                    row += 1

            # Full set product line
            if full_set_products:
                fp = full_set_products[0]
                row += 1
                ws.merge_cells(f'A{row}:B{row}')
                cell = ws.cell(row=row, column=1, value=f"SET: {fp.get('product_name', '')}  —  {fp.get('total_ordered', 0)} szt. sprzedanych")
                cell.font = Font(bold=True, size=10, color=_PURPLE)
                cell.fill = s['fullset_fill']
                cell.border = s['border']
                row += 1

            # Total sets sold summary
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws.cell(row=row, column=1, value=f"Łącznie setów: {total_sets_sold} ({ordered_sets} kompletnych + {full_set_sold} pojedynczych)")
            cell.font = Font(bold=True, size=10)
            cell.border = s['border']
            row += 2

    # Column widths - dynamic based on sets
    _set_col_widths(ws, {'A': 30, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20})
    # Wider columns for set matrices
    max_set_cols = max((sd.get('set_max_sets', 0) for sd in sets_info), default=0) if sets_info else 0
    for i in range(2, max_set_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = 'A3'


# ------------------------------------------
# Live Sheet 2: Produkty
# ------------------------------------------

def _build_live_products_sheet(wb, summary, s):
    ws = wb.create_sheet("Produkty")
    ws.sheet_properties.tabColor = "28A745"

    # Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Sprzedaż produktów — LIVE"
    cell.font = s['title_font']
    cell.alignment = s['left']

    has_financials = summary.get('total_revenue') is not None
    headers = ["#", "Produkt", "Typ", "Zamówień", "Ilość łączna"]
    if has_financials:
        headers.append("Przychód (PLN)")

    header_row = 3
    _write_header_row(ws, header_row, headers, s)

    products = summary.get('products_aggregated', [])
    for idx, prod in enumerate(products, 1):
        row = header_row + idx
        fill = s['alt_fill'] if idx % 2 == 0 else None

        if prod.get('is_bonus'):
            type_label = "Gratis"
        elif prod.get('is_full_set'):
            type_label = "Set"
        elif prod.get('is_custom'):
            type_label = "Ręczny"
        else:
            type_label = "Zwykły"

        _write_cell(ws, row, 1, idx, s, align='center', fill=fill)
        _write_cell(ws, row, 2, prod['product_name'], s, align='left', fill=fill)

        type_cell = _write_cell(ws, row, 3, type_label, s, align='center', fill=fill)
        if type_label == "Set":
            type_cell.fill = s['fullset_fill']
        elif type_label == "Ręczny":
            type_cell.fill = s['custom_fill']
        elif type_label == "Gratis":
            type_cell.fill = s['bonus_fill']

        _write_cell(ws, row, 4, prod.get('order_count', 0), s, align='center', fill=fill)
        _write_cell(ws, row, 5, prod['total_quantity'], s, align='center', bold=True, fill=fill)

        if has_financials:
            rev = prod.get('revenue', 0) or 0
            _write_cell(ws, row, 6, rev, s, align='right', fill=fill, fmt='#,##0.00')

    # Summary row
    if products:
        sum_row = header_row + len(products) + 2
        ws.cell(row=sum_row, column=1, value="SUMA").font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=5, value=sum(p['total_quantity'] for p in products)).font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=5).alignment = s['center']

        if has_financials:
            total_rev = sum(p.get('revenue', 0) or 0 for p in products)
            c = ws.cell(row=sum_row, column=6, value=total_rev)
            c.font = Font(bold=True, size=11)
            c.number_format = '#,##0.00'
            c.alignment = s['right']

        for col in range(1, len(headers) + 1):
            ws.cell(row=sum_row, column=col).border = Border(
                top=Side(style='medium', color=_PURPLE),
                bottom=Side(style='medium', color=_PURPLE),
            )

    _set_col_widths(ws, {'A': 6, 'B': 35, 'C': 12, 'D': 12, 'E': 14, 'F': 18})
    ws.freeze_panes = 'A4'


# ------------------------------------------
# Live Sheet 3: Zamówienia
# ------------------------------------------

def _build_live_orders_sheet(wb, summary, s):
    ws = wb.create_sheet("Zamówienia")
    ws.sheet_properties.tabColor = "4A90D9"

    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Szczegóły zamówień — LIVE"
    cell.font = s['title_font']
    cell.alignment = s['left']

    has_financials = summary.get('total_revenue') is not None

    headers = [
        "Nr zamówienia", "Klient", "Email", "Telefon", "Data",
        "Produkt", "Typ", "Ilość",
    ]
    if has_financials:
        headers.extend(["Cena jedn. (PLN)", "Wartość (PLN)"])

    header_row = 3
    _write_header_row(ws, header_row, headers, s)

    orders = summary.get('orders', [])
    row = header_row + 1
    prev_order_id = None

    for order in orders:
        order_id = order.get('order_id')
        items = order.get('order_items', [])
        if not items:
            items = [None]

        is_new_order = order_id != prev_order_id
        prev_order_id = order_id

        for item_idx, item in enumerate(items):
            if item_idx == 0:
                if is_new_order and row > header_row + 1:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row - 1, column=col).border = Border(
                            bottom=Side(style='thin', color=_PURPLE),
                            left=ws.cell(row=row - 1, column=col).border.left,
                            right=ws.cell(row=row - 1, column=col).border.right,
                            top=ws.cell(row=row - 1, column=col).border.top,
                        )

                _write_cell(ws, row, 1, order.get('order_number', ''), s, align='left', bold=True)

                customer_name = order.get('customer_name') or 'Gość'
                _write_cell(ws, row, 2, customer_name, s, align='left')
                _write_cell(ws, row, 3, order.get('customer_email') or '-', s, align='left')
                _write_cell(ws, row, 4, order.get('customer_phone') or '-', s, align='left')

                created_at = order.get('created_at')
                if created_at and hasattr(created_at, 'strftime'):
                    date_str = created_at.strftime('%d.%m.%Y %H:%M')
                else:
                    date_str = str(created_at) if created_at else '-'
                _write_cell(ws, row, 5, date_str, s, align='center')
            else:
                for col in range(1, 6):
                    _write_cell(ws, row, col, '', s, align='left')

            if item:
                product_name = item.get('product_name', '-')
                _write_cell(ws, row, 6, product_name, s, align='left')

                if item.get('is_bonus'):
                    type_label = "Gratis"
                    type_fill = s['bonus_fill']
                elif item.get('is_full_set'):
                    type_label = "Set"
                    type_fill = s['fullset_fill']
                elif item.get('is_custom'):
                    type_label = "Ręczny"
                    type_fill = s['custom_fill']
                else:
                    type_label = "Zwykły"
                    type_fill = None
                type_cell = _write_cell(ws, row, 7, type_label, s, align='center')
                if type_fill:
                    type_cell.fill = type_fill

                _write_cell(ws, row, 8, item.get('quantity', 0), s, align='center', bold=True)

                if has_financials:
                    price = item.get('price', 0) or 0
                    total = item.get('total', 0) or 0
                    _write_cell(ws, row, 9, price, s, align='right', fmt='#,##0.00')
                    _write_cell(ws, row, 10, total, s, align='right', fmt='#,##0.00')
            else:
                for col in range(6, len(headers) + 1):
                    _write_cell(ws, row, col, '-', s, align='center')

            row += 1

    # Summary
    if orders:
        row += 1
        ws.cell(row=row, column=1, value="PODSUMOWANIE").font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Zamówień:").font = s['bold']
        ws.cell(row=row, column=1).alignment = s['right']
        ws.cell(row=row, column=2, value=summary.get('total_orders', len(orders))).font = s['stat_value_font']

        if has_financials:
            total_rev = summary.get('total_revenue', 0)
            ws.cell(row=row, column=4, value="Przychód:").font = s['bold']
            ws.cell(row=row, column=4).alignment = s['right']
            c = ws.cell(row=row, column=5, value=total_rev)
            c.font = s['stat_value_font']
            c.number_format = '#,##0.00'

    widths = {
        'A': 18, 'B': 25, 'C': 28, 'D': 16, 'E': 18,
        'F': 30, 'G': 12, 'H': 10,
    }
    if has_financials:
        widths['I'] = 16
        widths['J'] = 16
    _set_col_widths(ws, widths)
    ws.freeze_panes = 'A4'


# ============================================
# Pre-order matrix report (Podsumowanie + Ilości)
# ============================================

_PREORDER_PRICE_FILL = "FFF7E6"      # soft yellow for pln/krw rows
_PREORDER_SUMA_FILL = "E8DAEF"       # violet for SUMA row


def _get_krw_pln_rate():
    """Zwraca aktualny kurs KRW/PLN (ile KRW za 1 PLN). None gdy niedostępny."""
    try:
        from utils.currency import get_exchange_rate
        data = get_exchange_rate('KRW')
        rate_krw_per_pln = data.get('rate')
        if not rate_krw_per_pln or rate_krw_per_pln <= 0:
            return None
        return round(1 / rate_krw_per_pln, 2)
    except Exception:
        return None


def _last_word(name):
    """Zwraca ostatni wyraz z nazwy produktu (po spacji)."""
    if not name:
        return ''
    return name.strip().split()[-1]


def _section_column_name(section):
    """Zwraca nazwę sekcji do nagłówka kolumny."""
    if section.section_type == 'variant_group':
        return section.variant_group.name if section.variant_group else 'Grupa wariantowa'
    if section.section_type == 'set':
        return section.set_name or 'Komplet'
    if section.section_type == 'product':
        return section.product.name if section.product else 'Produkt'
    return 'Sekcja'


def _section_products(section):
    """Zwraca listę produktów wchodzących w skład sekcji (do matching order items)."""
    if section.section_type == 'variant_group':
        return list(section.variant_group.products) if section.variant_group else []
    if section.section_type == 'set':
        products = []
        for item in section.get_set_items_ordered():
            products.extend(item.get_products())
        if section.set_product_id and section.set_product:
            products.append(section.set_product)
        return products
    if section.section_type == 'product':
        return [section.product] if section.product else []
    return []


def _price_pln_for_section(products):
    """Cena PLN do wiersza 'pln'. Range 'min-max' gdy różne ceny w grupie."""
    prices = []
    for p in products:
        if p and p.sale_price is not None:
            prices.append(float(p.sale_price))
    if not prices:
        return None
    p_min, p_max = min(prices), max(prices)
    if p_min == p_max:
        return _fmt_price(p_min)
    return f"{_fmt_price(p_min)}-{_fmt_price(p_max)}"


def _price_krw_for_section(products):
    """Cena KRW do wiersza 'krw'. Tylko gdy produkt ma purchase_currency='KRW'."""
    prices = []
    for p in products:
        if p and p.purchase_currency == 'KRW' and p.purchase_price is not None:
            prices.append(float(p.purchase_price))
    if not prices:
        return None
    p_min, p_max = min(prices), max(prices)
    if p_min == p_max:
        return _fmt_price(p_min, decimals=0)
    return f"{_fmt_price(p_min, decimals=0)}-{_fmt_price(p_max, decimals=0)}"


def _fmt_price(value, decimals=2):
    """Formatuj cenę usuwając zbędne zera po przecinku."""
    if value is None:
        return None
    if decimals == 0 or float(value).is_integer():
        return int(round(value))
    return round(value, decimals)


def _preorder_collect_data(page):
    """
    Zbiera dane potrzebne do zakładek Podsumowanie i Ilości.

    Returns:
        dict z kluczami:
            - orderable_sections: lista OfferSection (tylko product/variant_group/set)
            - customers: listę krotek (klucz, nazwa, user_id, email) w kolejności 1-go zakupu
            - matrix: {(customer_key, section_id): [(product_obj, quantity), ...]}
    """
    from modules.orders.models import Order, OrderItem

    orderable_types = ('product', 'variant_group', 'set')
    sections = [s for s in page.sections.order_by('sort_order').all()
                if s.section_type in orderable_types]

    # build product_id -> section map (pierwsza sekcja zawierająca produkt wygrywa)
    product_to_section = {}
    for sec in sections:
        for prod in _section_products(sec):
            if prod and prod.id not in product_to_section:
                product_to_section[prod.id] = sec.id

    orders = Order.query.filter_by(offer_page_id=page.id).filter(
        Order.status != 'anulowane'
    ).order_by(Order.created_at.asc()).all()

    customers_order = []
    seen_customers = set()
    matrix = {}

    for order in orders:
        customer_key = f'u{order.user_id}' if order.user_id else f'e{order.customer_email}'
        if customer_key not in seen_customers:
            seen_customers.add(customer_key)
            customers_order.append({
                'key': customer_key,
                'name': order.customer_name or order.customer_email or 'Gość',
            })

        for item in order.items:
            if item.is_bonus:
                continue
            if not item.product_id:
                continue
            section_id = product_to_section.get(item.product_id)
            if section_id is None:
                continue
            cell_key = (customer_key, section_id)
            matrix.setdefault(cell_key, []).append((item.product, int(item.quantity)))

    return {
        'sections': sections,
        'customers': customers_order,
        'matrix': matrix,
    }


def _format_variant_cell(entries):
    """Formatuj komórkę dla variant_group: 'Hongjoong ×2, San, Seonghwa'."""
    from collections import OrderedDict
    totals = OrderedDict()
    for product, qty in entries:
        label = _last_word(product.name) if product else ''
        totals[label] = totals.get(label, 0) + qty
    parts = []
    for label, qty in totals.items():
        parts.append(f'{label} ×{qty}' if qty > 1 else label)
    return ', '.join(parts)


def _format_product_cell(entries):
    """Formatuj komórkę dla section_type=product: suma ilości."""
    return sum(qty for _, qty in entries) or None


def _build_preorder_summary_sheet(wb, page, s):
    """Zakładka 'Podsumowanie' — macierz klient × sekcja z cenami, sumami i kursem."""
    ws = wb.create_sheet(title='Podsumowanie', index=0)
    ws.sheet_properties.tabColor = _PURPLE

    data = _preorder_collect_data(page)
    sections = data['sections']
    customers = data['customers']
    matrix = data['matrix']

    price_fill = PatternFill(start_color=_PREORDER_PRICE_FILL, end_color=_PREORDER_PRICE_FILL, fill_type='solid')
    suma_fill = PatternFill(start_color=_PREORDER_SUMA_FILL, end_color=_PREORDER_SUMA_FILL, fill_type='solid')

    # Row 1: headers
    headers = ['Klient'] + [_section_column_name(sec) for sec in sections] + ['Suma (PLN)', 'Suma (KRW)']
    _write_header_row(ws, 1, headers, s)

    # Precompute section prices + product sets
    section_products = [_section_products(sec) for sec in sections]
    pln_prices = [_price_pln_for_section(prods) for prods in section_products]
    krw_prices = [_price_krw_for_section(prods) for prods in section_products]

    # Row 2: pln
    _write_cell(ws, 2, 1, 'pln', s, align='center', bold=True, fill=price_fill)
    for idx, val in enumerate(pln_prices, start=2):
        _write_cell(ws, 2, idx, val, s, align='center', fill=price_fill)
    _write_cell(ws, 2, len(sections) + 2, None, s, fill=price_fill)
    _write_cell(ws, 2, len(sections) + 3, None, s, fill=price_fill)

    # Row 3: krw
    _write_cell(ws, 3, 1, 'krw', s, align='center', bold=True, fill=price_fill)
    for idx, val in enumerate(krw_prices, start=2):
        _write_cell(ws, 3, idx, val, s, align='center', fill=price_fill)
    _write_cell(ws, 3, len(sections) + 2, None, s, fill=price_fill)
    _write_cell(ws, 3, len(sections) + 3, None, s, fill=price_fill)

    # Rows: customers
    customer_totals_pln = []
    customer_totals_krw = []
    col_totals_pln = [0.0] * len(sections)
    col_totals_krw = [0.0] * len(sections)

    for i, customer in enumerate(customers):
        row = 4 + i
        _write_cell(ws, row, 1, customer['name'], s, align='left')
        total_pln = 0.0
        total_krw = 0.0

        for j, sec in enumerate(sections):
            col = 2 + j
            entries = matrix.get((customer['key'], sec.id), [])
            if not entries:
                _write_cell(ws, row, col, None, s, align='center')
                continue

            if sec.section_type == 'variant_group':
                cell_value = _format_variant_cell(entries)
            else:
                cell_value = _format_product_cell(entries)

            _write_cell(ws, row, col, cell_value, s, align='center')

            # money totals (use actual item prices)
            for product, qty in entries:
                if product and product.sale_price is not None:
                    val_pln = float(product.sale_price) * qty
                    total_pln += val_pln
                    col_totals_pln[j] += val_pln
                if product and product.purchase_currency == 'KRW' and product.purchase_price is not None:
                    val_krw = float(product.purchase_price) * qty
                    total_krw += val_krw
                    col_totals_krw[j] += val_krw

        total_pln_val = _fmt_price(total_pln) if total_pln else None
        total_krw_val = _fmt_price(total_krw, decimals=0) if total_krw else None
        _write_cell(ws, row, len(sections) + 2, total_pln_val, s, align='center', bold=True)
        _write_cell(ws, row, len(sections) + 3, total_krw_val, s, align='center', bold=True)

        customer_totals_pln.append(total_pln)
        customer_totals_krw.append(total_krw)

    # SUMA (PLN) row
    suma_row = 4 + len(customers)
    _write_cell(ws, suma_row, 1, 'SUMA (PLN)', s, align='left', bold=True, fill=suma_fill)
    for j, total in enumerate(col_totals_pln):
        _write_cell(ws, suma_row, 2 + j, _fmt_price(total) if total else None,
                    s, align='center', bold=True, fill=suma_fill)
    total_all_pln = sum(customer_totals_pln)
    total_all_krw = sum(customer_totals_krw)
    _write_cell(ws, suma_row, len(sections) + 2,
                _fmt_price(total_all_pln) if total_all_pln else None,
                s, align='center', bold=True, fill=suma_fill)
    _write_cell(ws, suma_row, len(sections) + 3,
                _fmt_price(total_all_krw, decimals=0) if total_all_krw else None,
                s, align='center', bold=True, fill=suma_fill)

    # Blank row then Kurs
    rate = _get_krw_pln_rate()
    rate_row = suma_row + 2
    _write_cell(ws, rate_row, 1, 'Kurs KRW/PLN:', s, align='left', bold=True)
    if rate:
        _write_cell(ws, rate_row, 2, round(rate), s, align='center', bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 24
    for j in range(len(sections)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 20
    ws.column_dimensions[get_column_letter(len(sections) + 2)].width = 12
    ws.column_dimensions[get_column_letter(len(sections) + 3)].width = 14

    ws.freeze_panes = 'B4'


def _build_preorder_quantities_sheet(wb, page, s):
    """Zakładka 'Ilości' — macierz klient × sekcja z łącznymi ilościami sztuk."""
    ws = wb.create_sheet(title='Ilości', index=1)
    ws.sheet_properties.tabColor = _PURPLE_LIGHT

    data = _preorder_collect_data(page)
    sections = data['sections']
    customers = data['customers']
    matrix = data['matrix']

    headers = ['Klient'] + [_section_column_name(sec) for sec in sections]
    _write_header_row(ws, 1, headers, s)

    for i, customer in enumerate(customers):
        row = 2 + i
        _write_cell(ws, row, 1, customer['name'], s, align='left')
        for j, sec in enumerate(sections):
            entries = matrix.get((customer['key'], sec.id), [])
            qty_total = sum(q for _, q in entries) if entries else None
            _write_cell(ws, row, 2 + j, qty_total, s, align='center')

    ws.column_dimensions['A'].width = 24
    for j in range(len(sections)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16

    ws.freeze_panes = 'B2'


# ============================================
# Helpers
# ============================================

def _format_period(starts_at, ends_at):
    """Format date period string."""
    if starts_at and ends_at:
        return f"{starts_at.strftime('%d.%m.%Y')} - {ends_at.strftime('%d.%m.%Y')}"
    elif starts_at:
        return f"od {starts_at.strftime('%d.%m.%Y')}"
    elif ends_at:
        return f"do {ends_at.strftime('%d.%m.%Y')}"
    return "-"


# ============================================
# General Orders Excel (unchanged)
# ============================================

def generate_orders_excel(orders, filename_prefix="zamowienia"):
    """
    Generuje ogólny plik Excel z listą zamówień.

    Args:
        orders: Lista zamówień (Order objects)
        filename_prefix: Prefix nazwy pliku

    Returns:
        BytesIO: Buffer z plikiem .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Zamówienia"

    # Style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7B2CBF", end_color="7B2CBF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Nagłówki
    headers = [
        "Lp.", "Nr zamówienia", "Klient", "Email", "Telefon",
        "Status", "Typ", "Data", "Wartość (PLN)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dane
    for idx, order in enumerate(orders, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=order.order_number)
        ws.cell(row=idx + 1, column=3, value=order.customer_name or "-")
        ws.cell(row=idx + 1, column=4, value=order.customer_email or "-")

        phone = order.user.phone if order.user else None
        ws.cell(row=idx + 1, column=5, value=phone or "-")

        ws.cell(row=idx + 1, column=6, value=order.status_display)
        ws.cell(row=idx + 1, column=7, value="Exclusive" if order.offer_page_id else "Standard")
        ws.cell(row=idx + 1, column=8, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=idx + 1, column=9, value=float(order.total_amount) if order.total_amount else 0)

    # Dostosuj szerokość kolumn
    column_widths = [6, 18, 25, 30, 15, 20, 12, 18, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Zapisz do bufora
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
