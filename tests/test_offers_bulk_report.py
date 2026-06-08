from openpyxl import Workbook

from utils.excel_export import _safe_sheet_title, _styles, _write_sets_matrix


# --- _write_sets_matrix (współdzielona MACIERZ SETÓW: LIVE + raport zbiorczy) ---

def _cells_text(ws):
    out = []
    for r in ws.iter_rows(values_only=True):
        for c in r:
            if isinstance(c, str):
                out.append(c)
    return out


def test_sets_matrix_writes_header_columns_and_customers():
    wb = Workbook()
    ws = wb.active
    sets_info = [{
        'set_name': 'Set A',
        'set_max_sets': 2,
        'has_limit': True,
        'ordered_sets': 1,
        'total_sets_sold': 1,
        'full_set_sold': 0,
        'products': [{
            'product_name': 'Photocard',
            'is_full_set': False,
            'slots': [
                {'filled': True, 'customer': 'Anna'},
                {'filled': False, 'customer': None},
            ],
        }],
    }]
    next_row, max_cols = _write_sets_matrix(ws, sets_info, _styles(), start_row=1)
    text = _cells_text(ws)
    assert 'MACIERZ SETÓW' in text
    assert 'Produkt' in text
    assert 'Set 1' in text and 'Set 2' in text
    assert 'Photocard' in text
    assert 'Anna' in text          # nazwa klienta w wypełnionym slocie
    assert max_cols == 2
    assert next_row > 1


def test_sets_matrix_empty_returns_start_row():
    wb = Workbook()
    ws = wb.active
    next_row, max_cols = _write_sets_matrix(ws, [], _styles(), start_row=5)
    # brak setów: nic nie wpisujemy poza nagłówkiem sekcji; kolumny = 0
    assert max_cols == 0


def test_short_name_unchanged():
    assert _safe_sheet_title('Lato 2026', set()) == 'Lato 2026'


def test_long_name_truncated_to_31():
    name = 'A' * 50
    title = _safe_sheet_title(name, set())
    assert len(title) <= 31
    assert title == 'A' * 31


def test_illegal_chars_replaced():
    title = _safe_sheet_title('Sprzedaż: K-pop / [2026]', set())
    for ch in '[]:*?/\\':
        assert ch not in title


def test_duplicate_gets_suffix():
    used = {'lato 2026'}
    title = _safe_sheet_title('Lato 2026', used)
    assert title != 'Lato 2026'
    assert title.lower() not in used


def test_duplicate_suffix_fits_31_chars():
    used = {('a' * 31).lower()}
    title = _safe_sheet_title('A' * 31, used)
    assert len(title) <= 31
    assert title.lower() not in used


def test_empty_name_gets_fallback():
    assert _safe_sheet_title('', set()) == 'Oferta'


def test_does_not_mutate_used_titles():
    used = {'lato 2026'}
    _safe_sheet_title('Lato 2026', used)
    assert used == {'lato 2026'}
