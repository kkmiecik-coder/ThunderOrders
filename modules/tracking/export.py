import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_visits_xlsx(campaign, visits):
    """Generuje plik XLSX z wizytami kampanii."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wizyty QR'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='462C1A', end_color='462C1A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'Kampania: {campaign.name}'
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:I2')
    ws['A2'] = f'URL: https://thunderorders.cloud/qr/{campaign.slug}'

    ws.merge_cells('A3:I3')
    ws['A3'] = f'Eksport: {__import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")}'

    headers = [
        'Data/Godzina', 'Typ urz.', 'Przegladarka', 'System',
        'Kraj', 'Miasto', 'IP', 'Unikalny', 'Referer'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, visit in enumerate(visits, 6):
        ws.cell(row=row_idx, column=1, value=visit.visited_at.strftime('%Y-%m-%d %H:%M:%S') if visit.visited_at else '')
        ws.cell(row=row_idx, column=2, value=visit.device_type or '')
        ws.cell(row=row_idx, column=3, value=visit.browser or '')
        ws.cell(row=row_idx, column=4, value=visit.os or '')
        ws.cell(row=row_idx, column=5, value=visit.country or '')
        ws.cell(row=row_idx, column=6, value=visit.city or '')
        ws.cell(row=row_idx, column=7, value=visit.ip_address or '')
        ws.cell(row=row_idx, column=8, value='Tak' if visit.is_unique else 'Nie')
        ws.cell(row=row_idx, column=9, value=visit.referer or '')

        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = thin_border

    column_widths = [20, 10, 15, 15, 15, 15, 18, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
