"""
Orders Module - Routes
======================

Routes for orders management (Admin + Client + API).
Includes HTMX endpoints for partial updates.
"""

import json
import os
from flask import render_template, request, redirect, url_for, flash, jsonify, abort, make_response, current_app, send_from_directory
from flask_login import login_required, current_user
from sqlalchemy import or_, and_, func
from datetime import datetime
from decimal import Decimal, InvalidOperation

from werkzeug.exceptions import HTTPException
from werkzeug.utils import secure_filename
from modules.orders import orders_bp
from modules.orders.models import (
    Order, OrderItem, OrderRefund,
    OrderStatus, OrderType, WmsStatus,
    ShippingRequestStatus, ShippingRequest, ShippingRequestOrder,
    PaymentConfirmation
)
from modules.products.models import Product
from modules.auth.models import User
from modules.orders.forms import (
    OrderFilterForm, OrderStatusForm,
    OrderTrackingForm, RefundForm, BulkActionForm,
    ShippingAddressForm, PickupPointForm
)
from modules.orders.utils import (
    generate_order_number, detect_courier, get_tracking_url,
    calculate_order_total, get_order_summary, zaloguj_blad_z_identyfikatorem
)
from modules.orders.wms_utils import COURIER_NAMES
from extensions import db
from utils.decorators import role_required
from utils.activity_logger import log_activity
# from modules.emails.sender import send_email  # Uncomment when email module is ready


# Mapowanie akcji na polskie opisy i ikony (współdzielone między widokiem klienta i admina)
ORDER_ACTION_CONFIG = {
    'order_created': {'label': 'Zamówienie utworzone', 'icon': '📦'},
    'order_status_change': {'label': 'Zmiana statusu', 'icon': '🔄'},
    'order_status_auto_updated': {'label': 'Automatyczna zmiana statusu', 'icon': '⚡'},
    'order_updated': {'label': 'Zaktualizowano zamówienie', 'icon': '✏️'},
    'order_item_added': {'label': 'Dodano produkt', 'icon': '➕'},
    'order_item_added_custom': {'label': 'Dodano produkt niestandardowy', 'icon': '➕'},
    'order_item_removed': {'label': 'Usunięto produkt', 'icon': '🗑️'},
    'order_item_deleted': {'label': 'Usunięto produkt', 'icon': '🗑️'},
    'order_item_updated': {'label': 'Zaktualizowano produkt', 'icon': '✏️'},
    'order_products_added': {'label': 'Dodano produkty', 'icon': '➕'},
    'order_field_updated': {'label': 'Zaktualizowano dane zamówienia', 'icon': '✏️'},
    'order_payment_updated': {'label': 'Zaktualizowano płatność', 'icon': '💳'},
    'tracking_number_added': {'label': 'Dodano numer śledzenia', 'icon': '🚚'},
    'tracking_number_updated': {'label': 'Zaktualizowano numer śledzenia', 'icon': '🚚'},
    'shipping_requested': {'label': 'Utworzono zlecenie wysyłki', 'icon': '📬'},
    'shipping_cost_updated': {'label': 'Zaktualizowano koszt wysyłki', 'icon': '💰'},
    'comment_added': {'label': 'Dodano komentarz', 'icon': '💬'},
    'order_cancelled': {'label': 'Anulowano zamówienie', 'icon': '🚫'},
    'order_completed': {'label': 'Zamówienie zakończone', 'icon': '🎉'},
    'refund_issued': {'label': 'Wystawiono zwrot', 'icon': '💸'},
    'payment_confirmation_uploaded': {'label': 'Przesłano potwierdzenie płatności', 'icon': '📤'},
    'payment_confirmation_reuploaded': {'label': 'Ponownie przesłano potwierdzenie', 'icon': '🔄'},
    'payment_confirmation_approved': {'label': 'Zatwierdzono płatność', 'icon': '✅'},
    'payment_confirmation_rejected': {'label': 'Odrzucono płatność', 'icon': '❌'},
    'shipment_added': {'label': 'Dodano przesyłkę', 'icon': '📦'},
    'shipment_deleted': {'label': 'Usunięto przesyłkę', 'icon': '🗑️'},
    'order_packed': {'label': 'Zamówienie spakowane', 'icon': '📦'},
    'proxy_shipping_distributed': {'label': 'Naliczono koszt wysyłki proxy', 'icon': '🚢'},
    'customs_vat_distributed': {'label': 'Naliczono cło/VAT', 'icon': '🏛️'},
    'offer_closure_fulfillment': {'label': 'Rozliczenie zamknięcia offer', 'icon': '📊'},
}

# Mapowanie etapów płatności na polskie nazwy
PAYMENT_STAGE_LABELS = {
    'product': 'Produkt',
    'korean_shipping': 'Wysyłka KR',
    'customs_vat': 'Cło/VAT',
    'domestic_shipping': 'Wysyłka PL',
}


# ====================
# ADMIN ROUTES
# ====================

@orders_bp.route('/admin/orders')
@login_required
@role_required('admin', 'mod')
def admin_list():
    """
    Admin orders list with filters and pagination.
    Supports quick filters (order type) and advanced filters.
    """
    # Initialize filter form
    filter_form = OrderFilterForm(request.args)

    # Populate status choices dynamically
    statuses = OrderStatus.query.filter_by(is_active=True).order_by(OrderStatus.sort_order).all()
    filter_form.status.choices = [(s.slug, s.name) for s in statuses]

    # Base query
    query = Order.query

    # Apply filters
    if filter_form.order_type.data:
        query = query.filter(Order.order_type == filter_form.order_type.data)

    if filter_form.status.data:
        query = query.filter(Order.status.in_(filter_form.status.data))

    if filter_form.date_from.data:
        query = query.filter(Order.created_at >= filter_form.date_from.data)

    if filter_form.date_to.data:
        # Add 1 day to include the end date
        from datetime import timedelta
        end_date = filter_form.date_to.data + timedelta(days=1)
        query = query.filter(Order.created_at < end_date)

    if filter_form.search.data:
        search_term = f"%{filter_form.search.data}%"
        query = query.join(Order.user, isouter=True).filter(
            or_(
                Order.order_number.like(search_term),
                Order.custom_name.like(search_term),
                db.func.concat(User.first_name, ' ', User.last_name).like(search_term),
                User.email.like(search_term)
            )
        )

    # Products filter - find orders containing selected products
    if filter_form.products.data:
        product_ids_str = filter_form.products.data
        try:
            product_ids = [int(pid.strip()) for pid in product_ids_str.split(',') if pid.strip()]
            if product_ids:
                # Subquery: orders that have at least one of these products
                from modules.orders.models import OrderItem
                subquery = db.session.query(OrderItem.order_id).filter(
                    OrderItem.product_id.in_(product_ids)
                ).distinct().subquery()
                query = query.filter(Order.id.in_(subquery))
        except ValueError:
            pass  # Invalid product IDs, skip filter

    # Payment stage filters
    from modules.orders.models import PaymentConfirmation
    stage_map = {
        'pay_e1': 'product',
        'pay_e2': 'korean_shipping',
        'pay_e3': 'customs_vat',
        'pay_e4': 'domestic_shipping',
    }
    for field_name, stage_value in stage_map.items():
        filter_val = getattr(filter_form, field_name).data
        if filter_val:
            if filter_val == 'none':
                # Orders with no PaymentConfirmation for this stage
                no_conf_subquery = db.session.query(PaymentConfirmation.order_id).filter(
                    PaymentConfirmation.payment_stage == stage_value
                ).subquery()
                query = query.filter(~Order.id.in_(no_conf_subquery))
            else:
                # Orders with specific status for this stage
                conf_subquery = db.session.query(PaymentConfirmation.order_id).filter(
                    PaymentConfirmation.payment_stage == stage_value,
                    PaymentConfirmation.status == filter_val
                ).subquery()
                query = query.filter(Order.id.in_(conf_subquery))

    # Sorting
    sort_by = request.args.get('sort', 'created_at')
    sort_order = request.args.get('order', 'desc')

    if sort_by == 'order_number':
        query = query.order_by(Order.order_number.desc() if sort_order == 'desc' else Order.order_number.asc())
    elif sort_by == 'total_amount':
        query = query.order_by(Order.total_amount.desc() if sort_order == 'desc' else Order.total_amount.asc())
    else:  # Default: created_at
        query = query.order_by(Order.created_at.desc() if sort_order == 'desc' else Order.created_at.asc())

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = int(filter_form.per_page.data) if filter_form.per_page.data else 20

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Get status counts for sidebar
    status_counts = db.session.query(
        Order.status,
        func.count(Order.id)
    ).group_by(Order.status).all()
    status_counts_dict = {status: count for status, count in status_counts}

    # Get all statuses with their counts
    all_statuses = OrderStatus.query.filter_by(is_active=True).order_by(OrderStatus.sort_order).all()
    statuses_with_counts = []
    total_count = 0
    for status in all_statuses:
        count = status_counts_dict.get(status.slug, 0)
        total_count += count
        statuses_with_counts.append({
            'slug': status.slug,
            'name': status.name,
            'badge_color': status.badge_color,
            'count': count
        })

    # Filter args without 'page' to avoid duplicate in pagination url_for
    filter_args = {k: v for k, v in request.args.items() if k != 'page'}

    from utils.supplier_order_state import get_supplier_states_for_orders
    supplier_states = get_supplier_states_for_orders(pagination.items)

    return render_template(
        'admin/orders/list.html',
        orders=pagination,
        filter_form=filter_form,
        statuses_with_counts=statuses_with_counts,
        total_orders_count=total_count,
        supplier_states=supplier_states,
        page_title='Zamówienia',
        filter_args=filter_args
    )


@orders_bp.route('/api/orders/create-for-client', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def api_create_order_for_client():
    """
    API endpoint to create a new empty order for a client.
    Used by the modal in orders list - creates order and returns order ID for redirect.
    """
    from modules.auth.models import User

    try:
        data = request.get_json()
        client_id = data.get('client_id')

        if not client_id:
            return jsonify({
                'success': False,
                'message': 'Nie podano ID klienta'
            }), 400

        # Validate client exists
        client = db.session.get(User, client_id)
        if not client or client.role != 'client':
            return jsonify({
                'success': False,
                'message': 'Nie znaleziono klienta'
            }), 404

        # Generate order number (use 'on_hand' type for manual orders)
        order_number = generate_order_number('on_hand')

        # Create empty order
        new_order = Order(
            order_number=order_number,
            user_id=client.id,
            order_type='on_hand',  # Manual orders use on_hand type
            status='nowe',
            total_amount=0
        )
        db.session.add(new_order)
        db.session.commit()

        # Log activity
        log_activity(
            user=current_user,
            action='order_created',
            entity_type='order',
            entity_id=new_order.id,
            new_value={'order_number': order_number, 'client': client.full_name}
        )

        return jsonify({
            'success': True,
            'message': f'Zamówienie {order_number} zostało utworzone',
            'order_id': new_order.id,
            'order_number': order_number,
            'redirect_url': url_for('orders.admin_detail', order_id=new_order.id)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas tworzenia zamówienia: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/create', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'mod')
def admin_create_order():
    """
    Admin order creation page.
    Creates a new manual order for a selected client.
    DEPRECATED: This page is kept for backwards compatibility.
    New flow uses API endpoint + redirect to order detail page.
    """
    from modules.auth.models import User
    from modules.products.models import Product

    client_id = request.args.get('client_id', type=int)

    # Validate client exists
    client = None
    if client_id:
        client = db.session.get(User, client_id)
        if not client or client.role != 'client':
            flash('Nie znaleziono klienta', 'error')
            return redirect(url_for('orders.admin_list'))

    if request.method == 'POST':
        # Handle order creation
        try:
            # Get client from form or URL
            form_client_id = request.form.get('client_id', type=int)
            if not client and form_client_id:
                client = db.session.get(User, form_client_id)

            if not client:
                flash('Wybierz klienta', 'error')
                return redirect(url_for('orders.admin_create_order'))

            # Get products from form
            product_ids = request.form.getlist('product_id[]', type=int)
            quantities = request.form.getlist('quantity[]', type=int)

            if not product_ids or not quantities:
                flash('Dodaj przynajmniej jeden produkt', 'error')
                return redirect(url_for('orders.admin_create_order', client_id=client.id))

            # Generate order number (use 'on_hand' type for manual orders)
            order_number = generate_order_number('on_hand')

            # Create order
            new_order = Order(
                order_number=order_number,
                user_id=client.id,
                order_type='on_hand',  # Manual orders use on_hand type
                status='nowe',
                total_amount=0
            )
            db.session.add(new_order)
            db.session.flush()  # Get order ID

            # Add order items
            total = 0
            for pid, qty in zip(product_ids, quantities):
                if qty <= 0:
                    continue
                product = db.session.get(Product, pid)
                if not product:
                    continue

                item_total = product.sale_price * qty
                total += item_total

                order_item = OrderItem(
                    order_id=new_order.id,
                    product_id=pid,
                    quantity=qty,
                    price=product.sale_price,
                    total=item_total
                )
                db.session.add(order_item)

            new_order.total_amount = total
            db.session.commit()

            # Log activity
            log_activity(
                user=current_user,
                action='order_created',
                entity_type='order',
                entity_id=new_order.id,
                new_value={'order_number': order_number, 'client': client.full_name}
            )

            flash(f'Zamówienie {order_number} zostało utworzone', 'success')
            return redirect(url_for('orders.admin_detail', order_id=new_order.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Błąd podczas tworzenia zamówienia: {str(e)}', 'error')
            return redirect(url_for('orders.admin_create_order', client_id=client_id))

    # GET request - show form
    # Get all active products for selection
    products = Product.query.filter_by(is_active=True).order_by(Product.name).all()

    return render_template(
        'admin/orders/create.html',
        client=client,
        products=products,
        page_title='Nowe zamówienie'
    )


@orders_bp.route('/admin/orders/<int:order_id>')
@login_required
@role_required('admin', 'mod')
def admin_detail(order_id):
    """
    Admin order detail page.
    Shows full order information, timeline, products, etc.
    """
    order = Order.query.get_or_404(order_id)

    # Eager load relationships
    # Sort items: fulfilled items first, unfulfilled items last
    order_items = sorted(
        OrderItem.query.filter_by(order_id=order_id).all(),
        key=lambda item: (
            2 if item.is_set_fulfilled is False else (1 if item.is_set_fulfilled is True else 0),
            item.id
        )
    )
    refunds = OrderRefund.query.filter_by(order_id=order_id).order_by(OrderRefund.created_at.desc()).all()

    # Forms
    status_form = OrderStatusForm()
    tracking_form = OrderTrackingForm(obj=order)
    refund_form = RefundForm()
    shipping_address_form = ShippingAddressForm(obj=order)
    pickup_point_form = PickupPointForm(obj=order)

    # Populate status choices
    statuses = OrderStatus.query.filter_by(is_active=True).order_by(OrderStatus.sort_order).all()
    status_form.status.choices = [(s.slug, s.name) for s in statuses]
    status_form.status.data = order.status

    # Build statuses list with colors for custom dropdown
    statuses_with_colors = [
        {'slug': s.slug, 'name': s.name, 'color': s.badge_color}
        for s in statuses
    ]

    # Set default refund amount to order total
    refund_form.amount.data = order.total_amount

    # Build timeline (merge comments and events)
    timeline = []

    # Add created event
    timeline.append({
        'type': 'created',
        'created_at': order.created_at,
        'icon': '📦',
        'message': 'Zamówienie utworzone'
    })

    # Add refunds
    for refund in refunds:
        timeline.append({
            'type': 'refund',
            'created_at': refund.created_at,
            'amount': refund.amount,
            'reason': refund.reason,
            'status': refund.status,
            'creator': refund.creator
        })

    # Add activity logs for this order
    from modules.admin.models import ActivityLog
    import json
    activity_logs = ActivityLog.query.filter_by(
        entity_type='order',
        entity_id=order.id
    ).order_by(ActivityLog.created_at).all()

    for log in activity_logs:
        # Parse new_value JSON if exists
        new_value_data = {}
        if log.new_value:
            try:
                new_value_data = json.loads(log.new_value)
            except:
                pass

        old_value_data = {}
        if log.old_value:
            try:
                old_value_data = json.loads(log.old_value)
            except:
                pass

        timeline.append({
            'type': 'activity',
            'created_at': log.created_at,
            'action': log.action,
            'user': log.user,
            'old_value': old_value_data,
            'new_value': new_value_data
        })

    # Maile wysłane w sprawie tego zamówienia (EmailLog). Osobne źródło od
    # ActivityLog, bo to przebieg techniczny, nie zmiana danych — ale w historii
    # zmian ląduje na tej samej osi czasu, żeby dało się zestawić „zmieniono
    # status o 14:02" z „mail o statusie wyszedł o 17:05, w drugiej próbie".
    from sqlalchemy import and_, or_
    from modules.admin.models import EmailLog

    # Maile wysyłkowe („paczka wysłana", „zmiana statusu zlecenia") logują się z
    # kontekstem ZLECENIA, nie zamówienia — a admin szuka ich właśnie tutaj. Bierzemy
    # zarówno zlecenie, w którym zamówienie fizycznie leży, jak i źródłowe: po
    # konsolidacji to dwa różne zlecenia, a mail do klienta idzie z numerem źródłowego.
    zlecenia_ids = set()
    for ro in order.shipping_request_orders:
        if ro.shipping_request_id:
            zlecenia_ids.add(ro.shipping_request_id)
        if ro.source_request_id:
            zlecenia_ids.add(ro.source_request_id)

    warunki = [and_(EmailLog.entity_type == 'order', EmailLog.entity_id == order.id)]
    if zlecenia_ids:
        warunki.append(and_(EmailLog.entity_type == 'shipping_request',
                            EmailLog.entity_id.in_(zlecenia_ids)))

    maile = EmailLog.query.filter(or_(*warunki)).order_by(EmailLog.created_at).all()

    for wpis in maile:
        timeline.append({
            'type': 'email',
            'created_at': wpis.created_at,
            'email': wpis,
        })

    # Sort timeline by date (oldest first, newest at bottom)
    timeline.sort(key=lambda x: x['created_at'], reverse=False)

    # Get categories and product series for add products modal
    from modules.products.models import Category, ProductSeries
    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    product_series = ProductSeries.query.filter_by(is_active=True).order_by(ProductSeries.name).all()

    # Get active payment methods from database
    from modules.payments.models import PaymentMethod
    payment_methods = PaymentMethod.get_active()

    # WMS session info for order detail
    wms_session_order = None
    if order.packed_at:
        from modules.orders.wms_models import WmsSessionOrder
        wms_session_order = WmsSessionOrder.query.filter_by(order_id=order.id).first()

    # Set probability for live offer sales (per order item, using stored set_number)
    set_probabilities = {}
    if order.order_type == 'exclusive' and order.offer_page_id:
        page_obj = order.offer_page
        if page_obj and not page_obj.is_fully_closed:
            from modules.offers.reservation import get_set_probabilities
            set_probabilities = get_set_probabilities(order)

    return render_template(
        'admin/orders/detail.html',
        order=order,
        order_items=order_items,
        timeline=timeline,
        status_form=status_form,
        statuses_with_colors=statuses_with_colors,
        tracking_form=tracking_form,
        refund_form=refund_form,
        shipping_address_form=shipping_address_form,
        pickup_point_form=pickup_point_form,
        categories=categories,
        product_series=product_series,
        payment_methods=payment_methods,
        wms_session_order=wms_session_order,
        set_probabilities=set_probabilities,
        action_config=ORDER_ACTION_CONFIG,
        payment_stage_labels=PAYMENT_STAGE_LABELS,
        page_title=f'Zamówienie {order.order_number}'
    )


# ====================
# ADMIN HTMX ENDPOINTS
# ====================

@orders_bp.route('/admin/orders/<int:order_id>/status', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_status(order_id):
    """
    HTMX endpoint for changing order status.
    Returns updated status badge HTML with HX-Trigger for toast.
    """
    order = Order.query.get_or_404(order_id)

    # Get status from form data (works with both WTForms and custom dropdown)
    new_status = request.form.get('status')

    if not new_status:
        response = make_response('<span class="badge badge-error">Błąd: brak statusu</span>', 400)
        response.headers['HX-Trigger'] = json.dumps({'showToast': {'message': 'Błąd: brak statusu', 'type': 'error'}})
        return response

    # Validate that status exists
    status_obj = OrderStatus.query.filter_by(slug=new_status, is_active=True).first()
    if not status_obj:
        response = make_response('<span class="badge badge-error">Błąd: nieprawidłowy status</span>', 400)
        response.headers['HX-Trigger'] = json.dumps({'showToast': {'message': 'Błąd: nieprawidłowy status', 'type': 'error'}})
        return response

    old_status = order.status
    old_status_name = order.status_display_name

    if old_status != new_status:
        order.status = new_status
        order.updated_at = datetime.now()

        from modules.orders.consolidation import (
            STATUSY_WYPINAJACE_Z_PACZKI, odepnij_anulowane_zamowienie)
        if new_status in STATUSY_WYPINAJACE_Z_PACZKI:
            # Anulowane zamówienie (i skierowane do zwrotu) nigdy nie spełni
            # bramek gotowości paczki zbiorczej ("all spakowane" / komplet E4) —
            # wypinamy je od razu, żeby nie zablokowało wysyłki pozostałym
            # uczestnikom.
            odepnij_anulowane_zamowienie(order)

        db.session.commit()

        # Auto-add to collection when delivered
        if new_status == 'dostarczone' and old_status != 'dostarczone':
            from modules.client.collection_utils import auto_add_order_to_collection
            try:
                auto_add_order_to_collection(order)
                db.session.commit()
            except Exception as e:
                current_app.logger.error(f'Collection auto-add error: {e}')

        # Activity log
        log_activity(
            user=current_user,
            action='order_status_change',
            entity_type='order',
            entity_id=order.id,
            old_value={'status': old_status},
            new_value={'status': new_status}
        )

        # Send email + push notification to customer
        from utils.email_manager import EmailManager
        EmailManager.notify_status_change(order, old_status_name, order.status_display_name)
        from utils.push_manager import PushManager
        PushManager.notify_status_change(order, old_status_name, order.status_display_name)

        # Return updated badge HTML with HX-Trigger for toast
        badge_html = f'<span class="badge" style="background-color: {order.status_badge_color}; color: #fff;" id="statusBadge">{order.status_display_name}</span>'
        response = make_response(badge_html)
        response.headers['HX-Trigger'] = json.dumps({
            'showToast': {
                'message': f'Status zmieniony: {old_status_name} → {order.status_display_name}',
                'type': 'success'
            }
        })
        return response

    # No change - return current badge
    return f'<span class="badge" style="background-color: {order.status_badge_color}; color: #fff;" id="statusBadge">{order.status_display_name}</span>'


@orders_bp.route('/admin/orders/<int:order_id>/shipping-address', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_shipping_address(order_id):
    """
    HTMX endpoint for updating shipping address.
    Returns success message.
    """
    order = Order.query.get_or_404(order_id)
    form = ShippingAddressForm()

    if form.validate_on_submit():
        order.shipping_name = form.shipping_name.data
        order.shipping_address = form.shipping_address.data
        order.shipping_postal_code = form.shipping_postal_code.data
        order.shipping_city = form.shipping_city.data
        order.shipping_voivodeship = form.shipping_voivodeship.data
        order.shipping_country = form.shipping_country.data or 'Polska'
        order.updated_at = datetime.now()
        db.session.commit()

        flash('Adres dostawy zaktualizowany', 'success')
        return '<div class="alert alert-success">Adres dostawy zapisany</div>'

    return '<div class="alert alert-error">Błąd podczas aktualizacji adresu</div>', 400


@orders_bp.route('/admin/orders/<int:order_id>/pickup-point', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_pickup_point(order_id):
    """
    HTMX endpoint for updating pickup point.
    Returns success message.
    """
    order = Order.query.get_or_404(order_id)
    form = PickupPointForm()

    if form.validate_on_submit():
        order.pickup_courier = form.pickup_courier.data
        order.pickup_point_id = form.pickup_point_id.data
        order.pickup_address = form.pickup_address.data
        order.pickup_postal_code = form.pickup_postal_code.data
        order.pickup_city = form.pickup_city.data
        order.updated_at = datetime.now()
        db.session.commit()

        flash('Punkt odbioru zaktualizowany', 'success')
        return '<div class="alert alert-success">Punkt odbioru zapisany</div>'

    return '<div class="alert alert-error">Błąd podczas aktualizacji punktu odbioru</div>', 400


@orders_bp.route('/admin/orders/<int:order_id>/tracking', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_tracking(order_id):
    """
    HTMX endpoint for updating tracking information.
    Returns updated tracking info HTML.
    """
    order = Order.query.get_or_404(order_id)
    form = OrderTrackingForm()

    if form.validate_on_submit():
        old_tracking = order.tracking_number
        old_courier = order.courier
        order.tracking_number = form.tracking_number.data
        order.courier = form.courier.data
        order.updated_at = datetime.now()
        db.session.commit()

        # Activity log
        action = 'tracking_number_added' if not old_tracking else 'tracking_number_updated'
        log_activity(
            user=current_user,
            action=action,
            entity_type='order',
            entity_id=order.id,
            old_value={'tracking_number': old_tracking, 'courier': old_courier},
            new_value={'tracking_number': order.tracking_number, 'courier': order.courier}
        )

        # Send tracking email if tracking number was added (not just updated)
        if order.tracking_number and not old_tracking:
            from utils.email_manager import EmailManager
            from utils.push_manager import PushManager
            EmailManager.notify_tracking_added(
                order,
                tracking_number=order.tracking_number,
                courier=order.courier,
                courier_name=COURIER_NAMES.get(order.courier, order.courier or 'Kurier')
            )
            PushManager.notify_tracking_added(
                order,
                tracking_number=order.tracking_number,
                courier_name=COURIER_NAMES.get(order.courier, order.courier or 'Kurier')
            )

        flash('Informacje o śledzeniu zaktualizowane', 'success')

        # Return updated tracking info HTML
        return render_template('admin/orders/_tracking_info.html', order=order)

    return '<div class="alert alert-error">Błąd podczas aktualizacji</div>', 400


@orders_bp.route('/admin/orders/<int:order_id>/refund', methods=['POST'])
@login_required
@role_required('admin')  # Only admin can issue refunds
def admin_issue_refund(order_id):
    """
    Issue refund for order.
    Changes order status to 'do_zwrotu' or 'czesciowo_zwrocone'.
    """
    order = Order.query.get_or_404(order_id)
    form = RefundForm()

    if form.validate_on_submit():
        # Create refund record
        refund = OrderRefund(
            order_id=order.id,
            amount=form.amount.data,
            reason=form.reason.data,
            status='pending',
            created_by=current_user.id
        )
        db.session.add(refund)

        # Update order status
        total_refunded = sum(r.amount for r in order.refunds if r.status == 'completed')
        total_refunded += form.amount.data

        if total_refunded >= order.total_amount:
            order.status = 'zwrocone'
        else:
            order.status = 'czesciowo_zwrocone'

        order.updated_at = datetime.now()
        db.session.commit()

        # Activity log
        # log_activity(
        #     user=current_user,
        #     action='refund_issued',
        #     entity_type='order',
        #     entity_id=order.id,
        #     new_value={
        #         'refund_id': refund.id,
        #         'amount': float(refund.amount),
        #         'reason': refund.reason
        #     }
        # )

        # Email notification
        # send_email(
        #     to=order.customer_email,
        #     template_type='refund_notification',
        #     context={
        #         'order': get_order_summary(order),
        #         'refund_amount': float(refund.amount),
        #         'refund_reason': refund.reason
        #     }
        # )

        flash(f'Zwrot w kwocie {refund.amount} PLN został utworzony', 'success')
        return redirect(url_for('orders.admin_detail', order_id=order.id))

    flash('Błąd podczas tworzenia zwrotu', 'error')
    return redirect(url_for('orders.admin_detail', order_id=order.id))


# ====================
# PAYMENT ENDPOINT
# ====================

@orders_bp.route('/admin/orders/<int:order_id>/payment', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_payment(order_id):
    """
    Update order payment amount.
    Returns JSON response.
    """
    order = Order.query.get_or_404(order_id)

    try:
        data = request.get_json()
        paid_amount = data.get('paid_amount')

        if paid_amount is None:
            return jsonify({
                'success': False,
                'message': 'Kwota płatności jest wymagana'
            }), 400

        # Convert to Decimal
        try:
            paid_amount = Decimal(str(paid_amount))
        except:
            return jsonify({
                'success': False,
                'message': 'Nieprawidłowa kwota'
            }), 400

        if paid_amount < 0:
            return jsonify({
                'success': False,
                'message': 'Kwota nie może być ujemna'
            }), 400

        # Store old value for logging
        old_paid_amount = order.paid_amount

        # Update payment
        order.paid_amount = paid_amount
        order.updated_at = datetime.now()
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='order_payment_updated',
            entity_type='order',
            entity_id=order.id,
            old_value={'paid_amount': float(old_paid_amount) if old_paid_amount else 0},
            new_value={'paid_amount': float(paid_amount)}
        )

        return jsonify({
            'success': True,
            'message': 'Płatność została zaktualizowana',
            'paid_amount': float(paid_amount),
            'is_fully_paid': order.is_fully_paid,
            'is_partially_paid': order.is_partially_paid,
            'is_overpaid': order.is_overpaid
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd: {str(e)}'
        }), 500


# ====================
# UPDATE ORDER FIELD ENDPOINT
# ====================

@orders_bp.route('/admin/orders/<int:order_id>/update-field', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_update_order_field(order_id):
    """
    Update a single field of an order.
    Returns JSON response with updated values.
    """
    order = Order.query.get_or_404(order_id)

    try:
        data = request.get_json()
        field = data.get('field')
        value = data.get('value')

        if not field:
            return jsonify({
                'success': False,
                'message': 'Nazwa pola jest wymagana'
            }), 400

        # Allowed fields that can be updated
        allowed_fields = ['delivery_method', 'shipping_cost', 'proxy_shipping_cost', 'customs_vat_sale_cost', 'payment_method', 'admin_notes']

        if field not in allowed_fields:
            return jsonify({
                'success': False,
                'message': f'Pole "{field}" nie może być aktualizowane'
            }), 400

        # Store old value for logging
        old_value = getattr(order, field)

        # Update the field
        if field in ('shipping_cost', 'proxy_shipping_cost', 'customs_vat_sale_cost'):
            from decimal import Decimal
            is_blank = value is None or (isinstance(value, str) and not value.strip())
            try:
                if is_blank:
                    # Cło/VAT ma trzy stany (patrz Order.has_customs_vat_stage):
                    # NULL = nieustalone, 0 = ustalono bez podatku, > 0 = z podatkiem.
                    # Puste pole to "nieustalone", a NIE decyzja "bez cła" — zapis 0
                    # skasowałby klientowi etap E3 i odblokował wysyłkę.
                    # Pozostałe koszty zachowują dotychczasowe zachowanie (0.00).
                    value = None if field == 'customs_vat_sale_cost' else Decimal('0.00')
                else:
                    value = Decimal(str(value))
            except:
                return jsonify({
                    'success': False,
                    'message': 'Nieprawidłowa kwota'
                }), 400

            if value is not None and value < 0:
                return jsonify({
                    'success': False,
                    'message': 'Kwota nie może być ujemna'
                }), 400

            # Ta sama blokada co w modalu Cło/VAT (modules/products/routes.py):
            # nie wolno wyzerować cła, które klient już opłacił albo zgłosił
            # do weryfikacji — powstałaby nadpłata do ręcznego zwrotu.
            if (field == 'customs_vat_sale_cost'
                    and old_value is not None and old_value > 0
                    and (value is None or value == 0)
                    and order.stage_3_status in ('approved', 'pending')):
                return jsonify({
                    'success': False,
                    'message': (f'Nie można wyzerować Cła/VAT — zamówienie '
                                f'{order.order_number} ma już opłacony ten etap.')
                }), 409

            setattr(order, field, value)
        elif field == 'delivery_method':
            order.delivery_method = value if value else None
        elif field == 'payment_method':
            order.payment_method = value if value else None
        elif field == 'admin_notes':
            order.admin_notes = value.strip() if value else None

        order.updated_at = datetime.now()
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='order_field_updated',
            entity_type='order',
            entity_id=order.id,
            old_value={field: str(old_value) if old_value is not None else None},
            new_value={field: str(value) if value is not None else None}
        )

        # Return updated values
        response_data = {
            'success': True,
            'message': 'Zaktualizowano pomyślnie',
            'field': field,
            'value': str(value) if value else None
        }

        # For cost fields, also return totals and payment status
        if field in ('shipping_cost', 'proxy_shipping_cost', 'customs_vat_sale_cost'):
            response_data['grand_total'] = float(order.grand_total)
            response_data['proxy_shipping_total'] = float(order.proxy_shipping_total)
            response_data['customs_vat_total'] = float(order.customs_vat_total)
            response_data['paid_amount'] = float(order.paid_amount) if order.paid_amount else 0
            response_data['is_fully_paid'] = order.is_fully_paid
            response_data['is_partially_paid'] = order.is_partially_paid
            response_data['is_overpaid'] = order.is_overpaid

        # Email notification for cost fields
        if field in ('proxy_shipping_cost', 'customs_vat_sale_cost', 'shipping_cost') and value and float(value) > 0:
            from utils.email_manager import EmailManager
            cost_type_map = {
                'proxy_shipping_cost': 'proxy_shipping',
                'customs_vat_sale_cost': 'customs_vat',
                'shipping_cost': 'domestic_shipping'
            }
            cost_type = cost_type_map[field]
            EmailManager.notify_cost_added(order, cost_type, float(value))
            from utils.push_manager import PushManager
            PushManager.notify_cost_added(order, cost_type, float(value))

        return jsonify(response_data)

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd: {str(e)}'
        }), 500


# ====================
# SHIPMENTS ENDPOINTS
# ====================

@orders_bp.route('/admin/orders/<int:order_id>/shipments', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_add_shipment(order_id):
    """
    Add a shipment to order.
    Returns JSON response.
    """
    from modules.orders.models import OrderShipment

    order = Order.query.get_or_404(order_id)

    try:
        data = request.get_json()
        tracking_number = data.get('tracking_number', '').strip()
        courier = data.get('courier', '').strip()

        if not tracking_number:
            return jsonify({
                'success': False,
                'message': 'Numer przesyłki jest wymagany'
            }), 400

        if not courier:
            return jsonify({
                'success': False,
                'message': 'Kurier jest wymagany'
            }), 400

        # Check if shipment already exists
        existing = OrderShipment.query.filter_by(
            order_id=order_id,
            tracking_number=tracking_number
        ).first()

        if existing:
            return jsonify({
                'success': False,
                'message': 'Przesyłka o tym numerze już istnieje'
            }), 400

        # Create new shipment
        shipment = OrderShipment(
            order_id=order_id,
            tracking_number=tracking_number,
            courier=courier,
            created_by=current_user.id
        )
        db.session.add(shipment)
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='shipment_added',
            entity_type='order',
            entity_id=order.id,
            new_value={
                'tracking_number': tracking_number,
                'courier': courier,
                'order_number': order.order_number
            }
        )

        # Send tracking email + push to customer
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        EmailManager.notify_tracking_added(
            order,
            tracking_number=shipment.tracking_number,
            courier=shipment.courier,
            courier_name=shipment.courier_display_name,
            tracking_url=shipment.tracking_url
        )
        PushManager.notify_tracking_added(
            order,
            tracking_number=shipment.tracking_number,
            courier_name=shipment.courier_display_name
        )

        return jsonify({
            'success': True,
            'message': 'Przesyłka została dodana',
            'shipment': {
                'id': shipment.id,
                'tracking_number': shipment.tracking_number,
                'courier': shipment.courier,
                'courier_name': shipment.courier_display_name,
                'tracking_url': shipment.tracking_url,
                'created_at': shipment.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas dodawania przesyłki: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/shipments/<int:shipment_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'mod')
def admin_delete_shipment(order_id, shipment_id):
    """
    Delete a shipment from order.
    Returns JSON response.
    """
    from modules.orders.models import OrderShipment

    order = Order.query.get_or_404(order_id)
    shipment = OrderShipment.query.filter_by(id=shipment_id, order_id=order_id).first_or_404()

    try:
        tracking_number = shipment.tracking_number
        courier = shipment.courier

        db.session.delete(shipment)
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='shipment_deleted',
            entity_type='order',
            entity_id=order.id,
            old_value={
                'tracking_number': tracking_number,
                'courier': courier,
                'order_number': order.order_number
            }
        )

        return jsonify({
            'success': True,
            'message': 'Przesyłka została usunięta'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas usuwania przesyłki: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/delete', methods=['DELETE', 'POST'])
@login_required
@role_required('admin')  # Only admin can delete
def admin_delete_order(order_id):
    """
    Delete order (admin only).
    Returns JSON response for AJAX/HTMX.
    """
    order = Order.query.get_or_404(order_id)

    # Check if order is linked to an active WMS session
    from modules.orders.wms_models import WmsSessionOrder, WmsSession
    active_wms = WmsSessionOrder.query.join(WmsSession).filter(
        WmsSessionOrder.order_id == order.id,
        WmsSession.status.in_(['active', 'paused'])
    ).first()
    if active_wms:
        return jsonify({
            'success': False,
            'message': f'Zamówienie {order.order_number} jest powiązane z aktywną sesją WMS i nie może zostać usunięte.'
        }), 400

    # Remove old WMS junction records (from completed/cancelled sessions)
    WmsSessionOrder.query.filter_by(order_id=order.id).delete()

    db.session.delete(order)
    db.session.commit()

    flash(f'Zamówienie {order.order_number} zostało usunięte', 'success')

    return jsonify({'success': True, 'message': 'Zamówienie usunięte'}), 200


# ====================
# BULK ACTIONS
# ====================

@orders_bp.route('/admin/orders/bulk/status', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def bulk_status_change():
    """
    Bulk status change for multiple orders.
    Returns JSON response with success count.
    """
    try:
        data = request.get_json()
        order_ids = data.get('order_ids', [])
        new_status = data.get('status')

        if not order_ids:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano żadnych zamówień'
            }), 400

        if not new_status:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano statusu'
            }), 400

        # Validate status exists
        status_obj = OrderStatus.query.filter_by(slug=new_status, is_active=True).first()
        if not status_obj:
            return jsonify({
                'success': False,
                'message': 'Nieprawidłowy status'
            }), 400

        # Update orders
        updated_count = 0
        email_queue = []
        for order_id in order_ids:
            order = db.session.get(Order, order_id)
            if order:
                old_status = order.status
                old_status_name = order.status_display_name
                if old_status != new_status:
                    order.status = new_status
                    order.updated_at = datetime.now()
                    updated_count += 1

                    from modules.orders.consolidation import (
                        STATUSY_WYPINAJACE_Z_PACZKI, odepnij_anulowane_zamowienie)
                    if new_status in STATUSY_WYPINAJACE_Z_PACZKI:
                        # Patrz komentarz w admin_update_status — to samo dotyczy
                        # akcji masowej: zamówienie anulowane albo do zwrotu wypina
                        # się z paczki zbiorczej, żeby nie blokować wysyłki innym
                        # uczestnikom.
                        odepnij_anulowane_zamowienie(order)

                    # Activity log
                    log_activity(
                        user=current_user,
                        action='order_status_change',
                        entity_type='order',
                        entity_id=order.id,
                        old_value={'status': old_status},
                        new_value={'status': new_status}
                    )

                    # Queue email notification
                    if order.customer_email:
                        email_queue.append({
                            'order': order,
                            'old_status': old_status_name,
                            'new_status': order.status_display_name
                        })

        db.session.commit()

        # Auto-add to collection when delivered (bulk)
        if new_status == 'dostarczone':
            from modules.client.collection_utils import auto_add_order_to_collection
            for oid in order_ids:
                o = db.session.get(Order, oid)
                if o:
                    try:
                        auto_add_order_to_collection(o)
                    except Exception as e:
                        current_app.logger.error(f'Collection auto-add error for order {oid}: {e}')
            db.session.commit()

        # Send email + push notifications after successful commit
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        for email_data in email_queue:
            EmailManager.notify_status_change(
                email_data['order'],
                email_data['old_status'],
                email_data['new_status']
            )
            PushManager.notify_status_change(
                email_data['order'],
                email_data['old_status'],
                email_data['new_status']
            )

        return jsonify({
            'success': True,
            'message': f'Zmieniono status {updated_count} zamówień na "{status_obj.name}"',
            'updated_count': updated_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas zmiany statusu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/bulk/delete', methods=['POST'])
@login_required
@role_required('admin')  # Only admin can delete
def bulk_delete():
    """
    Bulk delete multiple orders.
    Returns JSON response with success count.
    """
    try:
        data = request.get_json()
        order_ids = data.get('order_ids', [])

        if not order_ids:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano żadnych zamówień'
            }), 400

        # Delete orders (skip those linked to active WMS sessions)
        from modules.orders.wms_models import WmsSessionOrder, WmsSession
        deleted_count = 0
        deleted_numbers = []
        skipped_numbers = []
        for order_id in order_ids:
            order = db.session.get(Order, order_id)
            if order:
                active_wms = WmsSessionOrder.query.join(WmsSession).filter(
                    WmsSessionOrder.order_id == order.id,
                    WmsSession.status.in_(['active', 'paused'])
                ).first()
                if active_wms:
                    skipped_numbers.append(order.order_number)
                    continue
                # Remove old WMS junction records (from completed/cancelled sessions)
                WmsSessionOrder.query.filter_by(order_id=order.id).delete()
                deleted_numbers.append(order.order_number)
                db.session.delete(order)
                deleted_count += 1

        db.session.commit()

        # Activity log for bulk delete
        log_activity(
            user=current_user,
            action='orders_bulk_deleted',
            entity_type='order',
            entity_id=None,
            old_value={'order_numbers': deleted_numbers, 'count': deleted_count}
        )

        message = f'Usunięto {deleted_count} zamówień'
        if skipped_numbers:
            message += f'. Pominięto {len(skipped_numbers)} zamówień powiązanych z sesją WMS ({", ".join(skipped_numbers)})'

        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'skipped_count': len(skipped_numbers)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas usuwania zamówień: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/export')
@login_required
@role_required('admin', 'mod')
def export_orders():
    """
    Export selected orders to XLSX with nice formatting.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Get order IDs from query param
    ids_param = request.args.get('ids', '')

    if not ids_param:
        flash('Nie wybrano żadnych zamówień do eksportu', 'error')
        return redirect(url_for('orders.admin_list'))

    try:
        order_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
    except ValueError:
        flash('Nieprawidłowe ID zamówień', 'error')
        return redirect(url_for('orders.admin_list'))

    # Get orders
    orders = Order.query.filter(Order.id.in_(order_ids)).order_by(Order.created_at.desc()).all()

    if not orders:
        flash('Nie znaleziono zamówień do eksportu', 'error')
        return redirect(url_for('orders.admin_list'))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Zamówienia"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="5A189A", end_color="5A189A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_alignment = Alignment(vertical="center", wrap_text=True)
    currency_alignment = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    alt_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Header row
    headers = [
        'Numer zamówienia',
        'Data utworzenia',
        'Status',
        'Typ',
        'Klient',
        'Email klienta',
        'Telefon klienta',
        'Produkty',
        'Suma (PLN)',
        'Wysyłka (PLN)',
        'Razem (PLN)',
        'Wpłacono (PLN)',
        'Dostawa',
        'Płatność',
        'Uwagi admina'
    ]

    # Column widths
    column_widths = [18, 18, 15, 12, 20, 25, 15, 50, 12, 12, 12, 12, 15, 15, 30]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Set header row height
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_num, order in enumerate(orders, 2):
        # Get products string
        products_list = []
        for item in order.items:
            products_list.append(f"{item.product_name} x{item.quantity} ({item.price:.2f} PLN)")
        products_str = "\n".join(products_list)

        # Get customer info
        if order.user:
            customer_name = order.user.full_name
            customer_email = order.user.email
            customer_phone = order.user.phone or ''
        else:
            customer_name = 'Nieznany'
            customer_email = ''
            customer_phone = ''

        # Get type display name
        type_name = order.type_rel.name if order.type_rel else (order.order_type or '')

        # Get delivery method display
        delivery_display = order.delivery_method_display if hasattr(order, 'delivery_method_display') and order.delivery_method else (order.delivery_method or '')

        row_data = [
            order.order_number,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.status_display_name,
            type_name,
            customer_name,
            customer_email,
            customer_phone,
            products_str,
            float(order.total_amount) if order.total_amount else 0.00,
            float(order.shipping_cost) if order.shipping_cost else 0.00,
            float(order.grand_total) if order.grand_total else 0.00,
            float(order.paid_amount) if order.paid_amount else 0.00,
            delivery_display,
            order.payment_method or '',
            order.admin_notes or ''
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

            # Apply currency format to money columns
            if col_num in [9, 10, 11, 12]:
                cell.number_format = '#,##0.00 "PLN"'
                cell.alignment = currency_alignment
            else:
                cell.alignment = data_alignment

            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = alt_row_fill

        # Adjust row height for products column
        if products_str:
            line_count = len(products_str.split('\n'))
            ws.row_dimensions[row_num].height = max(20, min(line_count * 15, 100))

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=zamowienia_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return response


@orders_bp.route('/api/orders/bulk/info', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def bulk_orders_info():
    """
    Get info about selected orders for bulk actions modal.
    Returns order numbers and basic info.
    """
    try:
        data = request.get_json()
        order_ids = data.get('order_ids', [])

        if not order_ids:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano żadnych zamówień'
            }), 400

        orders = Order.query.filter(Order.id.in_(order_ids)).all()

        orders_info = []
        for order in orders:
            customer_name = order.user.full_name if order.user else 'Nieznany'
            orders_info.append({
                'id': order.id,
                'order_number': order.order_number,
                'customer_name': customer_name,
                'status': order.status_display_name,
                'status_color': order.status_badge_color,
                'total': float(order.total_amount) if order.total_amount else 0
            })

        return jsonify({
            'success': True,
            'orders': orders_info,
            'count': len(orders_info)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Błąd: {str(e)}'
        }), 500


# ====================
# API ENDPOINTS
# ====================

@orders_bp.route('/api/orders/detect-courier')
@login_required
@role_required('admin', 'mod')
def api_detect_courier():
    """
    API endpoint for courier auto-detection.
    Returns JSON with courier suggestion.
    """
    tracking_number = request.args.get('tracking', '')

    if not tracking_number:
        return jsonify({'courier': None, 'confidence': 'low', 'url': None})

    result = detect_courier(tracking_number)

    return jsonify(result)


# ====================
# CLIENT ROUTES
# ====================

def apply_payment_status_filter(query, variant):
    """Filtruje zapytanie o zamówienia po stanie opłacenia (wszystkie etapy E1–E4).

    Miarą jest `Order.total_to_pay` — pełna należność klienta obejmująca produkt,
    wysyłkę z Korei, cło/VAT i wysyłkę krajową. `paid_amount` akumuluje sumę
    zatwierdzonych wpłat ze wszystkich etapów, więc porównanie tych dwóch wartości
    odpowiada na pytanie „czy został jeszcze jakiś etap do opłacenia".

    Warianty:
        'paid'    — nic nie zostało do zapłaty (także nadpłata),
        'unpaid'  — cokolwiek zostało do zapłaty (obejmuje częściowo opłacone),
        'partial' — coś już wpłacono, ale nie wszystko.

    Zamówienia anulowane i w zwrocie nie są należnością — klient nie ma czego
    zapłacić, więc wypadają z 'unpaid' i 'partial' (parytet z kafelkiem
    „Do zapłaty" na dashboardzie, patrz modules/client/dashboard_service.py).

    Pusty/nieznany wariant zwraca zapytanie bez zmian.
    """
    from utils.offer_closure import CLOSED_ORDER_STATUSES

    if not variant:
        return query

    paid_amount = db.func.coalesce(Order.paid_amount, 0)

    if variant == 'paid':
        return query.filter(paid_amount >= Order.total_to_pay)
    if variant == 'unpaid':
        return query.filter(
            ~Order.status.in_(CLOSED_ORDER_STATUSES),
            paid_amount < Order.total_to_pay,
        )
    if variant == 'partial':
        return query.filter(
            ~Order.status.in_(CLOSED_ORDER_STATUSES),
            paid_amount > 0,
            paid_amount < Order.total_to_pay,
        )
    return query


@orders_bp.route('/client/orders')
@login_required
def client_list():
    """
    Client order history.
    Shows only orders belonging to current user.
    """
    # Filters
    status_filter = request.args.get('status')
    statuses_filter = request.args.get('statuses', '').strip()  # comma-separated list of statuses
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    search_query = request.args.get('search', '').strip()
    payment_status_filter = request.args.get('payment_status', '').strip()
    # Base query (only user's orders) with eager loading
    query = Order.query.filter_by(user_id=current_user.id).options(
        db.joinedload(Order.items).joinedload(OrderItem.product)
    )

    # Apply filters
    if statuses_filter:
        # Multiple statuses (comma-separated)
        status_list = [s.strip() for s in statuses_filter.split(',') if s.strip()]
        if status_list:
            query = query.filter(Order.status.in_(status_list))
    elif status_filter:
        query = query.filter(Order.status == status_filter)

    if date_from:
        query = query.filter(Order.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))

    if date_to:
        from datetime import timedelta
        end_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Order.created_at < end_date)

    # Search filter (by order number or custom name)
    if search_query:
        query = query.filter(
            or_(
                Order.order_number.ilike(f'%{search_query}%'),
                Order.custom_name.ilike(f'%{search_query}%')
            )
        )

    # Payment status filter
    query = apply_payment_status_filter(query, payment_status_filter)

    # Sorting
    query = query.order_by(Order.created_at.desc())

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Get statuses for filter dropdown
    statuses = OrderStatus.query.filter_by(is_active=True).order_by(OrderStatus.sort_order).all()

    filter_args = {k: v for k, v in request.args.items() if k != 'page'}

    from utils.supplier_order_state import get_supplier_states_for_orders
    supplier_states = get_supplier_states_for_orders(pagination.items)

    return render_template(
        'client/orders/list.html',
        orders=pagination,
        statuses=statuses,
        supplier_states=supplier_states,
        page_title='Moje zamówienia',
        filter_args=filter_args
    )


@orders_bp.route('/client/orders/<int:order_id>')
@login_required
def client_detail(order_id):
    """
    Client order detail page.
    User can only view their own orders.
    """
    order = Order.query.get_or_404(order_id)

    # Security check: User can only view their own orders
    if order.user_id != current_user.id:
        abort(403)

    # Load relationships
    order_items = OrderItem.query.filter_by(order_id=order_id).all()

    # Order history - WSZYSTKIE activity logs dla zamówienia
    from modules.admin.models import ActivityLog

    order_history = []
    activity_logs = ActivityLog.query.filter_by(
        entity_type='order',
        entity_id=order.id
    ).order_by(ActivityLog.created_at.desc()).all()

    for log in activity_logs:
        # Biała lista: klient ogląda tylko akcje, dla których mamy przygotowany
        # polski opis. Nieznana akcja renderowała się dotąd surową nazwą z bazy
        # ('order_status_auto_updated 📝'), więc każda nowa akcja techniczna
        # wyciekała klientowi na stronę w chwili dodania. Admin nadal widzi
        # wszystko — jego oś czasu buduje osobny kod (admin_detail).
        if log.action not in ORDER_ACTION_CONFIG:
            continue

        config = ORDER_ACTION_CONFIG[log.action]

        # Podstawowe dane zdarzenia
        history_item = {
            'created_at': log.created_at,
            'user_name': log.user.full_name if log.user else 'System',
            'action': log.action,
            'action_label': config['label'],
            'action_icon': config['icon'],
            'is_status_change': False,
        }

        # Parse new_value dla wzbogacenia etykiet
        new_value_data = {}
        if log.new_value:
            try:
                new_value_data = json.loads(log.new_value)
            except (json.JSONDecodeError, TypeError):
                pass

        # Obsługa zmian statusu (z kolorowym badge) - ręczne i automatyczne
        if log.action in ('order_status_change', 'order_status_auto_updated'):
            status_slug = new_value_data.get('status')
            status_obj = OrderStatus.query.filter_by(slug=status_slug).first()
            history_item['is_status_change'] = True
            history_item['status_name'] = status_obj.name if status_obj else status_slug
            history_item['status_color'] = status_obj.badge_color if status_obj else '#6B7280'

        # Wzbogacenie etykiet o kontekst z new_value
        elif log.action in ('payment_confirmation_uploaded', 'payment_confirmation_reuploaded',
                            'payment_confirmation_approved', 'payment_confirmation_rejected'):
            stage = new_value_data.get('payment_stage', '')
            amount = new_value_data.get('amount')
            stage_name = PAYMENT_STAGE_LABELS.get(stage, stage)
            extra = []
            if stage_name:
                extra.append(stage_name)
            if amount:
                extra.append(f"{amount} PLN")
            if extra:
                history_item['action_label'] = f"{config['label']} ({' — '.join(extra)})"

        elif log.action in ('proxy_shipping_distributed', 'customs_vat_distributed'):
            amount = new_value_data.get('amount')
            if amount:
                history_item['action_label'] = f"{config['label']} ({amount} PLN)"

        elif log.action == 'offer_closure_fulfillment':
            fulfilled = new_value_data.get('fulfilled_items', 0)
            total = new_value_data.get('total_items', 0)
            new_total = new_value_data.get('new_total_amount')
            parts = []
            if total:
                parts.append(f"{fulfilled}/{total} produktów")
            if new_total is not None:
                parts.append(f"kwota: {new_total} PLN")
            if parts:
                history_item['action_label'] = f"{config['label']} ({', '.join(parts)})"

        order_history.append(history_item)

    # Set probability for live offer sales
    set_probabilities = {}
    has_set_sections = False
    if order.order_type == 'exclusive' and order.offer_page_id:
        page_obj = order.offer_page
        if page_obj:
            from modules.offers.models import OfferSection
            has_set_sections = OfferSection.query.filter_by(
                offer_page_id=order.offer_page_id,
                section_type='set'
            ).first() is not None
            if not page_obj.is_fully_closed:
                from modules.offers.reservation import get_set_probabilities
                set_probabilities = get_set_probabilities(order)

    # --- Tracking map data ---
    tracking_statuses = [
        'dostarczone_proxy', 'w_drodze_polska', 'urzad_celny',
        'dostarczone_gom', 'spakowane', 'wyslane', 'dostarczone'
    ]
    show_tracking_map = order.status in tracking_statuses

    # Also show map for 'oczekujace' if product payment is approved (products ordered)
    tracking_proxy_status = None
    if not show_tracking_map and order.status == 'oczekujace' and order.product_payment_status == 'approved':
        from modules.products.models import ProxyOrderItem, ProxyOrder
        proxy_item = ProxyOrderItem.query.filter_by(order_id=order.id).first()
        if proxy_item:
            proxy_order = proxy_item.proxy_order
            if proxy_order and proxy_order.status in ('zamowiono', 'dostarczone_do_proxy'):
                show_tracking_map = True
                tracking_proxy_status = proxy_order.status

    status_timestamps = {}
    if show_tracking_map:
        from modules.admin.models import ActivityLog
        logs = ActivityLog.query.filter_by(
            entity_type='order',
            entity_id=order.id
        ).filter(
            ActivityLog.action.in_(['order_status_change', 'order_status_auto_updated'])
        ).order_by(ActivityLog.created_at.asc()).all()

        for log in logs:
            try:
                data = json.loads(log.new_value) if log.new_value else {}
                slug = data.get('status')
                if slug:
                    status_timestamps[slug] = log.created_at.strftime('%Y-%m-%dT%H:%M')
            except (ValueError, TypeError):
                pass

    # Shipping city for client marker — client_shipping_request, nie shipping_request:
    # po konsolidacji surowe shipping_request wskazuje paczkę zbiorczą z miastem obcej osoby.
    tracking_shipping_city = ''
    tracking_has_shipping = False
    if order.client_shipping_request:
        tracking_has_shipping = True
        tracking_shipping_city = order.client_shipping_request.shipping_city or ''

    return render_template(
        'client/orders/detail.html',
        order=order,
        order_items=order_items,
        order_history=order_history,
        set_probabilities=set_probabilities,
        has_set_sections=has_set_sections,
        show_tracking_map=show_tracking_map,
        status_timestamps=status_timestamps,
        tracking_shipping_city=tracking_shipping_city,
        tracking_has_shipping=tracking_has_shipping,
        tracking_proxy_status=tracking_proxy_status,
        page_title=f'Zamówienie {order.order_number}'
    )



@orders_bp.route('/client/orders/<int:order_id>/custom-name', methods=['POST'])
@login_required
def update_custom_name(order_id):
    """Update custom name for an order (AJAX)."""
    order = Order.query.get_or_404(order_id)

    if order.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Brak dostępu'}), 403

    data = request.get_json(silent=True) or {}
    custom_name = (data.get('custom_name') or '').strip()

    if len(custom_name) > 50:
        return jsonify({'success': False, 'error': 'Nazwa może mieć maksymalnie 50 znaków'}), 400

    order.custom_name = custom_name if custom_name else None
    db.session.commit()

    return jsonify({'success': True, 'custom_name': order.custom_name})


# ====================
# SETTINGS
# ====================

@orders_bp.route('/admin/orders/settings', methods=['GET'])
@login_required
@role_required('admin')
def settings():
    """
    Orders settings page - manage statuses, WMS statuses, payment methods, and offer closure settings.
    Only accessible to admins.
    """
    from modules.payments.models import PaymentMethod
    from modules.auth.models import Settings

    # Load all order statuses
    statuses = OrderStatus.query.order_by(OrderStatus.sort_order).all()

    # Load all WMS statuses
    wms_statuses = WmsStatus.query.order_by(WmsStatus.sort_order).all()

    # Load all payment methods
    payment_methods = PaymentMethod.query.order_by(PaymentMethod.sort_order, PaymentMethod.name).all()

    # Load offer closure settings
    def get_setting_value(key, default):
        setting = Settings.query.filter_by(key=key).first()
        return setting.value if setting else default

    offer_closure_settings = {
        'fully_fulfilled': get_setting_value('offer_closure_status_fully_fulfilled', 'oczekujace'),
        'partially_fulfilled': get_setting_value('offer_closure_status_partially_fulfilled', 'oczekujace'),
        'not_fulfilled': get_setting_value('offer_closure_status_not_fulfilled', 'anulowane')
    }

    # Load shipping request statuses
    shipping_request_statuses = ShippingRequestStatus.query.order_by(ShippingRequestStatus.sort_order).all()

    # Load shipping request allowed statuses
    shipping_request_allowed_json = get_setting_value('shipping_request_allowed_statuses', '["dostarczone_gom"]')
    try:
        shipping_request_allowed_statuses = json.loads(shipping_request_allowed_json) if shipping_request_allowed_json else []
    except (json.JSONDecodeError, TypeError):
        shipping_request_allowed_statuses = []

    # Load shipping request default status
    shipping_request_default_status = get_setting_value('shipping_request_default_status', '')

    # Load email notifications config
    email_notif_config_json = get_setting_value('email_notifications_config', '{}')
    try:
        email_notif_config = json.loads(email_notif_config_json) if isinstance(email_notif_config_json, str) else (email_notif_config_json or {})
    except (json.JSONDecodeError, TypeError):
        email_notif_config = {}

    # Admin notification recipients config (JSON: {disabled_admin_ids: [], extra_emails: ""})
    admin_notif_recipients_json = get_setting_value('admin_notification_recipients', '{}')
    try:
        admin_notif_recipients = json.loads(admin_notif_recipients_json) if isinstance(admin_notif_recipients_json, str) else (admin_notif_recipients_json or {})
    except (json.JSONDecodeError, TypeError):
        admin_notif_recipients = {}

    # Load all admin users for the email notifications tab
    from modules.auth.models import User as AuthUser
    admin_users = AuthUser.query.filter_by(role='admin', is_active=True).order_by(AuthUser.first_name).all()

    # OCR settings
    ocr_enabled = Settings.get_value('ocr_enabled', False)
    ocr_auto_approve_threshold = Settings.get_value('ocr_auto_approve_threshold', 90)
    ocr_suggest_threshold = Settings.get_value('ocr_suggest_threshold', 60)

    # Konfiguracja potwierdzeń dostawy (przypomnienie + automatyczne domknięcie)
    from modules.orders.delivery_config import pobierz_konfig_dostawy

    return render_template(
        'admin/orders/settings.html',
        statuses=statuses,
        wms_statuses=wms_statuses,
        payment_methods=payment_methods,
        offer_closure_settings=offer_closure_settings,
        shipping_request_statuses=shipping_request_statuses,
        shipping_request_allowed_statuses=shipping_request_allowed_statuses,
        shipping_request_default_status=shipping_request_default_status,
        email_notif_config=email_notif_config,
        admin_notif_recipients=admin_notif_recipients,
        admin_users=admin_users,
        ocr_enabled=ocr_enabled,
        ocr_auto_approve_threshold=ocr_auto_approve_threshold,
        ocr_suggest_threshold=ocr_suggest_threshold,
        maintenance_enabled=Settings.get_value('maintenance_mode', False),
        maintenance_message=Settings.get_value('maintenance_message', ''),
        maintenance_eta=Settings.get_value('maintenance_eta', ''),
        delivery_config=pobierz_konfig_dostawy(),
        page_title='Ustawienia zamówień'
    )


@orders_bp.route('/admin/orders/settings/ocr', methods=['POST'])
@login_required
@role_required('admin')
def update_ocr_settings():
    """Zapisz ustawienia OCR."""
    from modules.auth.models import Settings
    from utils.activity_logger import log_activity

    ocr_enabled = request.form.get('ocr_enabled') == 'on'
    auto_threshold = request.form.get('ocr_auto_approve_threshold', 90, type=int)
    suggest_threshold = request.form.get('ocr_suggest_threshold', 60, type=int)

    # Walidacja
    auto_threshold = max(0, min(100, auto_threshold))
    suggest_threshold = max(0, min(100, suggest_threshold))

    if suggest_threshold >= auto_threshold:
        flash('Próg sugestii musi być niższy niż próg auto-akceptacji.', 'error')
        return redirect(url_for('orders.settings') + '#tab-payment-methods')

    Settings.set_value('ocr_enabled', str(ocr_enabled).lower(), updated_by=current_user.id, type='boolean', description='Włącz/wyłącz OCR')
    Settings.set_value('ocr_auto_approve_threshold', str(auto_threshold), updated_by=current_user.id, type='integer', description='Próg auto-akceptacji OCR')
    Settings.set_value('ocr_suggest_threshold', str(suggest_threshold), updated_by=current_user.id, type='integer', description='Próg sugestii OCR')

    log_activity(
        user=current_user,
        action='settings_updated',
        entity_type='settings',
        new_value={
            'ocr_enabled': ocr_enabled,
            'ocr_auto_approve_threshold': auto_threshold,
            'ocr_suggest_threshold': suggest_threshold
        }
    )

    flash('Ustawienia OCR zostały zapisane.', 'success')
    return redirect(url_for('orders.settings') + '#tab-payment-methods')


@orders_bp.route('/admin/orders/settings/offer-closure', methods=['POST'])
@login_required
@role_required('admin')
def update_offer_closure_settings():
    """
    Update offer closure settings - configure automatic status changes after offer page closure.
    Only accessible to admins.
    """
    from modules.auth.models import Settings

    try:
        # Get form data
        status_fully = request.form.get('offer_closure_status_fully_fulfilled', '').strip()
        status_partially = request.form.get('offer_closure_status_partially_fulfilled', '').strip()
        status_not = request.form.get('offer_closure_status_not_fulfilled', '').strip()

        # Validate required fields
        if not status_fully or not status_partially or not status_not:
            flash('Wszystkie pola są wymagane', 'error')
            return redirect(url_for('orders.settings'))

        # Validate that statuses exist
        valid_statuses = [s.slug for s in OrderStatus.query.filter_by(is_active=True).all()]

        if status_fully not in valid_statuses:
            flash(f'Status "{status_fully}" nie istnieje lub jest nieaktywny', 'error')
            return redirect(url_for('orders.settings'))

        if status_partially not in valid_statuses:
            flash(f'Status "{status_partially}" nie istnieje lub jest nieaktywny', 'error')
            return redirect(url_for('orders.settings'))

        if status_not not in valid_statuses:
            flash(f'Status "{status_not}" nie istnieje lub jest nieaktywny', 'error')
            return redirect(url_for('orders.settings'))

        # Update or create settings
        def update_or_create_setting(key, value):
            setting = Settings.query.filter_by(key=key).first()
            if setting:
                setting.value = value
                setting.updated_at = datetime.now()
            else:
                # Create new setting if it doesn't exist
                setting = Settings(
                    key=key,
                    value=value,
                    type='string',
                    description=f'Auto-generated setting for {key}'
                )
                db.session.add(setting)

        update_or_create_setting('offer_closure_status_fully_fulfilled', status_fully)
        update_or_create_setting('offer_closure_status_partially_fulfilled', status_partially)
        update_or_create_setting('offer_closure_status_not_fulfilled', status_not)

        db.session.commit()

        # Activity log
        from utils.activity_logger import log_activity
        log_activity(
            user=current_user,
            action='settings_updated',
            entity_type='settings',
            entity_id=None,
            old_value=None,
            new_value={
                'fully_fulfilled': status_fully,
                'partially_fulfilled': status_partially,
                'not_fulfilled': status_not
            }
        )

        flash('Ustawienia zostały zapisane', 'success')
        return redirect(url_for('orders.settings') + '#tab-offer-closure')

    except Exception as e:
        db.session.rollback()
        flash(f'Błąd podczas zapisywania ustawień: {str(e)}', 'error')
        return redirect(url_for('orders.settings'))


# ============================================
# EMAIL NOTIFICATIONS SETTINGS
# ============================================

@orders_bp.route('/admin/orders/settings/email-notifications', methods=['POST'])
@login_required
@role_required('admin')
def update_email_notification_settings():
    """
    Update email notification settings - toggle on/off for each notification type
    and configure admin notification email addresses.
    """
    from modules.auth.models import Settings
    from utils.email_manager import EmailManager

    ALLOWED_KEYS = {
        'notify_order_confirmation', 'notify_status_change', 'notify_order_completed',
        'notify_tracking_added', 'notify_packing_photo', 'notify_order_cancelled',
        'notify_supplier_ordered', 'notify_supplier_cancelled',
        'notify_cost_added', 'notify_payment_approved', 'notify_payment_rejected',
        'notify_payment_reminder', 'notify_shipping_request_created',
        'notify_shipping_status_change', 'notify_offer_closure',
        'notify_new_offer_page', 'notify_back_in_stock',
        'notify_admin_new_order', 'notify_admin_payment_uploaded',
        'notify_delivery_confirmation', 'notify_delivery_confirmed',
        'notify_delivery_autoclosed', 'notify_admin_delivery_confirmed',
    }

    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Brak danych'}), 400

        # Build config dict from toggles (whitelist only allowed keys)
        toggles = data.get('toggles', {})
        config = {}
        for key in ALLOWED_KEYS:
            config[key] = bool(toggles.get(key, True))

        # Save email notifications config as JSON
        Settings.set_value(
            'email_notifications_config',
            json.dumps(config),
            updated_by=current_user.id,
            type='json',
            description='Email notification toggles (on/off per notification type)'
        )

        # Save admin notification recipients config
        recipients = {
            'disabled_admin_ids': [int(x) for x in data.get('disabled_admin_ids', [])],
            'extra_emails': data.get('extra_emails', '').strip(),
        }
        Settings.set_value(
            'admin_notification_recipients',
            json.dumps(recipients),
            updated_by=current_user.id,
            type='json',
            description='Admin notification recipients (disabled admins + extra emails)'
        )

        db.session.commit()

        # Clear cache so changes take effect immediately in this worker
        EmailManager.clear_email_config_cache()

        return jsonify({'success': True, 'message': 'Ustawienia powiadomień email zostały zapisane'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error saving email notification settings: {e}")
        return jsonify({'success': False, 'message': f'Błąd: {str(e)}'}), 500


# ============================================
# SHIPPING REQUEST SETTINGS
# ============================================

@orders_bp.route('/admin/orders/update-shipping-request-allowed-statuses', methods=['POST'])
@login_required
@role_required('admin')
def update_shipping_request_allowed_statuses():
    """
    Update list of order statuses that qualify for shipping request.
    Only accessible to admins.
    """
    from modules.auth.models import Settings

    try:
        # Get selected statuses (list of slugs)
        allowed_statuses = request.form.getlist('allowed_statuses')

        # Validate that all selected statuses exist
        valid_statuses = [s.slug for s in OrderStatus.query.filter_by(is_active=True).all()]

        # Filter only valid statuses
        validated_statuses = [s for s in allowed_statuses if s in valid_statuses]

        # Update or create setting
        setting = Settings.query.filter_by(key='shipping_request_allowed_statuses').first()
        if setting:
            setting.value = json.dumps(validated_statuses)
            setting.type = 'json'
            setting.updated_at = datetime.now()
        else:
            setting = Settings(
                key='shipping_request_allowed_statuses',
                value=json.dumps(validated_statuses),
                type='json',
                description='Lista statusów zamówień kwalifikujących się do zlecenia wysyłki'
            )
            db.session.add(setting)

        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='settings_updated',
            entity_type='settings',
            entity_id=None,
            old_value=None,
            new_value={'shipping_request_allowed_statuses': validated_statuses}
        )

        flash('Ustawienia zostały zapisane', 'success')
        return redirect(url_for('orders.settings') + '#tab-shipping-requests')

    except Exception as e:
        db.session.rollback()
        flash(f'Błąd podczas zapisywania ustawień: {str(e)}', 'error')
        return redirect(url_for('orders.settings') + '#tab-shipping-requests')


@orders_bp.route('/admin/orders/settings/shipping-request-default-status', methods=['POST'])
@login_required
@role_required('admin')
def update_shipping_request_default_status():
    """
    Update default status for new shipping requests.
    Only accessible to admins.
    """
    from modules.auth.models import Settings

    try:
        # Get selected status
        default_status = request.form.get('default_status', '').strip()

        # Validate that status exists
        if default_status:
            valid_statuses = [s.slug for s in ShippingRequestStatus.query.filter_by(is_active=True).all()]
            if default_status not in valid_statuses:
                flash('Wybrany status nie istnieje lub jest nieaktywny', 'error')
                return redirect(url_for('orders.settings') + '#tab-shipping-requests')

        # Update or create setting
        setting = Settings.query.filter_by(key='shipping_request_default_status').first()
        if setting:
            setting.value = default_status
            setting.updated_at = datetime.now()
        else:
            setting = Settings(
                key='shipping_request_default_status',
                value=default_status,
                type='string',
                description='Domyślny status dla nowych zleceń wysyłki'
            )
            db.session.add(setting)

        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='settings_updated',
            entity_type='settings',
            entity_id=None,
            old_value=None,
            new_value={'shipping_request_default_status': default_status}
        )

        flash('Ustawienia zostały zapisane', 'success')
        return redirect(url_for('orders.settings') + '#tab-shipping-requests')

    except Exception as e:
        db.session.rollback()
        flash(f'Błąd podczas zapisywania ustawień: {str(e)}', 'error')
        return redirect(url_for('orders.settings') + '#tab-shipping-requests')


# Etykiety pól konfiguracji dostawy do komunikatów walidacji — 1:1 z podpisami
# w templates/admin/orders/settings.html (sekcja „Potwierdzenie dostawy"). Admin
# widzi w formularzu opis pola, nie jego klucz z KLUCZE (modules/orders/delivery_config.py),
# więc komunikat błędu ma mówić tym samym językiem.
ETYKIETY_KONFIGURACJI_DOSTAWY = {
    'reminder_enabled': 'Wysyłaj przypomnienie o potwierdzeniu odbioru',
    'reminder_days': 'Przypomnienie po (dni od wysyłki)',
    'autocomplete_enabled': 'Domykaj niepotwierdzone zlecenia automatycznie',
    'autocomplete_days': 'Automatyczne domknięcie po (dni od wysyłki)',
    'autocomplete_batch': 'Maksymalnie zleceń domykanych na jeden przebieg',
    'review_window_days': 'Ocenę można wystawić przez (dni od dostarczenia)',
}


@orders_bp.route('/admin/settings/delivery', methods=['POST'])
@login_required
@role_required('admin')
def admin_update_delivery_settings():
    """Zapis ustawień potwierdzania dostawy (task 869efhwph).

    Settings.set_value() commituje wewnętrznie przy KAŻDYM wywołaniu (patrz
    modules/auth/models.py) — nie jest częścią jednej transakcji z resztą pętli.
    Dlatego zapis idzie dwuprzebiegowo: najpierw walidujemy komplet pól bez
    ani jednego zapisu do bazy, dopiero gdy cały formularz jest poprawny —
    zapisujemy. Inaczej błąd na którymś z kolejnych pól zostawiałby w bazie
    część już zapisanej, niekompletnej konfiguracji, mimo odpowiedzi 400.
    """
    from modules.auth.models import Settings
    from modules.orders.delivery_config import DOMYSLNE, KLUCZE, pobierz_konfig_dostawy

    dane = request.get_json() or {}

    # Przebieg 1: walidacja całego formularza — zero zapisów do bazy.
    do_zapisu = []
    for pole, klucz in KLUCZE.items():
        if pole not in dane:
            continue
        etykieta = ETYKIETY_KONFIGURACJI_DOSTAWY.get(pole, pole)
        if isinstance(DOMYSLNE[pole], bool):
            wartosc = dane[pole]
            # bool("false") == True w Pythonie (niepusty string jest prawdziwy) —
            # bez jawnego sprawdzenia typu endpoint przyjąłby dowolny śmieciowy
            # JSON jako włączenie przełącznika. Nieosiągalne z checkboksa w UI,
            # ale to publiczne API admina, więc odrzucamy 400 zamiast zgadywać
            # intencję.
            #
            # Tak, jest to SUROWSZE niż gałąź liczbowa niżej, która przez int()
            # przyjmuje też stringi („5"), i surowsze niż warstwa odczytu
            # (_jako_bool w delivery_config.py, która parsuje „true"/„on"/„1").
            # Ta asymetria jest zamierzona i nie ma jej po co „ujednolicać":
            #  * _jako_bool MUSI umieć stringi, bo tabela settings trzyma
            #    wartości jako tekst — czyta z bazy, nie z JSON-a;
            #  * „5" ma dokładnie jedno sensowne odczytanie jako liczba, więc
            #    int() niczego nie zgaduje;
            #  * dowolny string jako bool sensownego odczytania NIE ma — „false"
            #    czyta się dla człowieka jako fałsz, a dla Pythona jako prawda,
            #    i to był właśnie ten błąd.
            # JSON ma natywny typ logiczny, więc na wejściu możemy go wymagać.
            if not isinstance(wartosc, bool):
                return jsonify({
                    'success': False,
                    'message': f'Pole „{etykieta}" musi być wartością logiczną (prawda/fałsz)'
                }), 400
            do_zapisu.append((klucz, wartosc, 'boolean'))
        else:
            try:
                liczba = int(dane[pole])
            except (TypeError, ValueError):
                return jsonify({
                    'success': False,
                    'message': f'Pole „{etykieta}" musi być liczbą całkowitą'
                }), 400
            if liczba < 1:
                return jsonify({
                    'success': False,
                    'message': f'Pole „{etykieta}" musi wynosić co najmniej 1'
                }), 400
            do_zapisu.append((klucz, liczba, 'integer'))

    # Przebieg 2: komplet danych poprawny — dopiero teraz zapisujemy.
    for klucz, wartosc, typ in do_zapisu:
        Settings.set_value(klucz, wartosc, updated_by=current_user.id, type=typ)

    db.session.commit()
    return jsonify({'success': True, 'config': pobierz_konfig_dostawy()})


@orders_bp.route('/admin/orders/shipping-request-statuses/<int:status_id>', methods=['GET'])
@login_required
@role_required('admin')
def get_shipping_request_status(status_id):
    """Get shipping request status data for edit modal."""
    status = ShippingRequestStatus.query.get_or_404(status_id)
    return jsonify({
        'id': status.id,
        'name': status.name,
        'slug': status.slug,
        'badge_color': status.badge_color,
        'is_initial': status.is_initial,
        'is_active': status.is_active
    })


@orders_bp.route('/admin/orders/shipping-request-statuses/create', methods=['POST'])
@login_required
@role_required('admin')
def create_shipping_request_status():
    """Create new shipping request status."""
    from modules.orders.utils import generate_slug

    try:
        data = request.get_json()
        name = data.get('name', '').strip()

        if not name:
            return jsonify({'success': False, 'error': 'Nazwa jest wymagana'}), 400

        # Generate slug from name
        slug = generate_slug(name)

        # Check if slug already exists
        existing = ShippingRequestStatus.query.filter_by(slug=slug).first()
        if existing:
            return jsonify({'success': False, 'error': 'Status o takiej nazwie już istnieje'}), 400

        # Get max sort_order
        max_order = db.session.query(func.max(ShippingRequestStatus.sort_order)).scalar() or 0

        # Create new status
        status = ShippingRequestStatus(
            slug=slug,
            name=name,
            badge_color=data.get('badge_color', '#6B7280'),
            is_initial=data.get('is_initial', False),
            is_active=data.get('is_active', True),
            sort_order=max_order + 1
        )
        db.session.add(status)
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='shipping_request_status_created',
            entity_type='shipping_request_status',
            entity_id=status.id,
            old_value=None,
            new_value={'name': name, 'slug': slug}
        )

        return jsonify({'success': True, 'id': status.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/shipping-request-statuses/<int:status_id>', methods=['PUT'])
@login_required
@role_required('admin')
def update_shipping_request_status(status_id):
    """Update shipping request status."""
    status = ShippingRequestStatus.query.get_or_404(status_id)

    try:
        data = request.get_json()
        name = data.get('name', '').strip()

        if not name:
            return jsonify({'success': False, 'error': 'Nazwa jest wymagana'}), 400

        old_values = {
            'name': status.name,
            'badge_color': status.badge_color,
            'is_initial': status.is_initial,
            'is_active': status.is_active
        }

        status.name = name
        status.badge_color = data.get('badge_color', status.badge_color)
        status.is_initial = data.get('is_initial', status.is_initial)
        status.is_active = data.get('is_active', status.is_active)
        status.updated_at = datetime.now()

        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='shipping_request_status_updated',
            entity_type='shipping_request_status',
            entity_id=status.id,
            old_value=old_values,
            new_value={
                'name': status.name,
                'badge_color': status.badge_color,
                'is_initial': status.is_initial,
                'is_active': status.is_active
            }
        )

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/shipping-request-statuses/<int:status_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def delete_shipping_request_status(status_id):
    """Delete shipping request status."""
    status = ShippingRequestStatus.query.get_or_404(status_id)

    try:
        # Check if status is in use
        in_use_count = ShippingRequest.query.filter_by(status=status.slug).count()
        if in_use_count > 0:
            return jsonify({
                'success': False,
                'error': f'Nie można usunąć statusu - jest używany w {in_use_count} zleceniach'
            }), 400

        status_name = status.name

        db.session.delete(status)
        db.session.commit()

        # Activity log
        log_activity(
            user=current_user,
            action='shipping_request_status_deleted',
            entity_type='shipping_request_status',
            entity_id=status_id,
            old_value={'name': status_name},
            new_value=None
        )

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API ENDPOINTS (for modals)
# ============================================

@orders_bp.route('/api/orders/statuses/<int:status_id>')
@login_required
@role_required('admin')
def api_get_status(status_id):
    """Get status data for edit modal."""
    status = OrderStatus.query.get_or_404(status_id)
    return jsonify({
        'id': status.id,
        'name': status.name,
        'slug': status.slug,
        'badge_color': status.badge_color,
        'is_active': status.is_active
    })


@orders_bp.route('/admin/orders/statuses/create', methods=['POST'])
@login_required
@role_required('admin')
def create_status():
    """Create new order status."""
    from modules.orders.utils import generate_slug

    name = request.form.get('name', '').strip()
    badge_color = request.form.get('badge_color', '#6B7280')
    is_active = request.form.get('is_active') == 'on'

    # Validation
    errors = {}
    if not name:
        errors['name'] = 'Nazwa statusu jest wymagana'

    # Generate slug from name
    slug = generate_slug(name)

    # Check if slug already exists
    existing = OrderStatus.query.filter_by(slug=slug).first()
    if existing:
        errors['name'] = f'Status o nazwie "{name}" już istnieje (slug: {slug})'

    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    # Create new status
    status = OrderStatus(
        name=name,
        slug=slug,
        badge_color=badge_color,
        is_active=is_active,
        sort_order=OrderStatus.query.count()  # Add at the end
    )

    db.session.add(status)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Status "{name}" został utworzony',
        'status_id': status.id
    })


@orders_bp.route('/admin/orders/statuses/<int:status_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def edit_status(status_id):
    """Edit existing order status."""
    from modules.orders.utils import generate_slug

    status = OrderStatus.query.get_or_404(status_id)

    name = request.form.get('name', '').strip()
    badge_color = request.form.get('badge_color', '#6B7280')
    is_active = request.form.get('is_active') == 'on'

    # Validation
    errors = {}
    if not name:
        errors['name'] = 'Nazwa statusu jest wymagana'

    # Generate new slug from name
    new_slug = generate_slug(name)

    # Check if slug already exists (but not for this status)
    existing = OrderStatus.query.filter(
        OrderStatus.slug == new_slug,
        OrderStatus.id != status_id
    ).first()
    if existing:
        errors['name'] = f'Status o nazwie "{name}" już istnieje (slug: {new_slug})'

    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    # Update status
    status.name = name
    status.slug = new_slug
    status.badge_color = badge_color
    status.is_active = is_active

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Status "{name}" został zaktualizowany'
    })


@orders_bp.route('/admin/orders/statuses/<int:status_id>/check-usage')
@login_required
@role_required('admin')
def check_status_usage(status_id):
    """Check if status is used in any orders before deletion."""
    status = OrderStatus.query.get_or_404(status_id)
    orders_count = Order.query.filter_by(status=status.slug).count()

    # Get other available statuses for migration
    other_statuses = OrderStatus.query.filter(
        OrderStatus.id != status_id,
        OrderStatus.is_active == True
    ).order_by(OrderStatus.sort_order).all()

    return jsonify({
        'status_id': status.id,
        'status_name': status.name,
        'orders_count': orders_count,
        'can_delete_directly': orders_count == 0,
        'available_statuses': [
            {'id': s.id, 'slug': s.slug, 'name': s.name, 'badge_color': s.badge_color}
            for s in other_statuses
        ]
    })


@orders_bp.route('/admin/orders/statuses/<int:status_id>/migrate', methods=['POST'])
@login_required
@role_required('admin')
def migrate_status(status_id):
    """Migrate orders from one status to another and delete the old status."""
    try:
        status = OrderStatus.query.get_or_404(status_id)
        data = request.get_json()
        new_status_slug = data.get('new_status')

        if not new_status_slug:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano statusu zastępczego'
            }), 400

        # Verify new status exists
        new_status = OrderStatus.query.filter_by(slug=new_status_slug).first()
        if not new_status:
            return jsonify({
                'success': False,
                'message': 'Wybrany status zastępczy nie istnieje'
            }), 400

        # Zapisz ID zamówień PRZED masowym .update() — ten bypassuje ORM
        # (synchronize_session=False), więc żadne pojedyncze zamówienie nie
        # przejdzie przez zwykłą ścieżkę zmiany statusu. Jeśli status zastępczy
        # nie dojdzie już do wysyłki ('anulowane'/'do_zwrotu'), musimy zamówienia
        # dociągnąć i wypiąć z paczek zbiorczych ręcznie.
        from modules.orders.consolidation import STATUSY_WYPINAJACE_Z_PACZKI
        anulowane_ids = []
        if new_status_slug in STATUSY_WYPINAJACE_Z_PACZKI:
            anulowane_ids = [
                oid for (oid,) in db.session.query(Order.id).filter_by(status=status.slug).all()
            ]

        # Migrate all orders to new status
        orders_updated = Order.query.filter_by(status=status.slug).update(
            {'status': new_status_slug},
            synchronize_session=False
        )

        if anulowane_ids:
            from modules.orders.consolidation import odepnij_anulowane_zamowienie
            for oid in anulowane_ids:
                zamowienie = db.session.get(Order, oid)
                if zamowienie:
                    odepnij_anulowane_zamowienie(zamowienie)

        # Delete old status
        db.session.delete(status)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Przeniesiono {orders_updated} zamówień na status "{new_status.name}" i usunięto status "{status.name}"'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas migracji: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/statuses/<int:status_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def delete_status(status_id):
    """Delete order status (only if not used)."""
    try:
        status = OrderStatus.query.get_or_404(status_id)

        # Check if status is used in any orders
        orders_count = Order.query.filter_by(status=status.slug).count()
        if orders_count > 0:
            return jsonify({
                'success': False,
                'requires_migration': True,
                'orders_count': orders_count,
                'message': f'Status jest używany w {orders_count} zamówieniach. Wybierz status zastępczy.'
            }), 400

        db.session.delete(status)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Status "{status.name}" został usunięty'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas usuwania statusu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/statuses/reorder', methods=['POST'])
@login_required
@role_required('admin')
def reorder_statuses():
    """Update sort_order for statuses based on drag & drop."""
    try:
        data = request.get_json()
        statuses = data.get('statuses', [])

        if not statuses:
            return jsonify({
                'success': False,
                'message': 'Brak danych do aktualizacji'
            }), 400

        # Update sort_order for each status
        for status_data in statuses:
            status_id = status_data.get('id')
            sort_order = status_data.get('sort_order')

            if status_id is None or sort_order is None:
                continue

            status = db.session.get(OrderStatus, status_id)
            if status:
                status.sort_order = sort_order

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Kolejność statusów została zaktualizowana'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas aktualizacji kolejności: {str(e)}'
        }), 500


# ============================================
# WMS STATUSES API ENDPOINTS
# ============================================

@orders_bp.route('/api/orders/wms-statuses/<int:status_id>')
@login_required
@role_required('admin')
def api_get_wms_status(status_id):
    """Get WMS status data for edit modal."""
    status = WmsStatus.query.get_or_404(status_id)
    return jsonify({
        'id': status.id,
        'name': status.name,
        'slug': status.slug,
        'badge_color': status.badge_color,
        'is_active': status.is_active,
        'is_default': status.is_default,
        'is_picked': status.is_picked
    })


@orders_bp.route('/admin/orders/wms-statuses/create', methods=['POST'])
@login_required
@role_required('admin')
def create_wms_status():
    """Create new WMS status."""
    from modules.orders.utils import generate_slug

    name = request.form.get('name', '').strip()
    badge_color = request.form.get('badge_color', '#6B7280')
    is_active = request.form.get('is_active') == 'on'
    is_default = request.form.get('is_default') == 'on'
    is_picked = request.form.get('is_picked') == 'on'

    # Validation
    errors = {}
    if not name:
        errors['name'] = 'Nazwa statusu jest wymagana'

    # Generate slug from name
    slug = generate_slug(name)

    # Check if slug already exists
    existing = WmsStatus.query.filter_by(slug=slug).first()
    if existing:
        errors['name'] = f'Status WMS o nazwie "{name}" już istnieje'

    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    # If this is set as default, unset other defaults
    if is_default:
        WmsStatus.query.filter_by(is_default=True).update({'is_default': False})

    # Create new status
    status = WmsStatus(
        name=name,
        slug=slug,
        badge_color=badge_color,
        is_active=is_active,
        is_default=is_default,
        is_picked=is_picked,
        sort_order=WmsStatus.query.count()
    )

    db.session.add(status)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Status WMS "{name}" został utworzony',
        'status': {
            'id': status.id,
            'name': status.name,
            'slug': status.slug,
            'badge_color': status.badge_color,
            'is_active': status.is_active,
            'is_default': status.is_default,
            'is_picked': status.is_picked
        }
    })


@orders_bp.route('/admin/orders/wms-statuses/<int:status_id>/update', methods=['POST'])
@login_required
@role_required('admin')
def update_wms_status(status_id):
    """Update existing WMS status."""
    from modules.orders.utils import generate_slug

    status = WmsStatus.query.get_or_404(status_id)

    name = request.form.get('name', '').strip()
    badge_color = request.form.get('badge_color', '#6B7280')
    is_active = request.form.get('is_active') == 'on'
    is_default = request.form.get('is_default') == 'on'
    is_picked = request.form.get('is_picked') == 'on'

    # Validation
    errors = {}
    if not name:
        errors['name'] = 'Nazwa statusu jest wymagana'

    # Generate new slug if name changed
    new_slug = generate_slug(name)
    if new_slug != status.slug:
        existing = WmsStatus.query.filter_by(slug=new_slug).first()
        if existing:
            errors['name'] = f'Status WMS o nazwie "{name}" już istnieje'

    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    # If this is set as default, unset other defaults
    if is_default and not status.is_default:
        WmsStatus.query.filter(WmsStatus.id != status_id).filter_by(is_default=True).update({'is_default': False})

    # Update status
    status.name = name
    status.slug = new_slug
    status.badge_color = badge_color
    status.is_active = is_active
    status.is_default = is_default
    status.is_picked = is_picked

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Status WMS "{name}" został zaktualizowany',
        'status': {
            'id': status.id,
            'name': status.name,
            'slug': status.slug,
            'badge_color': status.badge_color,
            'is_active': status.is_active,
            'is_default': status.is_default,
            'is_picked': status.is_picked
        }
    })


@orders_bp.route('/admin/orders/wms-statuses/<int:status_id>/check-usage')
@login_required
@role_required('admin')
def check_wms_status_usage(status_id):
    """Check if WMS status is used in any order items before deletion."""
    status = WmsStatus.query.get_or_404(status_id)
    items_count = OrderItem.query.filter_by(wms_status=status.slug).count()

    # Get other available WMS statuses for migration
    other_statuses = WmsStatus.query.filter(
        WmsStatus.id != status_id,
        WmsStatus.is_active == True
    ).order_by(WmsStatus.sort_order).all()

    return jsonify({
        'status_id': status.id,
        'status_name': status.name,
        'items_count': items_count,
        'can_delete_directly': items_count == 0,
        'available_statuses': [
            {'id': s.id, 'slug': s.slug, 'name': s.name, 'badge_color': s.badge_color}
            for s in other_statuses
        ]
    })


@orders_bp.route('/admin/orders/wms-statuses/<int:status_id>/migrate', methods=['POST'])
@login_required
@role_required('admin')
def migrate_wms_status(status_id):
    """Migrate order items from one WMS status to another and delete the old status."""
    try:
        status = WmsStatus.query.get_or_404(status_id)
        data = request.get_json()
        new_status_slug = data.get('new_status')

        if not new_status_slug:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano statusu zastępczego'
            }), 400

        # Verify new status exists
        new_status = WmsStatus.query.filter_by(slug=new_status_slug).first()
        if not new_status:
            return jsonify({
                'success': False,
                'message': 'Wybrany status zastępczy nie istnieje'
            }), 400

        # Migrate all order items to new status
        items_updated = OrderItem.query.filter_by(wms_status=status.slug).update(
            {'wms_status': new_status_slug},
            synchronize_session=False
        )

        # Delete old status
        db.session.delete(status)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Przeniesiono {items_updated} pozycji na status "{new_status.name}" i usunięto status "{status.name}"'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas migracji: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/wms-statuses/<int:status_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def delete_wms_status(status_id):
    """Delete WMS status (only if not used)."""
    try:
        status = WmsStatus.query.get_or_404(status_id)

        # Check if status is in use
        items_count = OrderItem.query.filter_by(wms_status=status.slug).count()
        if items_count > 0:
            return jsonify({
                'success': False,
                'requires_migration': True,
                'items_count': items_count,
                'message': f'Status jest używany w {items_count} pozycjach zamówień. Wybierz status zastępczy.'
            }), 400

        db.session.delete(status)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Status WMS "{status.name}" został usunięty'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas usuwania statusu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/wms-statuses/reorder', methods=['POST'])
@login_required
@role_required('admin')
def reorder_wms_statuses():
    """Update sort_order for WMS statuses based on drag & drop."""
    try:
        data = request.get_json()
        statuses = data.get('statuses', [])

        if not statuses:
            return jsonify({
                'success': False,
                'message': 'Brak danych do aktualizacji'
            }), 400

        for status_data in statuses:
            status_id = status_data.get('id')
            sort_order = status_data.get('sort_order')

            if status_id is None or sort_order is None:
                continue

            status = db.session.get(WmsStatus, status_id)
            if status:
                status.sort_order = sort_order

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Kolejność statusów WMS została zaktualizowana'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas aktualizacji kolejności: {str(e)}'
        }), 500


# ============================================
# API - Pobieranie produktów do modala
# ============================================

@orders_bp.route('/admin/orders/api/products')
@login_required
@role_required('admin', 'mod')
def api_get_products():
    """API endpoint do pobierania produktów dla modala dodawania"""
    from modules.products.models import Product, Category, ProductSeries, ProductImage

    search = request.args.get('search', '').strip()
    category_id = request.args.get('category_id', type=int)
    series_id = request.args.get('series_id', type=int)

    query = Product.query.filter(Product.is_active == True)

    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                Product.name.ilike(search_term),
                Product.sku.ilike(search_term),
                Product.ean.ilike(search_term)
            )
        )

    if category_id:
        query = query.filter(Product.category_id == category_id)

    if series_id:
        query = query.filter(Product.series_id == series_id)

    products = query.order_by(Product.name).limit(50).all()

    products_data = []
    for product in products:
        primary_image = ProductImage.query.filter_by(
            product_id=product.id,
            is_primary=True
        ).first()

        if not primary_image:
            primary_image = ProductImage.query.filter_by(product_id=product.id).first()

        if primary_image:
            image_url = f'/static/uploads/products/compressed/{primary_image.filename}'
        else:
            image_url = '/static/img/placeholders/product.svg'

        category_name = product.category.name if product.category else None
        series_name = product.series.name if product.series else None

        products_data.append({
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'ean': product.ean,
            'sale_price': float(product.sale_price) if product.sale_price else 0,
            'quantity': product.quantity or 0,
            'category_name': category_name,
            'series_name': series_name,
            'image_url': image_url
        })

    return jsonify({
        'success': True,
        'products': products_data
    })


@orders_bp.route('/admin/orders/<int:order_id>/add-products', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_add_products(order_id):
    """Dodaje produkty do istniejącego zamówienia"""
    from modules.products.models import Product

    order = Order.query.get_or_404(order_id)

    try:
        data = request.get_json()
        products_to_add = data.get('products', [])

        if not products_to_add:
            return jsonify({
                'success': False,
                'message': 'Nie wybrano żadnych produktów'
            }), 400

        added_count = 0

        for item in products_to_add:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 1)

            if not product_id or quantity < 1:
                continue

            product = db.session.get(Product, product_id)
            if not product or not product.is_active:
                continue

            # Sprawdź czy produkt już jest w zamówieniu
            existing_item = OrderItem.query.filter_by(
                order_id=order_id,
                product_id=product_id
            ).first()

            if existing_item:
                # Dodaj do istniejącej pozycji
                existing_item.quantity += quantity
                existing_item.total = existing_item.quantity * existing_item.price
            else:
                # Dodaj nową pozycję
                new_item = OrderItem(
                    order_id=order_id,
                    product_id=product_id,
                    quantity=quantity,
                    price=product.sale_price or 0,
                    total=(product.sale_price or 0) * quantity
                )
                db.session.add(new_item)

            added_count += 1

        # Flush żeby nowe items były widoczne w relacji
        db.session.flush()

        # Odśwież order z bazy (aby items były aktualne)
        db.session.refresh(order)

        # Przelicz sumę zamówienia
        order.recalculate_total()

        db.session.commit()

        # Log aktywności
        log_activity(
            user=current_user,
            action='order_products_added',
            entity_type='order',
            entity_id=order.id,
            new_value={'added_products': added_count, 'order_number': order.order_number}
        )

        return jsonify({
            'success': True,
            'message': f'Dodano {added_count} produkt(ów) do zamówienia',
            'new_total': float(order.total_amount) if order.total_amount else 0
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas dodawania produktów: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/items/<int:item_id>', methods=['PUT'])
@login_required
@role_required('admin', 'mod')
def admin_update_item(order_id, item_id):
    """Aktualizuje produkt w zamówieniu (ilość, cena)"""
    order = Order.query.get_or_404(order_id)
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()

    try:
        data = request.get_json()
        quantity = data.get('quantity')
        price = data.get('price')

        if quantity is None or price is None:
            return jsonify({
                'success': False,
                'message': 'Wymagane pola: quantity, price'
            }), 400

        quantity = int(quantity)
        price = float(price)

        if quantity < 1:
            return jsonify({
                'success': False,
                'message': 'Ilość musi być większa niż 0'
            }), 400

        if price < 0:
            return jsonify({
                'success': False,
                'message': 'Cena nie może być ujemna'
            }), 400

        # Zapisz stare wartości do logu
        old_quantity = item.quantity
        old_price = float(item.price)

        # Aktualizuj item
        item.quantity = quantity
        item.price = price
        item.total = quantity * price

        # Przelicz sumę zamówienia
        db.session.flush()
        db.session.refresh(order)
        order.recalculate_total()

        db.session.commit()

        # Log aktywności
        log_activity(
            user=current_user,
            action='order_item_updated',
            entity_type='order',
            entity_id=order.id,
            old_value={
                'product_name': item.product_name,
                'quantity': old_quantity,
                'price': old_price
            },
            new_value={
                'product_name': item.product_name,
                'quantity': quantity,
                'price': price
            }
        )

        return jsonify({
            'success': True,
            'message': 'Produkt został zaktualizowany',
            'new_total': float(order.total_amount) if order.total_amount else 0,
            'item_total': float(item.total)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas aktualizacji produktu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/items/<int:item_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'mod')
def admin_delete_item(order_id, item_id):
    """Usuwa produkt z zamówienia"""
    order = Order.query.get_or_404(order_id)
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()

    try:
        # Zapisz dane do logu przed usunięciem
        product_name = item.product_name
        quantity = item.quantity
        price = float(item.price)

        # Usuń item
        db.session.delete(item)

        # Przelicz sumę zamówienia
        db.session.flush()
        db.session.refresh(order)
        order.recalculate_total()

        db.session.commit()

        # Log aktywności
        log_activity(
            user=current_user,
            action='order_item_deleted',
            entity_type='order',
            entity_id=order.id,
            old_value={
                'product_name': product_name,
                'quantity': quantity,
                'price': price
            }
        )

        return jsonify({
            'success': True,
            'message': f'Produkt "{product_name}" został usunięty',
            'new_total': float(order.total_amount) if order.total_amount else 0
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas usuwania produktu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/add-custom-product', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_add_custom_product(order_id):
    """
    Dodaje ręcznie wpisany produkt do zamówienia (bez product_id).
    Używane dla pełnych setów i innych custom produktów.
    """
    from decimal import Decimal

    order = Order.query.get_or_404(order_id)
    data = request.get_json()

    custom_name = (data.get('custom_name') or '').strip()
    custom_sku = (data.get('custom_sku') or '').strip() or None
    quantity = data.get('quantity', 0)
    price = data.get('price', 0)

    # Walidacja
    if not custom_name:
        return jsonify({'success': False, 'message': 'Podaj nazwę produktu'}), 400

    if not isinstance(quantity, int) or quantity <= 0:
        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Ilość musi być liczbą całkowitą większą od 0'}), 400

    try:
        price = Decimal(str(price))
        if price < 0:
            raise ValueError()
    except (ValueError, TypeError, InvalidOperation):
        return jsonify({'success': False, 'message': 'Nieprawidłowa cena'}), 400

    item_total = price * quantity

    try:
        # Utwórz OrderItem jako custom produkt (bez product_id)
        order_item = OrderItem(
            order_id=order.id,
            product_id=None,  # Brak linku do produktu
            custom_name=custom_name,
            custom_sku=custom_sku,
            is_custom=True,
            quantity=quantity,
            price=price,
            total=item_total,
            picked=False
        )

        db.session.add(order_item)

        # Przelicz sumę zamówienia
        db.session.flush()
        db.session.refresh(order)
        order.recalculate_total()

        db.session.commit()

        # Log aktywności
        log_activity(
            user=current_user,
            action='order_item_added_custom',
            entity_type='order',
            entity_id=order.id,
            new_value={
                'custom_name': custom_name,
                'custom_sku': custom_sku,
                'quantity': quantity,
                'price': float(price),
                'total': float(item_total)
            }
        )

        return jsonify({
            'success': True,
            'message': f'Produkt "{custom_name}" został dodany',
            'item_id': order_item.id,
            'new_total': float(order.total_amount) if order.total_amount else 0
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas dodawania produktu: {str(e)}'
        }), 500


@orders_bp.route('/admin/orders/<int:order_id>/add-gratis-product', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_add_gratis_product(order_id):
    """Dodaje produkt gratisowy (is_bonus=True, price=0) do zamówienia."""
    from modules.products.models import Product

    order = Order.query.get_or_404(order_id)

    try:
        data = request.get_json()
        product_id = data.get('product_id')
        quantity = data.get('quantity', 1)

        if not product_id:
            return jsonify({'success': False, 'message': 'Nie podano produktu'}), 400

        try:
            quantity = int(quantity)
            if quantity < 1:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Ilość musi być liczbą całkowitą większą od 0'}), 400

        product = db.session.get(Product, product_id)
        if not product:
            return jsonify({'success': False, 'message': 'Nie znaleziono produktu'}), 404

        order_item = OrderItem(
            order_id=order.id,
            product_id=product.id,
            quantity=quantity,
            price=0,
            total=0,
            is_bonus=True,
            picked=False
        )

        db.session.add(order_item)
        db.session.flush()
        db.session.refresh(order)
        order.recalculate_total()
        db.session.commit()

        log_activity(
            user=current_user,
            action='order_gratis_added',
            entity_type='order',
            entity_id=order.id,
            new_value={
                'product_id': product.id,
                'product_name': product.name,
                'quantity': quantity,
                'is_bonus': True
            }
        )

        return jsonify({
            'success': True,
            'message': f'Gratis "{product.name}" (x{quantity}) został dodany',
            'item_id': order_item.id,
            'new_total': float(order.total_amount) if order.total_amount else 0
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Błąd podczas dodawania gratisu: {str(e)}'
        }), 500


# ============================================
# PAYMENT METHODS CRUD (Settings Tab)
# ============================================

def _save_payment_method_logo(file, method_id, logo_type):
    """Zapisuje logo metody płatności i zwraca nazwę pliku."""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return None

    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment_methods')
    os.makedirs(upload_dir, exist_ok=True)

    filename = secure_filename(f"method_{method_id}_{logo_type}.{ext}")
    file.save(os.path.join(upload_dir, filename))
    return filename


@orders_bp.route('/admin/orders/payment-methods/create', methods=['POST'])
@login_required
@role_required('admin')
def create_payment_method():
    """Create new payment method."""
    from modules.payments.models import PaymentMethod

    name = request.form.get('name', '').strip()
    is_active = request.form.get('is_active') == 'on'

    if not name:
        return jsonify({'success': False, 'error': 'Nazwa metody jest wymagana'}), 400

    try:
        # Auto-assign sort_order (max + 1)
        max_sort_order = db.session.query(db.func.max(PaymentMethod.sort_order)).scalar() or -1
        sort_order = max_sort_order + 1

        method = PaymentMethod(
            name=name,
            recipient=request.form.get('recipient', '').strip() or None,
            account_number=request.form.get('account_number', '').strip() or None,
            account_number_label=request.form.get('account_number_label', '').strip() or None,
            code=request.form.get('code', '').strip() or None,
            code_label=request.form.get('code_label', '').strip() or None,
            transfer_title=request.form.get('transfer_title', '').strip() or None,
            additional_info=request.form.get('additional_info', '').strip() or None,
            is_active=is_active,
            sort_order=sort_order
        )

        db.session.add(method)
        db.session.flush()  # Potrzebujemy method.id do nazwy pliku

        # Upload logo light
        logo_light_file = request.files.get('logo_light')
        if logo_light_file and logo_light_file.filename:
            filename = _save_payment_method_logo(logo_light_file, method.id, 'light')
            if filename:
                method.logo_light = filename

        # Upload logo dark
        logo_dark_file = request.files.get('logo_dark')
        if logo_dark_file and logo_dark_file.filename:
            filename = _save_payment_method_logo(logo_dark_file, method.id, 'dark')
            if filename:
                method.logo_dark = filename

        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/payment-methods/<int:id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def edit_payment_method(id):
    """Edit payment method."""
    from modules.payments.models import PaymentMethod

    method = PaymentMethod.query.get_or_404(id)

    try:
        method.name = request.form.get('name', '').strip()
        method.recipient = request.form.get('recipient', '').strip() or None
        method.account_number = request.form.get('account_number', '').strip() or None
        method.account_number_label = request.form.get('account_number_label', '').strip() or None
        method.code = request.form.get('code', '').strip() or None
        method.code_label = request.form.get('code_label', '').strip() or None
        method.transfer_title = request.form.get('transfer_title', '').strip() or None
        method.additional_info = request.form.get('additional_info', '').strip() or None
        method.is_active = request.form.get('is_active') == 'on'
        # sort_order is managed by drag & drop, don't change it here

        # Upload logo light
        logo_light_file = request.files.get('logo_light')
        if logo_light_file and logo_light_file.filename:
            # Usuń stare logo jeśli istnieje
            if method.logo_light:
                old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment_methods', method.logo_light)
                if os.path.exists(old_path):
                    os.remove(old_path)
            filename = _save_payment_method_logo(logo_light_file, method.id, 'light')
            if filename:
                method.logo_light = filename

        # Upload logo dark
        logo_dark_file = request.files.get('logo_dark')
        if logo_dark_file and logo_dark_file.filename:
            # Usuń stare logo jeśli istnieje
            if method.logo_dark:
                old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment_methods', method.logo_dark)
                if os.path.exists(old_path):
                    os.remove(old_path)
            filename = _save_payment_method_logo(logo_dark_file, method.id, 'dark')
            if filename:
                method.logo_dark = filename

        # Usunięcie logo (jeśli zaznaczono checkbox "usuń")
        if request.form.get('remove_logo_light') == '1' and method.logo_light:
            old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment_methods', method.logo_light)
            if os.path.exists(old_path):
                os.remove(old_path)
            method.logo_light = None

        if request.form.get('remove_logo_dark') == '1' and method.logo_dark:
            old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment_methods', method.logo_dark)
            if os.path.exists(old_path):
                os.remove(old_path)
            method.logo_dark = None

        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/payment-methods/<int:id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_payment_method(id):
    """Delete payment method."""
    from modules.payments.models import PaymentMethod

    method = PaymentMethod.query.get_or_404(id)

    try:
        db.session.delete(method)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/payment-methods/list', methods=['GET'])
@login_required
@role_required('admin')
def get_payment_methods_list():
    """Get payment methods list HTML (for AJAX refresh)."""
    from modules.payments.models import PaymentMethod
    from flask import render_template_string

    payment_methods = PaymentMethod.query.order_by(PaymentMethod.sort_order, PaymentMethod.name).all()

    template = '''<!-- Data rows -->
{% if payment_methods %}
    {% for method in payment_methods %}
        <div class="payment-method-list-item" data-method-id="{{ method.id }}" draggable="true">
            <div class="payment-method-col-name">
                <div class="drag-handle">
                    <svg width="16" height="16" viewBox="0 0 16 16" fill="currentColor">
                        <path d="M2 3h12v2H2V3zm0 4h12v2H2V7zm0 4h12v2H2v-2z"/>
                    </svg>
                </div>
                <strong>{{ method.name }}</strong>
            </div>
            <div class="payment-method-col-account">
                <code>{{ method.account_number or '—' }}</code>
            </div>
            <div class="payment-method-col-status">
                {% if method.is_active %}
                    <span class="badge badge-success">Aktywny</span>
                {% else %}
                    <span class="badge badge-secondary">Nieaktywny</span>
                {% endif %}
            </div>
            <div class="payment-method-col-actions">
                <button type="button" class="action-link" onclick='openEditPaymentMethodModal({{ method.to_dict()|tojson }})'>Edytuj</button>
                <button type="button" class="action-link delete-link" onclick="deletePaymentMethod({{ method.id }}, &#39;{{ method.name }}&#39;)">Usuń</button>
            </div>
        </div>
    {% endfor %}
{% else %}
    <div class="empty-state">
        <p>Brak metod płatności. Dodaj pierwszą metodę.</p>
    </div>
{% endif %}'''

    return render_template_string(template, payment_methods=payment_methods)


@orders_bp.route('/admin/orders/payment-methods/reorder', methods=['POST'])
@login_required
@role_required('admin')
def reorder_payment_methods():
    """Reorder payment methods via drag & drop."""
    from modules.payments.models import PaymentMethod

    data = request.get_json()
    order = data.get('order', [])

    try:
        for item in order:
            method = db.session.get(PaymentMethod, item['id'])
            if method:
                method.sort_order = item['sort_order']

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>', methods=['GET'])
@login_required
@role_required('admin', 'mod')
def admin_get_shipping_request(shipping_request_id):
    """Get shipping request details as JSON."""
    from modules.orders.models import ShippingRequest

    sr = ShippingRequest.query.get_or_404(shipping_request_id)

    # Build orders list with shipping costs
    orders_data = []
    for ro in sr.request_orders:
        if ro.order:
            orders_data.append({
                'id': ro.order.id,
                'order_number': ro.order.order_number,
                'total_amount': float(ro.order.total_amount or 0),
                'shipping_cost': float(ro.order.shipping_cost or 0)
            })

    return jsonify({
        'id': sr.id,
        'request_number': sr.request_number,
        'status': sr.status,
        'status_display_name': sr.status_display_name,
        'courier': sr.courier,
        'tracking_number': sr.tracking_number,
        'parcel_size': sr.parcel_size,
        'calculated_shipping_cost': float(sr.calculated_shipping_cost or 0),
        'admin_notes': sr.admin_notes,
        'packaging_material_id': sr.packaging_material_id,
        'packaging_material': ({
            'id': sr.packaging_material.id,
            'name': sr.packaging_material.name,
            'type': sr.packaging_material.type,
            'type_display': sr.packaging_material.type_display,
            'size_category': sr.packaging_material.size_category,
            'size_display': sr.packaging_material.size_display,
            'sale_price': float(sr.packaging_material.sale_price) if sr.packaging_material.sale_price else None,
        } if sr.packaging_material else None),
        'client_package_preference': sr.client_package_preference,
        'client_notes': sr.client_notes,
        'address_type': sr.address_type,
        'shipping_name': sr.shipping_name,
        # Nazwa adresata odporna na typ dostawy — przy paczkomacie `shipping_name`
        # jest puste i modal pokazywał na liście miasto punktu zamiast człowieka.
        'addressee_name': sr.addressee_name,
        'shipping_address': sr.shipping_address,
        'shipping_postal_code': sr.shipping_postal_code,
        'shipping_city': sr.shipping_city,
        'shipping_voivodeship': sr.shipping_voivodeship,
        'pickup_courier': sr.pickup_courier,
        'pickup_point_id': sr.pickup_point_id,
        'pickup_address': sr.pickup_address,
        'pickup_postal_code': sr.pickup_postal_code,
        'pickup_city': sr.pickup_city,
        'orders': orders_data,
        # Modal nazywa przycisk destrukcyjny wg tego, co ten przycisk NAPRAWDĘ robi:
        # na paczce zbiorczej DELETE rozwiązuje konsolidację (zamówienia wracają do
        # właścicieli), na zwykłym zleceniu kasuje zlecenie.
        'is_consolidation': sr.is_consolidation,
        'is_consolidated_source': sr.is_consolidated_source,
        # To samo zdanie co na karcie WMS („Czeka na wycenę: Jagoda R."). Admin
        # wypełniający kwoty w modalu widzi listę zamówień bez podziału na ludzi —
        # bez tego nie wie, czyje pola zostawia puste. None na zwykłym zleceniu
        # i na paczce, w której wszyscy są już rozliczeni.
        'consolidation_block_note': sr.consolidation_block_note,
        'payment_deadline': sr.payment_deadline.isoformat() if sr.payment_deadline else None,
        'created_at': sr.created_at.isoformat() if sr.created_at else None
    })


def _sync_order_statuses_from_shipping_request(shipping_request, new_sr_status_slug):
    """
    Synchronizuje statusy zamówień klienta na podstawie zmiany statusu zlecenia wysyłki.
    Mapowanie: SR 'wyslane' → Order 'wyslane', SR 'dostarczone' → Order 'dostarczone'.

    Przejście na 'dostarczone' deleguje do dostarcz_zlecenie() — wcześniej ta funkcja
    ustawiała status zamówień sama i NIE dopisywała przedmiotów do kolekcji klienta,
    więc ręczna zmiana statusu na zleceniu gubiła kolekcję.

    status_juz_ustawiony=True, bo to WŁAŚNIE ta funkcja (a konkretnie jej wywołujący
    w _zapisz_zlecenie_wysylki / admin_bulk_status_shipping_requests) ustawia
    shipping_request.status na 'dostarczone' i commituje ZANIM w ogóle tu trafi —
    dostarcz_zlecenie() nie ma bez tej flagi jak odróżnić „status ustawiony przed
    chwilą przez wywołującego" od „zlecenie historyczne, dostarczone dawno temu bez
    delivered_at" (patrz komentarz przy strażniku w wms_utils.py).
    """
    from utils.email_manager import EmailManager
    from utils.push_manager import PushManager
    from modules.orders.wms_utils import (
        ZlecenieJuzDostarczone, ZlecenieZrodloweNieDomykane, dostarcz_zlecenie)

    if new_sr_status_slug == 'dostarczone':
        try:
            dostarcz_zlecenie(
                shipping_request, source='admin', user=current_user,
                status_juz_ustawiony=True)
        except (ZlecenieJuzDostarczone, ZlecenieZrodloweNieDomykane) as err:
            current_app.logger.info(f'Pominięto domknięcie dostawy: {err}')
        return

    if new_sr_status_slug != 'wyslane':
        return

    order_status_obj = OrderStatus.query.filter_by(slug='wyslane', is_active=True).first()
    if not order_status_obj:
        current_app.logger.warning("Order status 'wyslane' not found or inactive")
        return

    for ro in shipping_request.request_orders:
        order = ro.order
        if not order or order.status == 'wyslane':
            continue

        old_status_name = order.status_display_name
        order.status = 'wyslane'

        try:
            EmailManager.notify_status_change(order, old_status_name, order_status_obj.name)
            PushManager.notify_status_change(order, old_status_name, order_status_obj.name)
        except Exception as e:
            current_app.logger.error(f'Status sync email error for {order.order_number}: {e}')


def _status_logistyczny_dla_zrodla(sr, nowy_status):
    """Czy `nowy_status` to zmiana logistyczna zlecenia jadącego w paczce zbiorczej?

    Jedno miejsce na regułę wspólną dla obu tras zapisu statusu
    (`_zapisz_zlecenie_wysylki` i `admin_bulk_status_shipping_requests`), żeby nie
    rozjechały się jak dotąd — poprawka wprowadzona w jednej z nich obchodziła się
    bokiem drugiej.

    Podział jest ten sam, co w `propaguj_na_zrodla`: logistyka („spakowane",
    „wysłane", „dostarczone") jest własnością kartonu i schodzi na uczestników
    z paczki zbiorczej, więc ustawiona źródłu wprost albo rozjeżdża uczestników
    jednej przesyłki, albo — dla „dostarczone" — zostawia połowiczny stan, bo
    `dostarcz_zlecenie()` odrzuca źródło strażnikiem `ZlecenieZrodloweNieDomykane`
    JUŻ PO zapisie statusu. Statusy finansowe są indywidualne per uczestnik
    i wolno je zapisywać źródłom wprost (robi to `przeprowadz_uczestnikow_na_oplacenie`).
    """
    from modules.orders.consolidation import STATUSY_LOGISTYCZNE

    return sr.is_consolidated_source and nowy_status in STATUSY_LOGISTYCZNE


class BladZapisuZlecenia(Exception):
    """Odmowa zapisu zlecenia wysyłki z powodem nadającym się do pokazania adminowi.

    Zamiast gołego 500 (albo cichego pominięcia pozycji) endpoint przerywa zapis
    i oddaje konkretny komunikat: czego nie zapisano, którego zamówienia dotyczy
    i dlaczego. Właściciel produktu prosił wprost o taki poziom szczegółu.
    """

    def __init__(self, message, status_code=400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def _odmowa_zapisu(sr, powod, order=None, status_code=400):
    """Buduje komunikat odmowy w jednej konwencji dla całego zapisu zlecenia.

    Wzorzec: „Nie zapisano zlecenia WYS/000050 — zamówienie EX/00001046: <powód>".
    Numer zlecenia jest w komunikacie zawsze, bo modal zapisuje N zleceń w pętli
    i bez numeru admin nie wie, które z nich odpadło.
    """
    if order is not None:
        return BladZapisuZlecenia(
            f'Nie zapisano zlecenia {sr.request_number} — '
            f'zamówienie {order.order_number}: {powod}', status_code)
    return BladZapisuZlecenia(
        f'Nie zapisano zlecenia {sr.request_number} — {powod}', status_code)


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>', methods=['PUT'])
@login_required
@role_required('admin', 'mod')
def admin_update_shipping_request(shipping_request_id):
    """Zapis zlecenia wysyłki z modalu wyceny — cienka warstwa nad `_zapisz_zlecenie_wysylki`.

    Sam zapis siedzi w osobnej funkcji tylko po to, żeby dało się opakować go
    w try/except: wcześniej endpoint nie miał żadnego, więc każdy błąd bazy leciał
    jako puste 500 i front nie miał czego pokazać — admin widział „nic się nie stało".
    """
    from modules.orders.models import ShippingRequest

    sr = ShippingRequest.query.get_or_404(shipping_request_id)
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({
            'success': False,
            'error': f'Nie zapisano zlecenia {sr.request_number} — '
                     f'treść żądania nie jest poprawnym JSON-em.',
        }), 400

    try:
        return _zapisz_zlecenie_wysylki(sr, data)
    except BladZapisuZlecenia as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': e.message}), e.status_code
    except HTTPException:
        raise   # abort()/404 z głębi ma zostać tym, czym jest
    except Exception:
        db.session.rollback()
        # Identyfikator błędu wędruje i do loga (obok tracebacka), i do admina —
        # dzięki temu zgłoszenie „nie zapisało się" da się skorelować z konkretnym
        # wpisem w logu zamiast zgadywać. Tracebacka do przeglądarki nie wypuszczamy.
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd zapisu zlecenia {sr.request_number} '
            f'(id={sr.id}, user={getattr(current_user, "email", "?")})')
        return jsonify({
            'success': False,
            'error': f'Nie zapisano zlecenia {sr.request_number} — nieoczekiwany błąd serwera. '
                     f'Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
        }), 500


def _zapisz_zlecenie_wysylki(sr, data):
    """Właściwy zapis zlecenia wysyłki (pola, koszty, statusy, powiadomienia)."""
    old_tracking = sr.tracking_number
    old_status = sr.status

    # Payment deadline
    if 'payment_deadline' in data:
        deadline_str = data['payment_deadline']
        if deadline_str:
            from datetime import datetime
            try:
                sr.payment_deadline = datetime.fromisoformat(deadline_str)
            except (ValueError, TypeError):
                raise _odmowa_zapisu(
                    sr, f'nieczytelny termin płatności („{deadline_str}") — '
                        f'oczekiwano formatu RRRR-MM-DDTGG:MM')
        else:
            sr.payment_deadline = None

    # Update basic fields
    if 'status' in data:
        # Bliźniaczy strażnik do tego z admin_bulk_status_shipping_requests —
        # ta trasa (w odróżnieniu od tamtej) NIE jest martwa, woła ją modal
        # wyceny/wysyłki. Bez niego zapis zostawiał źródłu paczki zbiorczej
        # 'dostarczone' bez delivered_at: strażnik w dostarcz_zlecenie() odrzuca
        # takie zlecenie dopiero w synchronizacji statusów niżej, czyli PO
        # commicie samego statusu.
        #
        # Odmowa, nie ciche pominięcie pola — reszta tej funkcji trzyma tę samą
        # konwencję (patrz _odmowa_zapisu przy kosztach): admin ma wiedzieć,
        # czego nie zapisano i dlaczego, zamiast zobaczyć „zapisano" po zapisie
        # niepełnym.
        #
        # Warunek na zmianę wartości jest istotny: modal wysyła cały payload,
        # więc niezmieniony status w żądaniu nie może blokować zapisu kosztów
        # ani trackingu na zleceniu źródłowym.
        if data['status'] != sr.status and _status_logistyczny_dla_zrodla(sr, data['status']):
            paczka = sr.consolidated_into.request_number if sr.consolidated_into else '?'
            raise _odmowa_zapisu(
                sr, f'jedzie w paczce zbiorczej {paczka} — status logistyczny '
                    f'ustaw na samej paczce, zjedzie na nie propagacją')
        sr.status = data['status']
    if 'courier' in data:
        sr.courier = data['courier'] or None
    if 'tracking_number' in data:
        sr.tracking_number = data['tracking_number'] or None
    if 'parcel_size' in data:
        sr.parcel_size = data['parcel_size'] or None
    if 'admin_notes' in data:
        sr.admin_notes = data['admin_notes'] or None
    if 'packaging_material_id' in data:
        mat_id = data['packaging_material_id'] or None
        sr.packaging_material_id = mat_id
        # Wyprowadź gabaryt z materiału, o ile admin nie podał parcel_size jawnie w tym żądaniu.
        if mat_id and 'parcel_size' not in data:
            from modules.orders.wms_models import PackagingMaterial
            mat = db.session.get(PackagingMaterial, mat_id)
            if mat and mat.size_category:
                sr.parcel_size = mat.size_category
        # Bez zmian quantity_in_stock — magazyn obsługiwany przy pakowaniu zamówienia.

    # Update order shipping costs
    orders_with_new_cost = []
    if 'order_costs' in data:
        pozycje = data['order_costs']
        if not isinstance(pozycje, list):
            raise _odmowa_zapisu(sr, 'pole „order_costs" musi być listą kosztów zamówień')

        for cost_data in pozycje:
            if not isinstance(cost_data, dict):
                raise _odmowa_zapisu(
                    sr, f'pozycja kosztu ma nieoczekiwaną postać ({type(cost_data).__name__}) '
                        f'— oczekiwano obiektu z polami order_id i shipping_cost')

            order_id = cost_data.get('order_id')
            # Wcześniej brak/nieznane order_id znaczyło ciche pominięcie pozycji: admin
            # widział „zapisano", a kwota nie wchodziła do bazy.
            if order_id in (None, ''):
                raise _odmowa_zapisu(sr, 'jedna z pozycji kosztów nie wskazuje zamówienia '
                                         '(brak order_id)')

            order = db.session.get(Order, order_id)
            if not order:
                raise _odmowa_zapisu(sr, f'nie znaleziono zamówienia o identyfikatorze {order_id}')

            # Po konsolidacji to jedyne miejsce, gdzie admin mógłby ustawić kwotę E4
            # zamówieniu obcego klienta — modal renderuje tylko zamówienia tego zlecenia,
            # ale endpoint przyjmował dowolne ID.
            if order.id not in {ro.order_id for ro in sr.request_orders}:
                raise _odmowa_zapisu(sr, 'nie należy do tego zlecenia wysyłki', order=order)

            surowy_koszt = cost_data.get('shipping_cost', 0)
            if surowy_koszt in (None, ''):
                surowy_koszt = 0   # puste pole w modalu = brak kosztu = 0 zł
            try:
                nowy_koszt = Decimal(str(surowy_koszt)).quantize(Decimal('0.01'))
            except (InvalidOperation, ValueError, TypeError):
                raise _odmowa_zapisu(
                    sr, f'kwota wysyłki „{surowy_koszt}" nie jest liczbą', order=order)
            if nowy_koszt < 0:
                raise _odmowa_zapisu(
                    sr, f'kwota wysyłki nie może być ujemna (podano {nowy_koszt} zł)', order=order)

            old_cost = float(order.shipping_cost or 0)
            # Kolumna `orders.shipping_cost` jest NOT NULL, a cały moduł traktuje
            # ZERO jako „brak kosztu" (`(o.shipping_cost or 0) > 0` w consolidation.py,
            # routes.py i email_managerze) — semantyki „NULL = niewycenione" nigdzie nie ma.
            # Wpisanie None wywalało zapis na IntegrityError 1048 i admin dostawał gołe
            # 500: konsolidacja po raz pierwszy stawia w jednym modalu zlecenie wycenione
            # obok niewycenionego, którego pola renderują się puste i wracają jako 0.
            order.shipping_cost = nowy_koszt
            if nowy_koszt > 0 and float(nowy_koszt) != old_cost:
                orders_with_new_cost.append((order, float(nowy_koszt)))

    # Sync total_shipping_cost on SR from order costs
    if 'order_costs' in data:
        sr.total_shipping_cost = sr.calculated_shipping_cost

    # Zmiana statusu/kuriera/trackingu na paczce zbiorczej zjeżdża na zlecenia
    # źródłowe — przed commitem, bo helper nie commituje sam.
    from modules.orders.consolidation import propaguj_na_zrodla
    propaguj_na_zrodla(sr)

    db.session.commit()

    # Auto-status: czeka_na_wycene → czeka_na_oplacenie after pricing
    auto_status_changed = False
    if orders_with_new_cost and sr.is_consolidation:
        # Paczka zbiorcza nie ma własnego statusu finansowego — jej status to
        # minimum ze statusów uczestników. Podniesienie samej paczki zostawiłoby
        # źródła na „czeka na wycenę", a _sprawdz_oplacenie_konsolidacji podnosi
        # uczestnika na „opłacone" tylko z „czeka na opłacenie" — paczka nigdy nie
        # osiągała „opłacone" i WMS odrzucał wysyłkę (UNPAID_SR_STATUSES).
        from modules.orders.consolidation import przeprowadz_uczestnikow_na_oplacenie
        status_paczki_przed = sr.status
        if przeprowadz_uczestnikow_na_oplacenie(sr):
            auto_status_changed = sr.status != status_paczki_przed
            db.session.commit()
    elif sr.status == 'czeka_na_wycene' and orders_with_new_cost:
        has_any_cost = any(
            (ro.order.shipping_cost or 0) > 0
            for ro in sr.request_orders if ro.order
        )
        if has_any_cost:
            old_status = sr.status
            sr.status = 'czeka_na_oplacenie'
            auto_status_changed = True
            # Status finansowy, nie logistyczny — propaguj_na_zrodla świadomie NIE
            # zjedzie z nim w dół (finanse zostają indywidualne), ale wołamy ją tu
            # konsekwentnie z resztą miejsc zapisu, na wypadek zmiany trackingu/kuriera.
            propaguj_na_zrodla(sr)
            db.session.commit()

    # Sync order statuses based on SR status change
    if 'status' in data and data['status'] != old_status and not auto_status_changed:
        _sync_order_statuses_from_shipping_request(sr, data['status'])
        db.session.commit()
    elif auto_status_changed:
        _sync_order_statuses_from_shipping_request(sr, sr.status)
        db.session.commit()

    # Email + push notification for new domestic shipping costs
    if orders_with_new_cost:
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        for order, cost in orders_with_new_cost:
            try:
                EmailManager.notify_cost_added(order, 'domestic_shipping', cost)
                PushManager.notify_cost_added(order, 'domestic_shipping', cost)
            except Exception as e:
                current_app.logger.error(f'Błąd powiadomienia o koszcie wysyłki krajowej: {e}')

    # Activity log
    import json
    log_activity(
        user=current_user,
        action='shipping_request_updated',
        entity_type='shipping_request',
        entity_id=sr.id,
        new_value=json.dumps({
            'request_number': sr.request_number,
            'status': sr.status,
            'tracking_number': sr.tracking_number
        })
    )

    # Auto-create OrderShipment + JEDNO powiadomienie na paczkę, gdy numer właśnie doszedł.
    # Wpisy przesyłki powstają nadal per zamówienie — jedna jest tylko wiadomość
    # do klienta, bo fizycznie dostaje jeden karton.
    tracking_just_added = sr.tracking_number and not old_tracking
    if tracking_just_added:
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        from modules.orders.models import OrderShipment
        for order in sr.orders:
            existing = OrderShipment.query.filter_by(
                order_id=order.id,
                tracking_number=sr.tracking_number
            ).first()
            if not existing:
                shipment = OrderShipment(
                    order_id=order.id,
                    tracking_number=sr.tracking_number,
                    courier=sr.courier,
                    notes=f'Z zlecenia {sr.request_number}',
                    created_by=current_user.id
                )
                db.session.add(shipment)
        db.session.commit()

        courier_name = COURIER_NAMES.get(sr.courier, sr.courier or 'Kurier')
        try:
            EmailManager.notify_shipment_sent(
                sr, tracking_number=sr.tracking_number, courier=sr.courier,
                courier_name=courier_name, tracking_url=sr.tracking_url)
            PushManager.notify_shipment_sent(
                sr, tracking_number=sr.tracking_number, courier_name=courier_name)
        except Exception as e:
            current_app.logger.error(
                f'Błąd powiadomienia o wysyłce zlecenia {sr.request_number}: {e}')

    # Send status change email + push (skip if tracking was just added - that email already covers it)
    status_actually_changed = ('status' in data and data['status'] != old_status) or auto_status_changed
    if status_actually_changed and not tracking_just_added:
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        from modules.orders.models import ShippingRequestStatus
        try:
            EmailManager.notify_shipping_status_change(sr, old_status)
            new_status_obj = ShippingRequestStatus.query.filter_by(slug=sr.status).first()
            new_status_name = new_status_obj.name if new_status_obj else sr.status
            PushManager.notify_shipping_status_change(sr, new_status_name)
        except Exception as e:
            current_app.logger.error(f'Błąd powiadomienia o zmianie statusu zlecenia wysyłki: {e}')

    return jsonify({
        'success': True,
        'message': f'Zlecenie {sr.request_number} zostało zaktualizowane'
    })


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'mod')
def admin_delete_shipping_request(shipping_request_id):
    """Cancel/delete shipping request."""
    from modules.orders.models import ShippingRequest, ShippingRequestOrder

    sr = ShippingRequest.query.get_or_404(shipping_request_id)
    request_number = sr.request_number

    # Zlecenie źródłowe nie ma własnych zamówień i jest tylko widokiem dla klienta —
    # skasowanie go zostawiłoby paczkę z uczestnikiem, którego nie ma.
    if sr.is_consolidated_source:
        return jsonify({
            'success': False,
            'message': f'Zlecenie {sr.request_number} jedzie w paczce zbiorczej '
                       f'{sr.consolidated_into.request_number} — najpierw wypnij je z paczki.',
        }), 409

    # Kasowanie paczki zbiorczej: zamówienia muszą wrócić do właścicieli, inaczej
    # cascade='all, delete-orphan' zabierze powiązania zamówień obcych klientów.
    if sr.is_consolidation:
        from modules.orders.consolidation import (
            rozwiaz_konsolidacje, ConsolidationError, STATUSY_BEZ_EDYCJI)
        # Pre-check tylko po to, żeby dać komunikat pasujący do KASOWANIA — sam
        # rozwiaz_konsolidacje() i tak zablokowałby spakowaną paczkę, ale jego
        # komunikat ("nie można zmieniać jej składu") mówi o edycji składu, nie o
        # usuwaniu, i myliłby admina próbującego skasować spakowaną paczkę
        # (code review rundy 1).
        if sr.status in STATUSY_BEZ_EDYCJI:
            return jsonify({
                'success': False,
                'message': f'Paczka {sr.request_number} jest już spakowana — '
                           f'nie można jej skasować w tym stanie.',
            }), 409
        try:
            zrodla = rozwiaz_konsolidacje(sr)
            db.session.commit()
        except ConsolidationError as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': e.message}), e.status_code
        except Exception:
            db.session.rollback()
            blad_id = zaloguj_blad_z_identyfikatorem(
                f'Nieoczekiwany błąd rozwiązywania paczki {request_number} przy kasowaniu '
                f'(id={shipping_request_id})')
            return jsonify({
                'success': False,
                'message': f'Nie rozwiązano paczki {request_number} — nieoczekiwany błąd '
                           f'serwera. Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
            }), 500
        log_activity(
            user=current_user, action='shipping_request_consolidation_dissolved',
            entity_type='shipping_request',
            new_value={
                'consolidation_number': sr.request_number,
                'restored_numbers': [z.request_number for z in zrodla],
                'reason': 'delete',
            },
        )
        return jsonify({'success': True,
                        'message': 'Paczka zbiorcza rozwiązana, zlecenia wróciły do klientów'})

    # Check if shipping request is linked to an active WMS session
    from modules.orders.wms_models import WmsSessionShippingRequest, WmsSession
    active_wms = WmsSessionShippingRequest.query.join(WmsSession).filter(
        WmsSessionShippingRequest.shipping_request_id == sr.id,
        WmsSession.status.in_(['active', 'paused'])
    ).first()
    if active_wms:
        return jsonify({
            'success': False,
            'message': f'Zlecenie {request_number} jest powiązane z aktywną sesją WMS i nie może zostać usunięte.'
        }), 400

    # Kasowanie dotyka trzech tabel naraz (junction WMS, powiązania zamówień, samo
    # zlecenie) — bez try/except każdy błąd FK leciał do admina jako puste 500.
    try:
        # Remove old WMS junction records (from completed/cancelled sessions)
        WmsSessionShippingRequest.query.filter_by(shipping_request_id=sr.id).delete()

        # Remove all order associations (orders go back to pool)
        ShippingRequestOrder.query.filter_by(shipping_request_id=sr.id).delete()

        # Delete the shipping request
        db.session.delete(sr)
        db.session.commit()
    except Exception:
        db.session.rollback()
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd usuwania zlecenia {request_number} (id={shipping_request_id})')
        return jsonify({
            'success': False,
            'message': f'Nie usunięto zlecenia {request_number} — nieoczekiwany błąd serwera. '
                       f'Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
        }), 500

    # Activity log
    import json
    log_activity(
        user=current_user,
        action='shipping_request_cancelled',
        entity_type='shipping_request',
        entity_id=shipping_request_id,
        new_value=json.dumps({
            'request_number': request_number
        })
    )

    return jsonify({
        'success': True,
        'message': f'Zlecenie {request_number} zostało anulowane'
    })


@orders_bp.route('/admin/orders/shipping-requests/bulk-cancel', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_bulk_cancel_shipping_requests():
    """Bulk cancel/delete multiple shipping requests."""
    data = request.get_json()
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'error': 'Nie wybrano żadnych zleceń'}), 400

    from modules.orders.wms_models import WmsSessionShippingRequest, WmsSession
    from modules.orders.consolidation import rozwiaz_konsolidacje, ConsolidationError

    deleted_numbers = []
    skipped = []  # lista (numer_zlecenia, powod) — do czytelnego komunikatu dla admina

    # Zaznaczenie w UI przeżywa przełączenie filtra „consolidation=sources"
    # (trzymane w sessionStorage), więc admin może w jednym żądaniu zaznaczyć
    # zarówno paczkę zbiorczą, jak i jedno z jej źródeł. Dlatego PIERWSZY PRZEBIEG
    # rozwiązuje WSZYSTKIE zaznaczone paczki zbiorcze, zanim tkniemy cokolwiek
    # innego — inaczej wynik zależałby od kolejności ID w `ids`: źródło
    # przetworzone PRZED swoją paczką miałoby jeszcze consolidated_into_id
    # ustawione i trafiłoby do pominiętych, mimo że po rozwiązaniu paczki (kilka
    # linijek dalej w tej samej pętli) staje się zwykłym, skasowalnym zleceniem —
    # a to właśnie ono, nie coś innego, admin też zaznaczył do usunięcia
    # (code review rundy 1, task 17).
    pozostale = []
    for sr_id in ids:
        sr = db.session.get(ShippingRequest, sr_id)
        if not sr:
            continue
        if sr.is_consolidation:
            try:
                rozwiaz_konsolidacje(sr)
            except ConsolidationError as e:
                skipped.append((sr.request_number, e.message))
            else:
                deleted_numbers.append(sr.request_number)
                # Nic więcej do zrobienia z samymi źródłami tutaj — jeśli któreś
                # z nich jest też w `ids`, drugi przebieg znajdzie je już jako
                # zwykłe, niekonsolidowane zlecenie: rozwiaz_konsolidacje mutuje
                # te same obiekty ShippingRequest w tej samej sesji SQLAlchemy
                # (identity map), więc ich is_consolidated_source jest teraz False.
        else:
            pozostale.append(sr)

    for sr in pozostale:
        # Źródło, którego paczka NIE była (albo była, ale nie dała się rozwiązać)
        # w tym zaznaczeniu — nadal nie ma własnych zamówień, nie kasujemy.
        if sr.is_consolidated_source:
            skipped.append((
                sr.request_number,
                f'jedzie w paczce zbiorczej {sr.consolidated_into.request_number} — '
                f'najpierw wypnij je z paczki',
            ))
            continue
        active_wms = WmsSessionShippingRequest.query.join(WmsSession).filter(
            WmsSessionShippingRequest.shipping_request_id == sr.id,
            WmsSession.status.in_(['active', 'paused'])
        ).first()
        if active_wms:
            skipped.append((sr.request_number, 'powiązane z aktywną sesją WMS'))
            continue
        # Remove old WMS junction records (from completed/cancelled sessions)
        WmsSessionShippingRequest.query.filter_by(shipping_request_id=sr.id).delete()
        deleted_numbers.append(sr.request_number)
        # Remove all order associations (orders go back to pool)
        ShippingRequestOrder.query.filter_by(shipping_request_id=sr.id).delete()
        # Delete the shipping request
        db.session.delete(sr)

    db.session.commit()

    # Activity log
    log_activity(
        user=current_user,
        action='shipping_requests_bulk_cancelled',
        entity_type='shipping_request',
        new_value=json.dumps({
            'request_numbers': deleted_numbers,
            'count': len(deleted_numbers)
        })
    )

    message = f'Usunięto {len(deleted_numbers)} zleceń'
    if skipped:
        # Powód przy każdej pozycji — inaczej admin nie dowie się z interfejsu,
        # dlaczego coś przepadło (code review rundy 1, task 17).
        opisy = '; '.join(f'{numer} ({powod})' for numer, powod in skipped)
        message += f'. Pominięto {len(skipped)} zleceń: {opisy}'

    return jsonify({
        'success': True,
        'message': message,
        'skipped_count': len(skipped)
    })


@orders_bp.route('/admin/orders/shipping-requests/consolidation-preview')
@login_required
@role_required('admin', 'mod')
def admin_consolidation_preview():
    """Dane do modalu konsolidacji — pełne adresy i powody blokady.

    Modal nie może karmić się danymi z kart: karty nie mają kompletu adresów,
    a stan mógł się zmienić od załadowania strony.
    """
    from modules.orders.consolidation import waliduj_do_konsolidacji, ConsolidationError

    surowe = request.args.get('ids', '')
    ids = [int(x) for x in surowe.split(',') if x.strip().isdigit()]
    if not ids:
        return jsonify({'error': 'Nie wskazano zleceń'}), 400

    requests_list = ShippingRequest.query.filter(ShippingRequest.id.in_(ids)).all()

    pozycje = []
    for sr in requests_list:
        pozycje.append({
            'id': sr.id,
            'request_number': sr.request_number,
            # User.full_name jest null-safe i ma fallback na e-mail — ręczne sklejanie
            # first_name/last_name dawało pusty string dla klientów bez podanego imienia.
            'client_name': sr.user.full_name if sr.user else 'Brak klienta',
            'client_email': sr.user.email if sr.user else None,
            'client_phone': sr.user.phone if sr.user else None,
            'full_address': sr.full_address,
            'address_type': sr.address_type,
            'status': sr.status,
            'status_name': sr.status_display_name,
            # Modal liczy z tego „najmniej zaawansowany status" do ostrzeżenia —
            # tą samą miarą (sort_order), którą backend liczy status paczki.
            'status_sort_order': sr.status_rel.sort_order if sr.status_rel else 0,
            'orders_count': len(sr.display_orders),
            'is_consolidation': sr.is_consolidation,
            'has_tracking': bool(sr.tracking_number),
            # Potrzebne modalowi w trybie zarządzania gotową paczką (Task 14).
            'source_ids': [s.id for s in sr.consolidated_sources],
            'lead_source_request_id': sr.lead_source_request_id,
        })

    # Zaznaczenie do modalu (Task 14) może zawierać już istniejącą paczkę zbiorczą —
    # to scenariusz dopięcia, nie łączenia dwóch paczek. Bez target= waliduj_do_konsolidacji
    # nie odróżnia „ten wpis TO cel" od „to inna paczka" i zawsze odrzuca taki zestaw.
    target_w_zestawie = next((sr for sr in requests_list if sr.is_consolidation), None)

    blokady = []
    try:
        waliduj_do_konsolidacji(requests_list, target=target_w_zestawie)
    except ConsolidationError as e:
        blokady.append(e.message)

    return jsonify({'success': True, 'requests': pozycje, 'blocked': blokady})


def _powiadom_o_konsolidacji(zbiorcze):
    """Powiadomienia dla uczestników paczki. Pełna implementacja w utils/email_manager.py."""
    from utils.email_manager import EmailManager
    from utils.push_manager import PushManager
    try:
        EmailManager.notify_shipment_consolidated(zbiorcze)
        PushManager.notify_shipment_consolidated(zbiorcze)
    except Exception as e:
        current_app.logger.error(
            f'Błąd powiadomienia o konsolidacji {zbiorcze.request_number}: {e}')


@orders_bp.route('/admin/orders/shipping-requests/consolidate', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_consolidate_shipping_requests():
    """Tworzy paczkę zbiorczą albo dopina zlecenia do istniejącej."""
    from modules.orders.consolidation import (
        utworz_konsolidacje, dopnij_do_konsolidacji, ConsolidationError)

    data = request.get_json() or {}
    # Payload buduje modal (Task 14) na podstawie zaznaczenia kart — złośliwy albo
    # uszkodzony JSON nie może dać gołego 500 bez treści (nieobsłużony ValueError
    # z int()), tylko czytelny komunikat z kodem 400, zgodnie z resztą modułu
    # (patrz np. export_orders, admin_add_custom_product).
    try:
        ids = [int(x) for x in data.get('ids', [])]
        target_id = int(data['target_id']) if data.get('target_id') else None
        lead_id = int(data['lead_request_id']) if data.get('lead_request_id') else None
    except (ValueError, TypeError):
        return jsonify({'error': 'Nieprawidłowy format identyfikatorów zleceń — oczekiwano liczb całkowitych'}), 400

    try:
        if target_id:
            target = db.session.get(ShippingRequest, target_id)
            if not target:
                return jsonify({'error': 'Nie znaleziono paczki zbiorczej'}), 404
            dopnij_do_konsolidacji(target, [i for i in ids if i != target.id])
            zbiorcze = target
        else:
            if not lead_id:
                return jsonify({'error': 'Wskaż zlecenie wiodące'}), 400
            zbiorcze = utworz_konsolidacje(ids, lead_id)
        db.session.commit()
    except ConsolidationError as e:
        db.session.rollback()
        return jsonify({'error': e.message}), e.status_code
    except Exception:
        db.session.rollback()
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd konsolidacji zleceń {ids} '
            f'(target={target_id}, lead={lead_id})')
        czynnosc = ('dopiąć zleceń do paczki zbiorczej' if target_id
                    else 'utworzyć paczki zbiorczej')
        return jsonify({
            'error': f'Nie udało się {czynnosc} — nieoczekiwany błąd serwera. '
                     f'Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
        }), 500

    _powiadom_o_konsolidacji(zbiorcze)

    log_activity(
        user=current_user, action='shipping_requests_consolidated',
        entity_type='shipping_request', entity_id=zbiorcze.id,
        new_value={
            'consolidation_number': zbiorcze.request_number,
            'source_numbers': [s.request_number for s in zbiorcze.consolidated_sources],
            'lead_request_id': zbiorcze.lead_source_request_id,
        },
    )
    return jsonify({
        'success': True,
        'message': f'Utworzono paczkę zbiorczą {zbiorcze.request_number} '
                   f'z {len(zbiorcze.consolidated_sources)} zleceń',
        'consolidation_id': zbiorcze.id,
    })


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>/consolidation/lead',
                 methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_consolidation_change_lead(shipping_request_id):
    """Przełącza zlecenie wiodące — zmienia adres i adresata paczki."""
    from modules.orders.consolidation import zmien_wiodace, ConsolidationError

    target = db.session.get(ShippingRequest, shipping_request_id)
    if not target:
        return jsonify({'error': 'Nie znaleziono zlecenia'}), 404

    data = request.get_json() or {}
    try:
        lead_request_id = int(data.get('lead_request_id', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Nieprawidłowy identyfikator zlecenia wiodącego — oczekiwano liczby całkowitej'}), 400

    try:
        zmien_wiodace(target, lead_request_id)
        db.session.commit()
    except ConsolidationError as e:
        db.session.rollback()
        return jsonify({'error': e.message}), e.status_code
    except Exception:
        db.session.rollback()
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd zmiany zlecenia wiodącego paczki {target.request_number} '
            f'(lead_request_id={lead_request_id})')
        return jsonify({
            'error': f'Nie zmieniono adresata paczki {target.request_number} — '
                     f'nieoczekiwany błąd serwera. Identyfikator błędu: {blad_id} — '
                     f'podaj go przy zgłoszeniu.',
        }), 500

    log_activity(
        user=current_user, action='shipping_request_consolidation_lead_changed',
        entity_type='shipping_request', entity_id=target.id,
        new_value={'lead_request_id': target.lead_source_request_id},
    )
    # addressee_name, nie shipping_name — przy paczkomacie to drugie jest puste
    # i komunikat brzmiał „jest teraz None”.
    adresat = target.addressee_name
    return jsonify({
        'success': True,
        'message': (f'Adresatem paczki {target.request_number} jest teraz {adresat}'
                    if adresat else
                    f'Zmieniono zlecenie wiodące paczki {target.request_number}'),
    })


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>/consolidation/detach',
                 methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_consolidation_detach(shipping_request_id):
    """Wypina jedno zlecenie z paczki. Przy jednym uczestniku paczka znika."""
    from modules.orders.consolidation import wypnij_zlecenie, ConsolidationError

    target = db.session.get(ShippingRequest, shipping_request_id)
    if not target:
        return jsonify({'error': 'Nie znaleziono zlecenia'}), 404

    data = request.get_json() or {}
    numer = target.request_number
    try:
        source_id = int(data.get('source_id', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Nieprawidłowy identyfikator zlecenia do wypięcia — oczekiwano liczby całkowitej'}), 400

    try:
        rozwiazana = wypnij_zlecenie(target, source_id)
        db.session.commit()
    except ConsolidationError as e:
        db.session.rollback()
        return jsonify({'error': e.message}), e.status_code
    except Exception:
        db.session.rollback()
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd wypięcia zlecenia {source_id} z paczki {numer}')
        return jsonify({
            'error': f'Nie wypięto zlecenia z paczki {numer} — nieoczekiwany błąd serwera. '
                     f'Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
        }), 500

    log_activity(
        user=current_user, action='shipping_request_consolidation_detached',
        entity_type='shipping_request',
        new_value={'consolidation_number': numer, 'source_id': source_id,
                   'dissolved': rozwiazana},
    )
    return jsonify({
        'success': True, 'dissolved': rozwiazana,
        'message': ('Paczka została rozwiązana — został tylko jeden uczestnik'
                    if rozwiazana else 'Zlecenie wypięte z paczki'),
    })


@orders_bp.route('/admin/orders/shipping-requests/<int:shipping_request_id>/consolidation/dissolve',
                 methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_consolidation_dissolve(shipping_request_id):
    """Rozmontowuje paczkę zbiorczą — wszystkie zamówienia wracają do swoich zleceń."""
    from modules.orders.consolidation import rozwiaz_konsolidacje, ConsolidationError

    target = db.session.get(ShippingRequest, shipping_request_id)
    if not target:
        return jsonify({'error': 'Nie znaleziono zlecenia'}), 404

    numer = target.request_number
    try:
        zrodla = rozwiaz_konsolidacje(target)
        numery = [s.request_number for s in zrodla]
        db.session.commit()
    except ConsolidationError as e:
        db.session.rollback()
        return jsonify({'error': e.message}), e.status_code
    except Exception:
        db.session.rollback()
        blad_id = zaloguj_blad_z_identyfikatorem(
            f'Nieoczekiwany błąd rozwiązywania paczki zbiorczej {numer} (id={target.id})')
        return jsonify({
            'error': f'Nie rozwiązano paczki {numer} — nieoczekiwany błąd serwera. '
                     f'Identyfikator błędu: {blad_id} — podaj go przy zgłoszeniu.',
        }), 500

    log_activity(
        user=current_user, action='shipping_request_consolidation_dissolved',
        entity_type='shipping_request',
        new_value={'consolidation_number': numer, 'restored_numbers': numery},
    )
    return jsonify({
        'success': True,
        'message': f'Paczka {numer} rozwiązana — zlecenia wróciły do samodzielnej wysyłki',
    })


@orders_bp.route('/admin/orders/shipping-requests/bulk-status', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_bulk_status_shipping_requests():
    """Zbiorcza zmiana statusu wielu zleceń wysyłki.

    UWAGA: trasa jest MARTWA — po żadnej stronie (`static/js/`, `templates/`) nic
    jej nie woła; jedyne wywołania to testy. Zostaje, bo to spójne z resztą API
    adminowego wejście i nic nie kosztuje, ale przy szacowaniu wagi błędów w tej
    funkcji nie licz „widoczności na produkcji" — dziś dosięgnąć jej można wyłącznie
    ręcznym żądaniem. (Notatka po tym, jak pierwsza fala sprzątania długu opisała
    tutejszy UnboundLocalError jako awarię produkcyjną — nią nie był.)
    """
    data = request.get_json()
    ids = data.get('ids', [])
    new_status = data.get('status')

    if not ids:
        return jsonify({'error': 'Nie wybrano żadnych zleceń'}), 400

    if not new_status:
        return jsonify({'error': 'Nie wybrano nowego statusu'}), 400

    # Verify status exists
    status_obj = ShippingRequestStatus.query.filter_by(slug=new_status, is_active=True).first()
    if not status_obj:
        return jsonify({'error': 'Nieprawidłowy status'}), 400

    from modules.orders.consolidation import propaguj_na_zrodla

    updated_count = 0
    skipped_source_count = 0
    changed_requests = []  # (ShippingRequest, old_status) for email notifications
    for sr_id in ids:
        sr = db.session.get(ShippingRequest, sr_id)
        if not sr:
            continue
        # Pomijamy WYŁĄCZNIE statusy logistyczne, nie każdy status. Logistyka
        # jest własnością kartonu: propaguj_na_zrodla() kopiuje na źródła tylko
        # STATUSY_LOGISTYCZNE, a dostarcz_zlecenie() odrzuca źródło strażnikiem
        # ZlecenieZrodloweNieDomykane — ustawiony tu wprost status zostawiłby
        # połowiczny stan ('dostarczone' bez delivered_at, bez kaskady na
        # zamówienia i bez kolekcji), którego nic już nie podnosi (cron filtruje
        # status=='wyslane'), albo rozjechałby uczestników jednego kartonu.
        # Finanse są odwrotnie — indywidualne per uczestnik: propagacja jest dla
        # nich CELOWO wyłączona, a przeprowadz_uczestnikow_na_oplacenie() wprost
        # zapisuje status finansowy na źródłach. Pomijanie ich (tak robiła
        # pierwsza wersja tej poprawki) czyniło z ustawienia źródłu np.
        # 'oplacone' cichy no-op, a rada „zmień status całej paczki" była
        # nieprawdziwa — paczka takiego statusu na źródła nie zjedzie.
        # Pominięte źródło niczego nie traci: jeśli w tym samym zaznaczeniu jest
        # jego paczka zbiorcza, propaguj_na_zrodla() niżej i tak skopiuje na nie
        # docelowy stan paczki, niezależnie od kolejności ID.
        if _status_logistyczny_dla_zrodla(sr, new_status):
            skipped_source_count += 1
            continue
        old_status = sr.status
        if old_status != new_status:
            changed_requests.append((sr, old_status))
        sr.status = new_status
        updated_count += 1
        # Zmiana zbiorcza może objąć paczkę konsolidacyjną — jej źródła muszą
        # dostać ten sam stan, zanim padnie wspólny commit poniżej.
        propaguj_na_zrodla(sr)

    db.session.commit()

    # Sync order statuses for changed shipping requests
    if changed_requests:
        for sr, _old_sr_status in changed_requests:
            _sync_order_statuses_from_shipping_request(sr, new_status)
        db.session.commit()

    # Send status change emails + push
    if changed_requests:
        from utils.email_manager import EmailManager
        from utils.push_manager import PushManager
        # UWAGA: ShippingRequestStatus jest już zaimportowany na poziomie modułu
        # (patrz góra pliku). Lokalny import tej samej nazwy tutaj czynił ją
        # lokalną dla CAŁEJ funkcji (reguła zasięgu Pythona) — walidacja statusu
        # ~30 linii wyżej odwoływała się więc do zmiennej, która w tym momencie
        # jeszcze nie istniała, i rzucała UnboundLocalError. Padało każde żądanie,
        # które przeszło wcześniejsze bramki: rola (403 z dekoratora), brak `ids`
        # i brak `status` (400) wracały normalnie i do tej linii nie docierały —
        # więc nie „przy KAŻDYM wywołaniu", jak głosił poprzedni komentarz.
        # Nie dodawaj importu z powrotem.
        for sr, old_status in changed_requests:
            try:
                EmailManager.notify_shipping_status_change(sr, old_status)
                new_status_obj = ShippingRequestStatus.query.filter_by(slug=sr.status).first()
                new_status_name = new_status_obj.name if new_status_obj else sr.status
                PushManager.notify_shipping_status_change(sr, new_status_name)
            except Exception as e:
                current_app.logger.error(f'Błąd powiadomienia o zmianie statusu zlecenia {sr.request_number}: {e}')

    # Activity log
    log_activity(
        user=current_user,
        action='shipping_requests_bulk_status_change',
        entity_type='shipping_request',
        new_value=json.dumps({
            'ids': ids,
            'new_status': new_status,
            'count': updated_count
        })
    )

    komunikat = f'Zmieniono status {updated_count} zleceń na "{status_obj.name}"'
    if skipped_source_count:
        # Rada jest prawdziwa tylko dla statusów logistycznych i tylko takie tu
        # pomijamy — status paczki zbiorczej zjeżdża na uczestników propagacją.
        komunikat += (
            f' (pominięto {skipped_source_count} — jadą w paczce zbiorczej; '
            f'status logistyczny ustaw na samej paczce, zjedzie na nie propagacją)'
        )

    return jsonify({
        'success': True,
        'message': komunikat,
        'skipped_source_count': skipped_source_count,
    })


@orders_bp.route('/admin/orders/shipping-requests/export-inpost', methods=['POST'])
@login_required
@role_required('admin', 'mod')
def admin_export_shipping_requests_inpost():
    """Buduje plik CSV do masowego nadania przesyłek w panelu InPost.

    Zwraca treść pliku w JSON (front sam tworzy plik do pobrania), żeby razem
    z nim przekazać ostrzeżenia o zleceniach pominiętych lub niekompletnych.
    """
    from modules.orders.inpost_export import build_inpost_csv, count_exported_rows

    data = request.get_json() or {}
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'error': 'Nie wybrano żadnych zleceń'}), 400

    shipping_requests = ShippingRequest.query.filter(
        ShippingRequest.id.in_(ids),
        # Zlecenie źródłowe ma własny adres i gabaryt, więc trafiłoby do pliku
        # jako druga przesyłka na tę samą paczkę — realny koszt u kuriera.
        ShippingRequest.consolidated_into_id.is_(None),
    ).order_by(ShippingRequest.request_number).all()

    if not shipping_requests:
        return jsonify({'error': 'Nie znaleziono zaznaczonych zleceń'}), 404

    csv_text, warnings = build_inpost_csv(shipping_requests)

    # Zlecenia źródłowe odpadły w zapytaniu wyżej po consolidated_into_id — bez
    # tego admin nie wiedziałby, dlaczego zaznaczył np. 5 pozycji, a plik ma 3
    # wiersze. Etykieta i tak jedzie z paczką zbiorczą, więc to nie błąd, tylko
    # informacja.
    found_ids = {sr.id for sr in shipping_requests}
    excluded_ids = set(ids) - found_ids
    if excluded_ids:
        excluded_sources = ShippingRequest.query.filter(
            ShippingRequest.id.in_(excluded_ids),
            ShippingRequest.consolidated_into_id.isnot(None),
        ).order_by(ShippingRequest.request_number).all()
        for sr in excluded_sources:
            paczka = sr.consolidated_into.request_number if sr.consolidated_into else '?'
            warnings.append(
                f'{sr.request_number} — jedzie w paczce zbiorczej {paczka}, '
                f'pominięto (etykieta jest już w pliku dla tej paczki)'
            )

    exported = count_exported_rows(csv_text)
    from modules.orders.models import get_local_now
    filename = f'inpost_{get_local_now().strftime("%Y-%m-%d_%H%M")}.csv'

    log_activity(
        user=current_user,
        action='shipping_requests_exported_inpost',
        entity_type='shipping_request',
        new_value=json.dumps({
            'ids': ids,
            'exported': exported,
            # Liczone od WSZYSTKICH zaznaczonych ID, nie tylko tych, które przeszły
            # filtr źródeł — inaczej log nie mówi prawdy o realnej liczbie pominiętych.
            'skipped': len(ids) - exported,
        })
    )

    return jsonify({
        'success': True,
        'csv': csv_text,
        'filename': filename,
        'exported': exported,
        'warnings': warnings,
    })


@orders_bp.route('/admin/orders/shipping-request-statuses/list', methods=['GET'])
@login_required
@role_required('admin', 'mod')
def admin_list_shipping_request_statuses():
    """List all active shipping request statuses."""
    from modules.orders.models import ShippingRequestStatus

    statuses = ShippingRequestStatus.query.filter_by(is_active=True).order_by(ShippingRequestStatus.sort_order).all()

    return jsonify([{
        'id': s.id,
        'slug': s.slug,
        'name': s.name,
        'badge_color': s.badge_color,
        'is_initial': s.is_initial
    } for s in statuses])


# ====================
# ADMIN: OPINIE O DOSTAWIE (task 869efhwph)
# ====================

@orders_bp.route('/admin/shipping-requests/opinie')
@login_required
@role_required('admin', 'mod')
def admin_delivery_reviews():
    """Lista opinii klientów o dostawie (task 869efhwph).

    JOIN po ShippingRequest wykorzystujemy też do eager-loadu (contains_eager) —
    inaczej każdy wiersz tabeli w szablonie (numer zlecenia, sposób domknięcia)
    doklejałby własne zapytanie i przy dłuższej liście opinii zrobiłby z tego N+1.

    Paginacja przez wspólny `utils/pagination.py` — ten sam mechanizm, co lista
    produktów, użytkowników i zleceń wysyłki: `page` z query stringu, wybór „ile
    na stronie" zapamiętywany na czas sesji logowania, opcja „Wszystkie".
    Bez paginacji `.all()` ładowało całą tabelę, która rośnie monotonicznie
    i nic jej nie czyści.

    Do szablonu idzie CAŁY obiekt paginacji, nie sama lista pozycji: pierwsza
    wersja tej zmiany oddawała `pagination.items`, więc widok cichcem ucinał się
    na 20 wierszach i admin przy 25 opiniach nie miał skąd wiedzieć, że pozostałe
    (w tym ewentualne oceny 1–2 z reklamacjami) w ogóle istnieją. Szablon
    renderuje makro `pagination_nav` z components/_pagination.html, które pokazuje
    licznik wszystkich wyników nawet wtedy, gdy strona jest jedna.
    """
    from sqlalchemy.orm import contains_eager, joinedload

    from modules.orders.review_models import DeliveryReview
    from utils.pagination import paginate_with_choice, resolve_per_page

    zapytanie = (
        DeliveryReview.query
        .join(ShippingRequest, DeliveryReview.shipping_request_id == ShippingRequest.id)
        .options(contains_eager(DeliveryReview.shipping_request), joinedload(DeliveryReview.user))
    )

    ocena = request.args.get('rating', type=int)
    if ocena in (1, 2, 3, 4, 5):
        zapytanie = zapytanie.filter(DeliveryReview.rating == ocena)

    tylko_z_komentarzem = request.args.get('with_comment') == '1'
    if tylko_z_komentarzem:
        zapytanie = zapytanie.filter(DeliveryReview.comment.isnot(None))

    strona = request.args.get('page', 1, type=int)
    per_page = resolve_per_page('delivery_reviews', default=20)
    paginacja = paginate_with_choice(
        zapytanie.order_by(DeliveryReview.created_at.desc()), strona, per_page)

    return render_template(
        'admin/orders/delivery_reviews.html',
        title='Opinie o dostawie',
        opinie=paginacja.items,
        paginacja=paginacja,
        per_page=per_page,
        wybrana_ocena=ocena,
        tylko_z_komentarzem=tylko_z_komentarzem,
    )


# ====================
# ADMIN SHIPPING REQUESTS LIST
# ====================

@orders_bp.route('/admin/orders/shipping-requests')
@login_required
@role_required('admin', 'mod')
def admin_shipping_requests_list():
    """
    Redirect old shipping requests page to WMS dashboard shipping tab.
    Preserves query parameters (status, search, page).
    """
    args = request.args.to_dict()
    args['tab'] = 'shipping'
    return redirect(url_for('orders.wms_dashboard', **args))


# ============================================
# PAYMENT PROOF FILE SERVING (zabezpieczony)
# ============================================

@orders_bp.route('/payment-proof/<filename>')
@login_required
def serve_payment_proof(filename):
    """
    Serwuje pliki dowodów płatności z autoryzacją.
    - Admin/Mod: dostęp do wszystkich plików
    - Client: dostęp tylko do plików powiązanych z własnymi zamówieniami
    """
    from flask import send_from_directory

    # Znajdź potwierdzenie po nazwie pliku
    confirmation = PaymentConfirmation.query.filter_by(proof_file=filename).first()
    if not confirmation:
        abort(404)

    # Sprawdź uprawnienia
    if current_user.role in ('admin', 'mod'):
        pass  # Admin/mod widzi wszystko
    elif current_user.role == 'client':
        # Klient widzi tylko pliki powiązane z jego zamówieniami
        order = confirmation.order
        if not order or order.user_id != current_user.id:
            abort(403)
    else:
        abort(403)

    upload_folder = os.path.join(current_app.root_path, 'uploads', 'payment_confirmations')
    return send_from_directory(upload_folder, filename)
